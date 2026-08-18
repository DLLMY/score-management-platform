import io
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Any, Optional, Callable

import math
import csv
import urllib.parse

"""
统一Excel服务模块
支持：大数据量分批导出、格式验证、错误定位、中文文件名
"""
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
try:
    OPENPYXL_WRITER_AVAILABLE = True
except ImportError:
    OPENPYXL_WRITER_AVAILABLE = False
BATCH_SIZE = 5000
MAX_EXPORT_ROWS = 500000


class ExcelExportService:
    """统一Excel导出服务 - 支持大数据量分批处理"""

    @staticmethod
    def _sanitize_filename(filename: str) -> str:
        """清理文件名中的非法字符"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, "_")
        if len(filename) > 100:
            filename = filename[:100]
        return filename

    @staticmethod
    def _build_content_disposition(filename: str) -> str:
        """构建支持中文的Content-Disposition头 (RFC 5987)"""
        safe_filename = ExcelExportService._sanitize_filename(filename)
        encoded = urllib.parse.quote(safe_filename, safe="")
        return f"attachment; filename*=UTF-8''{encoded}"

    @staticmethod
    def _convert_value(value: Any) -> Any:
        """统一类型转换，确保值可安全写入Excel"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, bool):
            return "是" if value else "否"
        if isinstance(value, (list, tuple)):
            return ", ".join(str(v) for v in value)
        if isinstance(value, dict):
            return str(value)
        if isinstance(value, float):
            return round(value, 10)
        if isinstance(value, str):
            # S8 修复: Excel 公式注入（= + - @ 开头 → 打开即执行公式/外链）
            stripped = value.lstrip()
            if stripped.startswith(("=", "+", "-", "@")):
                return "'" + value
        return value

    @staticmethod
    def export_large_dataset(
        data_getter: Callable[[int, int], List[Dict[str, Any]]],
        total_count: int,
        headers: List[str],
        filename: str,
        sheet_name: str = "数据",
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> io.BytesIO:
        """
        大数据量分批导出（流式写入）
        注意：write_only 模式下不支持样式设置，以性能优先
        Args:
            data_getter: 数据获取函数 get_data(page, page_size) -> List[Dict]
            total_count: 总数据量
            headers: 表头列表
            filename: 文件名
            sheet_name: 工作表名
            progress_callback: 进度回调 (current, total)
        Returns:
            Excel文件字节流
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装")
        if total_count > MAX_EXPORT_ROWS:
            total_count = MAX_EXPORT_ROWS
        # 先创建带样式的表头文件
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        # 追加数据行
        current_row = 2
        total_batches = math.ceil(total_count / BATCH_SIZE)
        exported = 0
        for batch_idx in range(total_batches):
            page = batch_idx + 1
            batch_data = data_getter(page, BATCH_SIZE)
            if not batch_data:
                break
            for row_data in batch_data:
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    converted = ExcelExportService._convert_value(value)
                    cell = ws.cell(row=current_row, column=col_idx, value=converted)
                    cell.alignment = center_align
                current_row += 1
            exported += len(batch_data)
            if progress_callback:
                progress_callback(exported, total_count)
        # 设置列宽
        for col, header in enumerate(headers, 1):
            max_len = len(str(header)) + 2
            for row_data in data_getter(1, min(100, total_count)):  # 采样前100行计算列宽
                value = row_data.get(header, "")
                converted = ExcelExportService._convert_value(value)
                cell_len = len(str(converted)) + 2
                if cell_len > max_len:
                    max_len = cell_len
            ws.column_dimensions[get_column_letter(col)].width = min(max_len, 50)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        headers: List[str],
        filename: str = None,
        sheet_name: str = "Sheet1",
    ) -> io.BytesIO:
        """通用Excel导出（小数据量，<=5000行）"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装")
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        for row_idx, row_data in enumerate(data, start=2):
            for col, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                converted = ExcelExportService._convert_value(value)
                cell = ws.cell(row=row_idx, column=col, value=converted)
                cell.alignment = center_align
                cell.border = thin_border
        for col, header in enumerate(headers, 1):
            max_len = len(str(header)) + 2
            for row_data in data:
                value = row_data.get(header, "")
                converted = ExcelExportService._convert_value(value)
                cell_len = len(str(converted)) + 2
                if cell_len > max_len:
                    max_len = cell_len
            ws.column_dimensions[get_column_letter(col)].width = min(max_len, 50)
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_to_csv(
        data: List[Dict[str, Any]],
        headers: List[str],
        filename: str = None,
    ) -> io.StringIO:
        """CSV导出"""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row_data in data:
            writer.writerow([row_data.get(h, "") for h in headers])
        output.seek(0)
        return output


class ExcelImportService:
    """统一Excel导入服务 - 含格式验证与错误定位"""

    @staticmethod
    def _parse_cell_value(value: Any) -> Any:
        """将Excel单元格值转换为Python原生类型"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            return value
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    @staticmethod
    def parse_excel_file(file_content: bytes) -> Dict[str, Any]:
        """
        解析Excel文件并返回结构化数据和验证结果
        Returns:
            {
                'headers': List[str],
                'data': List[Dict[str, Any]],
                'total_rows': int,
                'errors': List[Dict]  # 格式错误信息
            }
        """
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "openpyxl库未安装"}
        try:
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if not rows:
                return {"success": False, "error": "文件为空", "total_rows": 0, "data": [], "headers": []}
            headers = [str(h).strip() if h else f"列{i+1}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]
            parsed_data = []
            errors = []
            empty_rows = 0
            for row_idx, row in enumerate(data_rows, start=2):
                if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                    empty_rows += 1
                    continue
                row_dict = {}
                for col_idx, value in enumerate(row):
                    header = headers[col_idx] if col_idx < len(headers) else f"列{col_idx+1}"
                    row_dict[header] = ExcelImportService._parse_cell_value(value)
                parsed_data.append(row_dict)
            return {
                "success": True,
                "headers": headers,
                "data": parsed_data,
                "total_rows": len(parsed_data),
                "empty_rows": empty_rows,
                "errors": errors,
            }
        except Exception as e:
            return {"success": False, "error": f"文件解析失败: {str(e)}", "total_rows": 0, "data": [], "headers": []}

    @staticmethod
    def validate_required_fields(
        data: List[Dict[str, Any]],
        required_fields: List[str],
        field_labels: Dict[str, str] = None,
    ) -> Dict[str, Any]:
        """
        验证必填字段
        Args:
            data: 解析后的数据列表
            required_fields: 必填字段列表
            field_labels: 字段显示名映射
        Returns:
            {
                'valid': bool,
                'errors': [{row, field, message}],
                'valid_count': int,
                'invalid_count': int
            }
        """
        if field_labels is None:
            field_labels = {}
        errors = []
        valid_data = []
        invalid_count = 0
        for idx, row in enumerate(data, start=1):
            row_errors = []
            for field in required_fields:
                value = row.get(field, "")
                if value is None or str(value).strip() == "":
                    label = field_labels.get(field, field)
                    row_errors.append({"row": idx, "field": field, "message": f"{label}不能为空", "value": value})
            if row_errors:
                errors.extend(row_errors)
                invalid_count += 1
            else:
                valid_data.append(row)
        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "valid_count": len(valid_data),
            "invalid_count": invalid_count,
            "total_count": len(data),
            "valid_data": valid_data,
        }

    @staticmethod
    def generate_error_excel(errors: List[Dict], filename: str = "导入错误数据.xlsx") -> io.BytesIO:
        """生成包含错误详情的Excel文件"""
        if not OPENPYXL_AVAILABLE:
            return io.BytesIO()
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "错误详情"
        headers = ["行号", "字段", "错误信息", "原始值"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, error in enumerate(errors, start=2):
            ws.cell(row=row_idx, column=1, value=error.get("row", ""))
            ws.cell(row=row_idx, column=2, value=error.get("field", ""))
            ws.cell(row=row_idx, column=3, value=error.get("message", ""))
            ws.cell(row=row_idx, column=4, value=str(error.get("value", ""))[:50])
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
        wb.save(output)
        output.seek(0)
        return output


excel_export_service = ExcelExportService()
excel_import_service = ExcelImportService()
