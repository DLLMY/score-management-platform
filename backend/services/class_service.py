from datetime import datetime
from models import db, ClassInfo, User, Admin, AdminClass, get_by_id, cascade_delete_related_records
from utils.permission import get_allowed_classes
from utils.db_session import db_session_scope
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json
import re
import io


class ClassService:

    def __init__(self):
        pass

    def _can_access_class(self, class_name, admin=None):
        if admin is None:
            return False
        allowed_classes = get_allowed_classes(admin.id)
        if allowed_classes is None:
            return True
        return class_name in allowed_classes

    def _get_head_teacher_name(self, head_teacher_id):
        if not head_teacher_id:
            return None
        admin = get_by_id(Admin, head_teacher_id)
        return admin.real_name if admin else None

    def _get_student_count(self, class_info):
        return User.query.filter(
            User.class_name == class_info.name,
            User.is_active,
        ).count()

    def _build_class_response(self, class_info):
        # B3 收敛：基础列走 ClassInfo.to_dict，计算字段叠加（2026-08-30）
        data = class_info.to_dict()
        data["head_teacher_name"] = self._get_head_teacher_name(class_info.head_teacher_id)
        data["student_count"] = self._get_student_count(class_info)
        return data

    def get_class_list(self, page=1, per_page=10, keyword=None, admin=None):
        try:
            allowed_classes = get_allowed_classes(admin.id) if admin else None

            query = ClassInfo.query
            if allowed_classes is not None:
                query = query.filter(ClassInfo.name.in_(allowed_classes))

            if keyword:
                query = query.filter(
                    db.or_(
                        ClassInfo.name.like(f"%{keyword}%"),
                        ClassInfo.grade.like(f"%{keyword}%"),
                        ClassInfo.description.like(f"%{keyword}%"),
                    )
                )

            total = query.count()

            classes = query.order_by(ClassInfo.name).paginate(
                page=page, per_page=per_page, error_out=False
            )

            return {
                "classes": [self._build_class_response(c) for c in classes.items],
                "pagination": {
                    "page": page,
                    "per_page": per_page,
                    "total": total,
                    "pages": (total + per_page - 1) // per_page,
                },
            }
        except Exception:
            raise

    def create_class(self, data):
        if ClassInfo.query.filter_by(name=data.get("name")).first():
            return {"success": False, "message": "班级名称已存在"}, 400

        class_info = ClassInfo(
            name=data.get("name"),
            grade=data.get("grade"),
            description=data.get("description"),
            head_teacher_id=data.get("head_teacher_id"),
            is_active=data.get("is_active", True),
        )

        with db_session_scope(detach=False):
            db.session.add(class_info)
            db.session.flush()

            if class_info.head_teacher_id:
                admin_link = AdminClass(
                    admin_id=class_info.head_teacher_id,
                    class_info_id=class_info.id,
                    is_primary=True,
                    assigned_at=datetime.now(),
                )
                db.session.add(admin_link)

            response = self._build_class_response(class_info)

        return response, 201

    def get_class(self, class_id, admin=None):
        class_info = ClassInfo.query.get_or_404(class_id)
        if not self._can_access_class(class_info.name, admin=admin):
            return {"success": False, "message": "无权查看该班级"}, 403
        return self._build_class_response(class_info)

    def update_class(self, class_id, data, admin=None):
        class_info = ClassInfo.query.get_or_404(class_id)
        if not self._can_access_class(class_info.name, admin=admin):
            return {"success": False, "message": "无权修改该班级"}, 403

        old_name = class_info.name
        new_name = data.get("name", class_info.name)

        if new_name != old_name and not self._can_access_class(new_name, admin=admin):
            return {
                "success": False,
                "message": "无权将班级重命名为该名称",
            }, 403

        if new_name != old_name and ClassInfo.query.filter_by(name=new_name).first():
            return {"success": False, "message": "班级名称已存在"}, 400

        with db_session_scope(detach=False):
            class_info.name = new_name
            class_info.grade = data.get("grade", class_info.grade)
            class_info.description = data.get("description", class_info.description)
            class_info.is_active = data.get("is_active", class_info.is_active)
            class_info.updated_at = datetime.now()

            new_head_teacher_id = data.get("head_teacher_id", class_info.head_teacher_id)
            if new_head_teacher_id != class_info.head_teacher_id:
                if class_info.head_teacher_id:
                    prev_link = AdminClass.query.filter_by(
                        admin_id=class_info.head_teacher_id,
                        class_info_id=class_info.id,
                    ).first()
                    if prev_link:
                        prev_link.is_primary = False

                if new_head_teacher_id:
                    admin_link = AdminClass.query.filter_by(
                        admin_id=new_head_teacher_id,
                        class_info_id=class_info.id,
                    ).first()
                    if admin_link:
                        admin_link.is_primary = True
                    else:
                        admin_link = AdminClass(
                            admin_id=new_head_teacher_id,
                            class_info_id=class_info.id,
                            is_primary=True,
                            assigned_at=datetime.now(),
                        )
                        db.session.add(admin_link)

                class_info.head_teacher_id = new_head_teacher_id

            if old_name != class_info.name:
                User.query.filter_by(class_name=old_name).update({"class_name": class_info.name})

            # Build response inside session to avoid DetachedInstanceError
            response = self._build_class_response(class_info)

        return response

    def delete_class(self, class_id, admin=None):
        class_info = ClassInfo.query.get_or_404(class_id)
        if not self._can_access_class(class_info.name, admin=admin):
            return {"success": False, "message": "无权删除该班级"}, 403

        with db_session_scope(detach=False):
            AdminClass.query.filter_by(class_info_id=class_id).delete()
            # 班级存在多级子表依赖（如 class_info -> seating_chart -> seating_seat、
            # class_info -> duty_group -> duty_assignment），需递归清理；
            # user.class_info_id 等可空外键只解除引用，不会删除学生
            cascade_delete_related_records(ClassInfo, class_id)
            db.session.delete(class_info)

        return {"success": True, "message": "班级删除成功"}

    def validate_associations(self):
        issues = []
        all_classes = ClassInfo.query.all()

        for cls in all_classes:
            head_teacher_id = cls.head_teacher_id

            if head_teacher_id:
                admin_link = AdminClass.query.filter_by(
                    admin_id=head_teacher_id, class_info_id=cls.id
                ).first()

                if not admin_link:
                    issues.append(
                        {
                            "type": "missing_link",
                            "class_name": cls.name,
                            "class_id": cls.id,
                            "head_teacher_id": head_teacher_id,
                            "message": (
                                f'班级 "{cls.name}" 的班主任ID({head_teacher_id})'
                                "在AdminClass表中缺少关联记录"
                            ),
                        }
                    )
                elif not admin_link.is_primary:
                    issues.append(
                        {
                            "type": "not_primary",
                            "class_name": cls.name,
                            "class_id": cls.id,
                            "head_teacher_id": head_teacher_id,
                            "message": (f'班级 "{cls.name}" 的班主任关联is_primary未设置为True'),
                        }
                    )

            primary_link = AdminClass.query.filter_by(class_info_id=cls.id, is_primary=True).first()

            if primary_link and cls.head_teacher_id != primary_link.admin_id:
                issues.append(
                    {
                        "type": "mismatch",
                        "class_name": cls.name,
                        "class_id": cls.id,
                        "head_teacher_id": cls.head_teacher_id,
                        "primary_admin_id": primary_link.admin_id,
                        "message": (
                            f'班级 "{cls.name}" 的head_teacher_id('
                            f"{cls.head_teacher_id})与AdminClass主班关联("
                            f"{primary_link.admin_id})不一致"
                        ),
                    }
                )

        return {
            "success": True,
            "total_classes": len(all_classes),
            "issues_found": len(issues),
            "issues": issues,
        }

    def fix_associations(self):
        fixed_count = 0
        issues_fixed = []
        all_classes = ClassInfo.query.all()

        with db_session_scope(detach=False):
            for cls in all_classes:
                head_teacher_id = cls.head_teacher_id

                if head_teacher_id:
                    admin_link = AdminClass.query.filter_by(
                        admin_id=head_teacher_id, class_info_id=cls.id
                    ).first()

                    if not admin_link:
                        new_link = AdminClass(
                            admin_id=head_teacher_id,
                            class_info_id=cls.id,
                            is_primary=True,
                            assigned_at=datetime.now(),
                        )
                        db.session.add(new_link)
                        fixed_count += 1
                        issues_fixed.append(
                            {
                                "class_name": cls.name,
                                "action": "created_link",
                                "message": (f'为班级 "{cls.name}" 创建了AdminClass关联记录'),
                            }
                        )
                    elif not admin_link.is_primary:
                        admin_link.is_primary = True
                        fixed_count += 1
                        issues_fixed.append(
                            {
                                "class_name": cls.name,
                                "action": "set_primary",
                                "message": (
                                    f'将班级 "{cls.name}" 的AdminClass关联is_primary设置为True'
                                ),
                            }
                        )

                primary_link = AdminClass.query.filter_by(
                    class_info_id=cls.id, is_primary=True
                ).first()

                if primary_link and cls.head_teacher_id != primary_link.admin_id:
                    cls.head_teacher_id = primary_link.admin_id
                    fixed_count += 1
                    issues_fixed.append(
                        {
                            "class_name": cls.name,
                            "action": "sync_head_teacher",
                            "message": (
                                f'同步班级 "{cls.name}" 的head_teacher_id为AdminClass主班管理员'
                            ),
                        }
                    )

        return {
            "success": True,
            "fixed_count": fixed_count,
            "issues_fixed": issues_fixed,
        }

    def export_classes(self, keyword=None, export_format="json"):
        query = ClassInfo.query
        if keyword:
            query = query.filter(
                db.or_(
                    ClassInfo.name.like(f"%{keyword}%"),
                    ClassInfo.grade.like(f"%{keyword}%"),
                )
            )

        classes = query.order_by(ClassInfo.name).all()
        export_data = []

        for cls in classes:
            head_teacher = self._get_head_teacher_name(cls.head_teacher_id)
            student_count = self._get_student_count(cls)

            export_data.append(
                {
                    "name": cls.name,
                    "grade": cls.grade,
                    "description": cls.description,
                    "head_teacher_id": cls.head_teacher_id,
                    "head_teacher_name": head_teacher,
                    "student_count": student_count,
                    "is_active": "是" if cls.is_active else "否",
                    "created_at": (cls.created_at.isoformat() if cls.created_at else None),
                    "updated_at": (cls.updated_at.isoformat() if cls.updated_at else None),
                }
            )

        if export_format == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "班级数据"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            headers = [
                "班级名称",
                "年级",
                "描述",
                "班主任ID",
                "班主任姓名",
                "学生数量",
                "是否启用",
                "创建时间",
                "更新时间",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, item in enumerate(export_data, 2):
                ws.cell(row=row_idx, column=1, value=item["name"]).border = thin_border
                ws.cell(row=row_idx, column=2, value=item["grade"]).border = thin_border
                ws.cell(row=row_idx, column=3, value=item["description"]).border = thin_border
                ws.cell(row=row_idx, column=4, value=item["head_teacher_id"]).border = thin_border
                ws.cell(row=row_idx, column=5, value=item["head_teacher_name"]).border = thin_border
                ws.cell(row=row_idx, column=6, value=item["student_count"]).border = thin_border
                ws.cell(row=row_idx, column=7, value=item["is_active"]).border = thin_border
                ws.cell(row=row_idx, column=8, value=item["created_at"]).border = thin_border
                ws.cell(row=row_idx, column=9, value=item["updated_at"]).border = thin_border

            column_widths = [15, 10, 30, 12, 12, 10, 10, 20, 20]
            for i, width in enumerate(column_widths, 1):
                col_letter = chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"
                ws.column_dimensions[col_letter].width = width

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            filename = f'classes_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}' ".xlsx"
            return {
                "type": "excel",
                "data": buf,
                "filename": filename,
            }
        else:
            output = {
                "export_time": datetime.now().isoformat(),
                "total": len(export_data),
                "data": export_data,
            }

            json_str = json.dumps(output, ensure_ascii=False, indent=2)
            buf = io.BytesIO(json_str.encode("utf-8"))
            buf.seek(0)

            filename = f'classes_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}' ".json"
            return {
                "type": "json",
                "data": buf,
                "filename": filename,
            }

    def import_classes(self, import_list, config=None):
        default_mappings = [
            {
                "source_field": "班级名称",
                "target_field": "name",
                "field_type": "string",
                "required": True,
            },
            {
                "source_field": "年级",
                "target_field": "grade",
                "field_type": "string",
            },
            {
                "source_field": "描述",
                "target_field": "description",
                "field_type": "string",
            },
            {
                "source_field": "班主任ID",
                "target_field": "head_teacher_id",
                "field_type": "integer",
                "relation": "admin",
            },
            {
                "source_field": "班主任姓名",
                "target_field": "head_teacher_name",
                "field_type": "string",
                "relation": "admin",
            },
            {
                "source_field": "是否启用",
                "target_field": "is_active",
                "field_type": "boolean",
            },
        ]

        field_mappings = config.field_mappings if config else default_mappings
        validation_rules = config.validation_rules if config else []
        conflict_strategy = config.conflict_strategy if config else "update"

        success_count = 0
        failed_count = 0
        messages = []

        def validate_item(item):
            errors = []
            for rule in validation_rules:
                field = rule["field"]
                rule_type = rule["rule_type"]
                params = rule.get("params", {})
                message = rule.get("message", f"{field}验证失败")
                value = item.get(field)

                if rule_type == "required" and value is None:
                    errors.append(message)
                elif (
                    rule_type == "max_length" and value and len(str(value)) > params.get("max", 100)
                ):
                    errors.append(message)
                elif rule_type == "min_length" and value and len(str(value)) < params.get("min", 1):
                    errors.append(message)
                elif (
                    rule_type == "regex"
                    and value
                    and not re.match(params.get("pattern", ""), str(value))
                ):
                    errors.append(message)
            return errors

        def resolve_relations(item):
            resolved = item.copy()
            validation_errors = []

            for mapping in field_mappings:
                relation = mapping.get("relation")

                if relation == "admin":
                    head_teacher_name = item.get("head_teacher_name")
                    head_teacher_id = item.get("head_teacher_id")

                    if head_teacher_name and head_teacher_id:
                        validation_errors.append("不能同时提供班主任姓名和班主任ID")
                    elif head_teacher_name:
                        if (
                            not isinstance(head_teacher_name, str)
                            or len(head_teacher_name.strip()) == 0
                        ):
                            validation_errors.append("班主任姓名格式无效，必须为非空字符串")
                        elif len(head_teacher_name.strip()) > 50:
                            validation_errors.append("班主任姓名长度超过限制（最大50字符）")
                        else:
                            admin = Admin.query.filter(
                                Admin.real_name == head_teacher_name.strip()
                            ).first()
                            if not admin:
                                admin = Admin.query.filter(
                                    Admin.username == head_teacher_name.strip()
                                ).first()
                            if not admin:
                                validation_errors.append(
                                    f'班主任 "{head_teacher_name}" 在系统中不存在'
                                )
                            else:
                                if admin.role not in ["admin", "teacher"]:
                                    validation_errors.append(
                                        f'用户 "{head_teacher_name}" 的角色不是'
                                        "管理员或教师，无法担任班主任"
                                    )
                                resolved["head_teacher_id"] = admin.id
                    elif head_teacher_id:
                        if not isinstance(head_teacher_id, (int, str)):
                            validation_errors.append("班主任ID格式无效，必须为数字")
                        else:
                            try:
                                admin_id = int(head_teacher_id)
                                admin = get_by_id(Admin, admin_id)
                                if not admin:
                                    validation_errors.append(
                                        f'班主任ID "{head_teacher_id}" 在系统中不存在'
                                    )
                                else:
                                    if admin.role not in ["admin", "teacher"]:
                                        validation_errors.append(
                                            f'用户ID "{head_teacher_id}" 的角色不是'
                                            "管理员或教师，无法担任班主任"
                                        )
                                    resolved["head_teacher_id"] = admin.id
                            except ValueError:
                                validation_errors.append("班主任ID格式无效，必须为有效数字")

            resolved["_validation_errors"] = validation_errors
            return resolved

        with db_session_scope(detach=False):
            for item in import_list:
                try:
                    errors = validate_item(item)
                    if errors:
                        failed_count += 1
                        messages.append(
                            {
                                "name": item.get("name", "未知"),
                                "action": "failed",
                                "message": (f'验证失败: {", ".join(errors)}'),
                                "row_data": item,
                                "error_fields": list(
                                    set(
                                        [
                                            rule["field"]
                                            for rule in validation_rules
                                            if item.get(rule["field"]) is None
                                        ]
                                    )
                                ),
                            }
                        )
                        continue

                    resolved_item = resolve_relations(item)

                    relation_errors = resolved_item.get("_validation_errors", [])
                    if relation_errors:
                        failed_count += 1
                        messages.append(
                            {
                                "name": item.get("name", "未知"),
                                "action": "failed",
                                "message": (f'关联验证失败: {", ".join(relation_errors)}'),
                                "row_data": item,
                                "error_fields": (
                                    ["head_teacher_name", "head_teacher_id"]
                                    if "head_teacher_name" in item or "head_teacher_id" in item
                                    else []
                                ),
                            }
                        )
                        continue

                    existing = ClassInfo.query.filter_by(name=resolved_item["name"]).first()

                    if existing:
                        if conflict_strategy == "skip":
                            messages.append(
                                {
                                    "name": resolved_item["name"],
                                    "action": "skipped",
                                    "message": (f'班级 "{resolved_item["name"]}" 已存在，已跳过'),
                                }
                            )
                            continue
                        elif conflict_strategy == "update":
                            existing.grade = resolved_item.get("grade", existing.grade)
                            existing.description = resolved_item.get(
                                "description", existing.description
                            )
                            existing.is_active = resolved_item.get("is_active", existing.is_active)
                            existing.updated_at = datetime.now()

                            if "head_teacher_id" in resolved_item:
                                existing.head_teacher_id = resolved_item["head_teacher_id"]
                                if resolved_item["head_teacher_id"]:
                                    admin_link = AdminClass.query.filter_by(
                                        admin_id=resolved_item["head_teacher_id"],
                                        class_info_id=existing.id,
                                    ).first()
                                    if admin_link:
                                        admin_link.is_primary = True
                                    else:
                                        admin_link = AdminClass(
                                            admin_id=resolved_item["head_teacher_id"],
                                            class_info_id=existing.id,
                                            is_primary=True,
                                            assigned_at=datetime.now(),
                                        )
                                        db.session.add(admin_link)

                            messages.append(
                                {
                                    "name": resolved_item["name"],
                                    "action": "updated",
                                    "message": (f'班级 "{resolved_item["name"]}" 已更新'),
                                }
                            )
                    else:
                        new_class = ClassInfo(
                            name=resolved_item["name"],
                            grade=resolved_item.get("grade"),
                            description=resolved_item.get("description"),
                            head_teacher_id=resolved_item.get("head_teacher_id"),
                            is_active=resolved_item.get("is_active", True),
                        )
                        db.session.add(new_class)
                        db.session.flush()

                        if resolved_item.get("head_teacher_id"):
                            admin_link = AdminClass(
                                admin_id=resolved_item["head_teacher_id"],
                                class_info_id=new_class.id,
                                is_primary=True,
                                assigned_at=datetime.now(),
                            )
                            db.session.add(admin_link)

                        messages.append(
                            {
                                "name": resolved_item["name"],
                                "action": "created",
                                "message": f'班级 "{resolved_item["name"]}" 已创建',
                            }
                        )

                    success_count += 1
                except Exception as e:
                    failed_count += 1
                    messages.append(
                        {
                            "name": item.get("name", "未知"),
                            "action": "failed",
                            "message": f"导入失败: {str(e)}",
                        }
                    )

        return {
            "success": True,
            "total": len(import_list),
            "success_count": success_count,
            "failed_count": failed_count,
            "messages": messages,
        }


class_service = ClassService()
