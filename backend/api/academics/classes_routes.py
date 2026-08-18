from flask_restx import Namespace, Resource, fields
from flask import send_file, request
from services.class_service import class_service
from utils.permission import requires_permission, get_current_admin
from utils.response import APIResponse
from utils.logger import log_operation

from models import User
from models import ClassInfo
from models import get_by_id

ns_classes = Namespace("classes", description="班级管理相关操作")

ns_classes.parser = ns_classes.parser()
ns_classes.parser.add_argument("page", type=int, location="args", default=1, help="页码")
ns_classes.parser.add_argument("per_page", type=int, location="args", default=10, help="每页数量")
ns_classes.parser.add_argument("keyword", type=str, location="args", help="搜索关键词")


def _audit_class_op(op, name, result, payload, target_id=None):
    """班级 CRUD 审计（成功才记录，异常全吞不打断业务）。"""
    try:
        if isinstance(result, dict) and result.get("success") is False:
            return
        if target_id is None and isinstance(result, dict) and isinstance(result.get("data"), dict):
            target_id = result["data"].get("id")
        log_operation(
            op,
            "class",
            target_id,
            f"{op} 班级: {name}",
            after_data=payload,
        )
    except Exception:  # noqa: BLE001 - 审计失败不影响业务
        pass


class_model = ns_classes.model(
    "ClassInfo",
    {
        "id": fields.Integer(readOnly=True, description="班级ID"),
        "name": fields.String(required=True, description="班级名称"),
        "grade": fields.String(description="年级"),
        "description": fields.String(description="描述"),
        "head_teacher_id": fields.Integer(description="班主任ID"),
        "is_active": fields.Boolean(description="是否启用"),
    },
)

class_response = ns_classes.model(
    "ClassResponse",
    {
        "id": fields.Integer(description="班级ID"),
        "name": fields.String(description="班级名称"),
        "grade": fields.String(description="年级"),
        "description": fields.String(description="描述"),
        "head_teacher_id": fields.Integer(description="班主任ID"),
        "head_teacher_name": fields.String(description="班主任姓名"),
        "student_count": fields.Integer(description="学生数量"),
        "is_active": fields.Boolean(description="是否启用"),
        "created_at": fields.String(description="创建时间"),
        "updated_at": fields.String(description="更新时间"),
    },
)


@ns_classes.route("/")
class ClassList(Resource):

    @ns_classes.doc("list_classes")
    @ns_classes.response(200, "成功")
    @requires_permission("class.view")
    def get(self):
        args = ns_classes.parser.parse_args()
        page = args.get("page", 1)
        per_page = args.get("per_page", 10)
        keyword = args.get("keyword")
        admin = get_current_admin()
        result = class_service.get_class_list(page, per_page, keyword, admin=admin)  # noqa: F841
        return APIResponse.success(data=result)

    @ns_classes.doc("create_class")
    @ns_classes.expect(class_model)
    @ns_classes.response(201, "创建成功", class_response)
    @requires_permission("class.manage")
    def post(self):
        data = ns_classes.payload
        result = class_service.create_class(data)
        _audit_class_op("class.create", data.get("name"), result, data)
        return result


@ns_classes.route("/<int:id>")
@ns_classes.param("id", "班级ID")
class ClassResource(Resource):

    @ns_classes.doc("get_class")
    @ns_classes.response(200, "成功", class_response)
    @ns_classes.response(404, "班级不存在")
    @requires_permission("class.view")
    def get(self, id):
        admin = get_current_admin()
        return class_service.get_class(id, admin=admin)

    @ns_classes.doc("update_class")
    @ns_classes.expect(class_model)
    @ns_classes.response(200, "更新成功")
    @ns_classes.response(404, "班级不存在")
    @requires_permission("class.manage")
    def put(self, id):
        data = ns_classes.payload
        admin = get_current_admin()
        result = class_service.update_class(id, data, admin=admin)
        _audit_class_op("class.update", data.get("name"), result, data, target_id=id)
        return result

    @ns_classes.doc("delete_class")
    @requires_permission("class.manage")
    def delete(self, id):
        admin = get_current_admin()
        result = class_service.delete_class(id, admin=admin)
        _audit_class_op("class.delete", f"id={id}", result, None, target_id=id)
        return result


@ns_classes.route("/<class_param>/students")
class ClassStudents(Resource):

    @ns_classes.doc("get_class_students", description="获取班级学生列表")
    @requires_permission("class.view")
    def get(self, class_param):
        """
        获取班级学生列表

        支持通过班级ID或班级名称查询
        """
        try:
            # 尝试解析为整数ID，如果失败则作为班级名称处理
            try:
                class_id = int(class_param)
                class_info = get_by_id(ClassInfo, class_id)
                class_name = class_info.name if class_info else class_param
            except ValueError:
                class_name = class_param

            students = User.query.filter_by(class_name=class_name).all()

            student_list = []
            for student in students:
                student_list.append(
                    {
                        "id": student.id,
                        "name": student.name,
                        "card_id": student.card_id,
                        "current_score": student.current_score,
                        "class_name": student.class_name,
                        "phone": student.phone,
                        "gender": student.gender,
                    }
                )

            return APIResponse.success(data={"students": student_list})
        except Exception as e:
            return APIResponse.error(message=str(e))


@ns_classes.route("/validate-associations")
class ValidateAssociations(Resource):

    @ns_classes.doc("validate_associations")
    @requires_permission("class.manage")
    def get(self):
        return class_service.validate_associations()

    @ns_classes.doc("fix_associations")
    @requires_permission("class.manage")
    def post(self):
        return class_service.fix_associations()


@ns_classes.route("/export")
class ClassExport(Resource):

    @ns_classes.doc("export_classes", description="导出班级数据")
    @requires_permission("class.view")
    def get(self):
        args = ns_classes.parser.parse_args()
        keyword = args.get("keyword")
        export_format = request.args.get("format", "json").lower()

        result = class_service.export_classes(keyword, export_format)  # noqa: F841

        if result["type"] == "excel":
            return send_file(
                result["data"],
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=result["filename"],
            )
        else:
            return send_file(
                result["data"],
                mimetype="application/json",
                as_attachment=True,
                download_name=result["filename"],
            )


@ns_classes.route("/import")
class ClassImport(Resource):

    @ns_classes.doc("import_classes", description="导入班级数据")
    @requires_permission("class.manage")
    def post(self):
        from models import ImportConfig, get_by_id

        content_type = request.content_type or ""
        config_id = request.args.get("config_id", type=int)

        config = None
        if config_id:
            config = get_by_id(ImportConfig, config_id)
        else:
            config = ImportConfig.query.filter(
                ImportConfig.module_name == "classes",
                ImportConfig.is_default,
                ImportConfig.is_active,
            ).first()

        import_list = []

        if "multipart/form-data" in content_type:
            if "file" not in request.files:
                return APIResponse.bad_request(message="请上传文件")

            file = request.files["file"]
            if not file.filename:
                return APIResponse.bad_request(message="请选择文件")

            filename = file.filename.lower()
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                from openpyxl import load_workbook

                wb = load_workbook(file)
                ws = wb.active

                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)

                col_map = {}
                for idx, header in enumerate(headers):
                    if header:
                        col_map[header] = idx

                for row_idx in range(2, ws.max_row + 1):
                    row_data = {}
                    for header, col_idx in col_map.items():
                        row_data[header] = ws.cell(row=row_idx, column=col_idx + 1).value

                    default_mappings = [
                        {
                            "source_field": "班级名称",
                            "target_field": "name",
                            "field_type": "string",
                            "required": True,
                        },
                        {"source_field": "年级", "target_field": "grade", "field_type": "string"},
                        {
                            "source_field": "描述",
                            "target_field": "description",
                            "field_type": "string",
                        },
                        {
                            "source_field": "班主任ID",
                            "target_field": "head_teacher_id",
                            "field_type": "integer",
                            "relation": "admin",
                        },
                        {
                            "source_field": "班主任姓名",
                            "target_field": "head_teacher_name",
                            "field_type": "string",
                            "relation": "admin",
                        },
                        {
                            "source_field": "是否启用",
                            "target_field": "is_active",
                            "field_type": "boolean",
                        },
                    ]
                    field_mappings = config.field_mappings if config else default_mappings
                    default_values = config.default_values if config else {}

                    mapped_item = {}
                    for mapping in field_mappings:
                        source_val = row_data.get(mapping["source_field"])
                        target_field = mapping["target_field"]
                        field_type = mapping.get("field_type", "string")

                        if source_val is None:
                            if mapping.get("required"):
                                break
                            source_val = mapping.get(
                                "default_value", default_values.get(target_field)
                            )

                        if field_type == "boolean":
                            if isinstance(source_val, str):
                                mapped_item[target_field] = source_val in [
                                    "是",
                                    "true",
                                    "True",
                                    "1",
                                ]
                            else:
                                mapped_item[target_field] = bool(source_val)
                        elif field_type == "integer":
                            mapped_item[target_field] = int(source_val) if source_val else None
                        else:
                            mapped_item[target_field] = source_val

                    if mapped_item.get("name"):
                        import_list.append(mapped_item)
            else:
                return APIResponse.bad_request(message="仅支持 .xlsx 或 .xls 格式")
        elif "application/json" in content_type:
            data = request.json
            if not data or "data" not in data:
                return APIResponse.bad_request(message="导入数据格式错误")
            import_list = data["data"]
        else:
            return APIResponse.bad_request(message="不支持的文件格式")

        return class_service.import_classes(import_list, config)
