from flask import request
import openpyxl
from flask_restx import Namespace, Resource, fields
from utils.response import APIResponse
from models import db, Exam, Score, User, Admin, Subject, get_by_id
from utils.permission import requires_permission
from io import BytesIO
from services.excel_service import excel_import_service
from services.academics_service import academics_service


def _resolve_subject_id(subject_name, subject_id):
    """将科目名称或科目ID解析为 subject.id；均缺失返回 None。"""
    if subject_id:
        return subject_id
    if subject_name:
        sub = Subject.query.filter_by(name=subject_name).first()
        if sub:
            return sub.id
        sub = Subject.query.filter_by(code=subject_name).first()
        if sub:
            return sub.id
    return None


ns_exam_import = Namespace("exam-import", description="成绩导入增强功能")

exam_import_request = ns_exam_import.model(
    "ExamImportRequest",
    {
        "exam_id": fields.Integer(required=True, description="考试ID"),
        "entered_by": fields.Integer(description="录入人ID"),
        "update_existing": fields.Boolean(description="是否更新已存在的成绩", default=True),
        "validate_score": fields.Boolean(description="是否验证分数范围", default=True),
    },
)


def _parse_score_excel(file_content: bytes) -> dict:
    """统一解析成绩Excel文件，返回 headers, parsed_rows (list of dicts), total_count"""
    result = excel_import_service.parse_excel_file(file_content)  # noqa: F841
    if not result.get("success"):
        raise ValueError(result.get("error", "文件解析失败"))
    headers = result.get("headers", [])
    parsed_rows = result.get("data", [])
    return {"headers": headers, "parsed_rows": parsed_rows, "total_count": len(parsed_rows)}


class ScoreImportHelper:
    HEADER_MAPPING = {
        "card_id": ["card_id", "学号", "卡号", "id", "学生id"],
        "student_name": ["student_name", "姓名", "学生姓名"],
        "class_name": ["class_name", "班级", "班级名称"],
        "subject": ["subject", "科目", "考试科目"],
        "score": ["score", "分数", "成绩"],
        "full_score": ["full_score", "满分", "总分"],
        "remark": ["remark", "备注", "说明"],
    }

    @staticmethod
    def validate_headers(headers) -> dict:
        errors = []
        if not headers:
            errors.append("Excel文件没有数据行")
            return {"valid": False, "errors": errors, "headers": headers}

        header_lower = [str(h).lower().strip() if h else "" for h in headers]
        required_fields = ["card_id", "subject", "score"]
        for field in required_fields:
            found = False
            for alias in ScoreImportHelper.HEADER_MAPPING.get(field, []):
                if alias.lower() in header_lower:
                    found = True
                    break
            if not found:
                errors.append(f"缺少必需列: {field}（或中文表头）")
        return {"valid": len(errors) == 0, "errors": errors, "headers": headers}

    @staticmethod
    def validate_excel_format(sheet) -> dict:
        """验证Excel格式是否正确（兼容旧接口）"""
        if sheet.max_row < 2:
            return {"valid": False, "errors": ["Excel文件没有数据行"], "headers": []}
        headers = [cell.value for cell in sheet[1]]
        return ScoreImportHelper.validate_headers(headers)

    @staticmethod
    def find_column_index(headers, target_column: str) -> int:
        """查找列的索引，支持模糊匹配和中英文表头"""
        aliases = ScoreImportHelper.HEADER_MAPPING.get(target_column, [target_column])

        for i, header in enumerate(headers):
            if header:
                header_str = str(header).lower().strip()
                for alias in aliases:
                    if alias.lower() in header_str:
                        return i
        return -1

    @staticmethod
    def parse_score_value(value) -> float:
        """解析分数值"""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def validate_score_range(score: float, full_score: float = 100) -> tuple:
        """验证分数是否在合理范围内"""
        if score is None:
            return False, "分数为空"

        if score < 0:
            return False, "分数不能为负数"

        if score > full_score * 1.5:
            return False, f"分数超过满分的150% ({full_score * 1.5})"

        return True, "valid"


@ns_exam_import.route("/validate")
class ValidateImportFile(Resource):

    @ns_exam_import.doc("validate_import_file", description="验证导入文件格式")
    @requires_permission("score.entry")
    def post(self):
        """
        验证导入文件格式

        上传Excel文件后，先验证格式是否正确。
        """
        if "file" not in request.files:
            return APIResponse.error(message="没有上传文件"), 400

        file = request.files["file"]
        exam_id = request.args.get("exam_id", type=int)

        if not exam_id:
            return APIResponse.error(message="缺少考试ID"), 400

        try:
            exam = get_by_id(Exam, exam_id)
            if not exam:
                return APIResponse.error(message="考试不存在"), 404

            file_content = file.read()
            parsed = _parse_score_excel(file_content)
            headers = parsed["headers"]
            parsed_rows = parsed["parsed_rows"]
            total_count = parsed["total_count"]

            validation = ScoreImportHelper.validate_headers(headers)

            if not validation["valid"]:
                return APIResponse.error(
                    message="文件格式验证失败", errors=validation["errors"], status_code=400
                )

            data_preview = []
            card_id_idx = ScoreImportHelper.find_column_index(headers, "card_id")
            subject_idx = ScoreImportHelper.find_column_index(headers, "subject")
            score_idx = ScoreImportHelper.find_column_index(headers, "score")

            for i, row_data in enumerate(parsed_rows[:5]):
                card_id = row_data.get(headers[card_id_idx]) if card_id_idx >= 0 else None
                subject = (
                    row_data.get(headers[subject_idx]) if subject_idx >= 0 else None
                )  # noqa: F841
                score_val = row_data.get(headers[score_idx]) if score_idx >= 0 else None

                student = User.query.filter_by(card_id=str(card_id)).first() if card_id else None

                data_preview.append(
                    {
                        "row": i + 2,
                        "card_id": str(card_id) if card_id else None,
                        "student_name": student.name if student else "未找到",
                        "subject": subject,
                        "score": score_val,
                        "status": "ready" if student else "student_not_found",
                    }
                )

            return APIResponse.success(
                message="文件格式验证通过",
                data={
                    "exam_name": exam.name,
                    "subjects": exam.subjects,
                    "preview": data_preview,
                    "total_rows": total_count,
                },
            )

        except Exception as e:
            return APIResponse.error(message=f"验证失败: {str(e)}", status_code=500)


@ns_exam_import.route("/preview")
class PreviewImportData(Resource):

    @ns_exam_import.doc("preview_import_data", description="预览导入数据")
    @requires_permission("score.entry")
    def post(self):
        """
        预览导入数据

        返回导入数据的预览，不实际写入数据库。
        """
        if "file" not in request.files:
            return APIResponse.error(message="没有上传文件"), 400

        file = request.files["file"]
        exam_id = request.form.get("exam_id", type=int)

        if not exam_id:
            return APIResponse.error(message="缺少考试ID"), 400

        try:
            exam = get_by_id(Exam, exam_id)
            if not exam:
                return APIResponse.error(message="考试不存在"), 404

            file_content = file.read()
            parsed = _parse_score_excel(file_content)
            headers = parsed["headers"]
            parsed_rows = parsed["parsed_rows"]
            total_count = parsed["total_count"]

            validation = ScoreImportHelper.validate_headers(headers)
            if not validation["valid"]:
                return APIResponse.error(
                    message="文件格式错误", errors=validation["errors"], status_code=400
                )

            card_id_idx = ScoreImportHelper.find_column_index(headers, "card_id")
            subject_idx = ScoreImportHelper.find_column_index(headers, "subject")
            score_idx = ScoreImportHelper.find_column_index(headers, "score")
            full_score_idx = ScoreImportHelper.find_column_index(headers, "full_score")
            remark_idx = ScoreImportHelper.find_column_index(headers, "remark")

            results = []
            errors = []

            for i, row_data in enumerate(parsed_rows):
                card_id = (
                    str(row_data.get(headers[card_id_idx], "")).strip()
                    if card_id_idx >= 0 and row_data.get(headers[card_id_idx])
                    else None
                )
                subject = (
                    row_data.get(headers[subject_idx]) if subject_idx >= 0 else None
                )  # noqa: F841
                score_val = (
                    ScoreImportHelper.parse_score_value(row_data.get(headers[score_idx]))
                    if score_idx >= 0
                    else None
                )
                full_score = (
                    ScoreImportHelper.parse_score_value(row_data.get(headers[full_score_idx]))
                    if full_score_idx >= 0
                    else 100
                )
                remark = (
                    str(row_data.get(headers[remark_idx], "")).strip()
                    if remark_idx >= 0 and row_data.get(headers[remark_idx])
                    else None
                )

                if not card_id:
                    errors.append(f"行{i+2}: 学号为空")
                    continue

                student = User.query.filter_by(card_id=card_id).first()
                if not student:
                    errors.append(f"行{i+2}: 学号{card_id}不存在")
                    continue

                if not subject:
                    errors.append(f"行{i+2}: 科目为空")
                    continue
                # F2 修复: preview 此前未解析 subject_id → NameError 恒 500；与 execute 保持一致
                subject_id = _resolve_subject_id(subject, None)
                if subject_id is None:
                    errors.append(f"行{i+2}: 科目「{subject}」未配置")
                    continue

                is_valid, msg = ScoreImportHelper.validate_score_range(score_val, full_score)
                if not is_valid:
                    errors.append(f"行{i+2}: {student.name}-{subject} - {msg}")

                existing_score = Score.query.filter_by(
                    exam_id=exam_id, student_id=student.id, subject_id=subject_id
                ).first()

                results.append(
                    {
                        "row": i + 2,
                        "card_id": card_id,
                        "student_name": student.name,
                        "class_name": student.class_name,
                        "subject": subject,
                        "score": score_val,
                        "full_score": full_score,
                        "remark": remark,
                        "will_update": existing_score is not None,
                        "will_insert": existing_score is None,
                    }
                )

            return APIResponse.success(
                message=f"预览完成，共{total_count}行",
                data={
                    "results": results[:50],
                    "errors": errors[:20],
                    "summary": {
                        "total": len(results),
                        "will_insert": sum(1 for r in results if r["will_insert"]),
                        "will_update": sum(1 for r in results if r["will_update"]),
                        "error_count": len(errors),
                    },
                },
            )

        except Exception as e:
            return APIResponse.error(message=f"预览失败: {str(e)}", status_code=500)


@ns_exam_import.route("/execute")
class ExecuteImport(Resource):

    @ns_exam_import.doc("execute_import", description="执行成绩导入")
    @requires_permission("score.entry")
    def post(self):
        """
        执行成绩导入

        实际写入数据库。
        """
        if "file" not in request.files:
            return APIResponse.error(message="没有上传文件"), 400

        file = request.files["file"]
        exam_id = request.form.get("exam_id", type=int)
        entered_by = request.form.get("entered_by", type=int, default=1)
        update_existing = request.form.get("update_existing", "true").lower() == "true"
        validate_score = request.form.get("validate_score", "true").lower() == "true"

        if not exam_id:
            return APIResponse.error(message="缺少考试ID"), 400

        try:
            exam = get_by_id(Exam, exam_id)
            if not exam:
                return APIResponse.error(message="考试不存在"), 404

            if exam.status != "published":
                return APIResponse.error(message="考试未发布，无法导入成绩"), 400

            file_content = file.read()
            parsed = _parse_score_excel(file_content)
            headers = parsed["headers"]
            parsed_rows = parsed["parsed_rows"]

            validation = ScoreImportHelper.validate_headers(headers)
            if not validation["valid"]:
                return APIResponse.error(
                    message="文件格式错误", errors=validation["errors"], status_code=400
                )

            result = academics_service.execute_score_import(
                exam_id=exam_id,
                entered_by=entered_by,
                update_existing=update_existing,
                validate_score=validate_score,
                parsed_rows=parsed_rows,
                headers=headers,
            )

            return APIResponse.success(
                message="导入完成",
                data=result,
            )

        except Exception as e:
            db.session.rollback()
            return APIResponse.error(message=f"导入失败: {str(e)}", status_code=500)


@ns_exam_import.route("/template")
class DownloadTemplate(Resource):

    @ns_exam_import.doc("download_template", description="下载导入模板")
    @requires_permission("score.view")
    def get(self):
        """
        下载成绩导入Excel模板
        """
        exam_id = request.args.get("exam_id", type=int)
        class_id = request.args.get("class_id", type=int)

        exam = None
        if exam_id:
            exam = get_by_id(Exam, exam_id)

        subjects = exam.subjects if exam else ["语文", "数学", "英语"]

        # 获取学生列表
        if class_id:
            # 如果指定了班级，获取该班级所有学生
            students = (
                User.query.filter_by(class_id=class_id, role="student").order_by(User.card_id).all()
            )
        else:
            # 如果没有指定班级，获取所有学生（用于全校考试）
            students = (
                User.query.filter_by(role="student").order_by(User.class_name, User.card_id).all()
            )

        wb = openpyxl.Workbook()

        # 1. 创建成绩导入表
        sheet_data = wb.active
        sheet_data.title = "成绩导入"

        # 2. 创建说明表
        sheet_notes = wb.create_sheet(title="填写说明")

        # --- 填写说明表 ---
        notes_header_style = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
        notes_header_fill = openpyxl.styles.PatternFill(
            start_color="4A90D9", end_color="4A90D9", fill_type="solid"
        )
        notes_header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        notes_data = [
            ["列名", "说明", "填写方式", "示例"],
            ["学号", "学生的学号，系统自动填入，请勿修改", "系统自动", "202401001"],
            ["姓名", "学生姓名，系统自动填入，仅作参考", "系统自动", "张三"],
            ["班级", "班级名称，系统自动填入", "系统自动", "高一(1)班"],
            ["科目", "科目名称，系统自动填入对应考试科目", "系统自动", "语文"],
            ["分数", "学生成绩，教师必须填写，必须为数字", "教师填写", "85"],
            ["满分", "科目满分，默认100，可修改", "默认100", "100"],
            ["备注", "成绩备注信息，可选填写", "可选", "进步明显"],
        ]

        for row_idx, row_data in enumerate(notes_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = sheet_notes.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 1:
                    cell.font = notes_header_style
                    cell.fill = notes_header_fill
                    cell.alignment = notes_header_alignment

        # 设置列宽
        sheet_notes.column_dimensions["A"].width = 20
        sheet_notes.column_dimensions["B"].width = 40
        sheet_notes.column_dimensions["C"].width = 10
        sheet_notes.column_dimensions["D"].width = 20

        # 添加考试信息
        if exam:
            sheet_notes.cell(row=10, column=1, value="考试信息")
            sheet_notes.cell(row=10, column=1).font = openpyxl.styles.Font(bold=True)
            sheet_notes.cell(row=11, column=1, value=f"考试名称: {exam.name}")
            sheet_notes.cell(row=12, column=1, value=f'考试科目: {", ".join(exam.subjects)}')
            if exam.start_time:
                sheet_notes.cell(
                    row=13,
                    column=1,
                    value=f'开始时间: {exam.start_time.strftime("%Y-%m-%d %H:%M")}',
                )
            if exam.end_time:
                sheet_notes.cell(
                    row=14, column=1, value=f'结束时间: {exam.end_time.strftime("%Y-%m-%d %H:%M")}'
                )

        # --- 成绩导入表 ---
        header_style = openpyxl.styles.Font(bold=True, size=11, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(
            start_color="4A90D9", end_color="4A90D9", fill_type="solid"
        )
        header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        headers = ["学号", "姓名", "班级", "科目", "分数", "满分", "备注"]
        sheet_data.append(headers)

        # 应用样式
        for col_idx in range(1, len(headers) + 1):
            cell = sheet_data.cell(row=1, column=col_idx)
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 设置列宽
        col_widths = [15, 12, 15, 12, 10, 10, 20]
        for col_idx, width in enumerate(col_widths, 1):
            sheet_data.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 添加学生数据 - 为每个学生生成所有科目的成绩行
        for student in students:
            for subject in subjects:
                sheet_data.append(
                    [
                        student.card_id,  # 学号（已填入）
                        student.name,  # 姓名（已填入）
                        student.class_name or "",  # 班级（已填入）
                        subject,  # 科目（已填入）
                        "",  # 分数（教师填写）
                        100,  # 满分（默认100）
                        "",  # 备注（可选填写）
                    ]
                )

        # 添加数据验证 - 确保分数是数字
        from openpyxl.worksheet.datavalidation import DataValidation

        score_validation = DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2="200",
            allow_blank=True,
            errorTitle="分数无效",
            error="请输入有效的分数（0-200之间）",
        )
        sheet_data.add_data_validation(score_validation)
        score_validation.add("E2:E1000")

        # 冻结首行
        sheet_data.freeze_panes = "A2"

        # 添加自动筛选
        sheet_data.auto_filter.ref = sheet_data.dimensions

        # 添加条件格式 - 分数列
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Font, PatternFill

        # 红色 - 不及格（假设满分100，60分以下）
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        red_rule = CellIsRule(operator="lessThan", formula=["60"], fill=red_fill, font=red_font)

        # 绿色 - 优秀（90分以上）
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100")
        green_rule = CellIsRule(
            operator="greaterThan", formula=["90"], fill=green_fill, font=green_font
        )

        sheet_data.conditional_formatting.add("E2:E1000", red_rule)
        sheet_data.conditional_formatting.add("E2:E1000", green_rule)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import send_file

        filename = f'成绩导入模板_{exam.name if exam else "通用"}.xlsx'
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )


@ns_exam_import.route("/history")
class ImportHistory(Resource):

    @ns_exam_import.doc("get_import_history", description="获取导入历史")
    @requires_permission("score.view")
    def get(self):
        """
        获取成绩导入历史记录
        """
        exam_id = request.args.get("exam_id", type=int)
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 20, type=int)

        query = Score.query.filter(Score.entered_by.isnot(None))

        if exam_id:
            query = query.filter_by(exam_id=exam_id)

        pagination = query.order_by(Score.entered_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )

        results = []
        for score in pagination.items:
            student = get_by_id(User, score.student_id)
            exam = get_by_id(Exam, score.exam_id)
            entered_by_admin = get_by_id(Admin, score.entered_by) if score.entered_by else None

            results.append(
                {
                    "id": score.id,
                    "exam_name": exam.name if exam else None,
                    "student_name": student.name if student else None,
                    "student_card_id": student.card_id if student else None,
                    "subject": score.subject_rel.name if score.subject_rel else "",
                    "score": score.score,
                    "full_score": score.full_score,
                    "status": score.status,
                    "rank": None,  # R7: Score.rank 列废弃，无排名上下文
                    "entered_by": entered_by_admin.username if entered_by_admin else None,
                    "entered_at": score.entered_at.isoformat() if score.entered_at else None,
                }
            )

        return APIResponse.success(
            data={
                "data": results,
                "total": pagination.total,
                "page": page,
                "per_page": per_page,
                "pages": pagination.pages,
            }
        )
