import json
import io
from flask_restx import Namespace, Resource, fields
from flask import request, send_file
from models import CourseSchedule, ClassInfo, Subject, ClassPeriod, Admin, ImportConfig, get_by_id
from services.class_time_checker import ClassTimeChecker
from services.academics_service import academics_service
from utils.permission import requires_permission, get_allowed_classes, get_current_admin
from utils.response import APIResponse
from utils.api_cache_middleware import cached_api, invalidate_cache
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ns_course_schedule = Namespace("course-schedules", description="课程表相关操作")

ns_course_schedule.parser = ns_course_schedule.parser()
ns_course_schedule.parser.add_argument(
    "class_info_id", type=int, location="args", required=False, help="班级ID"
)
ns_course_schedule.parser.add_argument(
    "day_of_week", type=int, location="args", required=False, help="星期"
)
ns_course_schedule.parser.add_argument(
    "period_number", type=int, location="args", required=False, help="节次编号"
)
ns_course_schedule.parser.add_argument(
    "is_active", type=bool, location="args", required=False, help="是否启用"
)
ns_course_schedule.parser.add_argument(
    "teacher_name", type=str, location="args", required=False, help="教师姓名"
)
ns_course_schedule.parser.add_argument(
    "classroom", type=str, location="args", required=False, help="教室"
)

course_schedule_model = ns_course_schedule.model(
    "CourseSchedule",
    {
        "id": fields.Integer(readOnly=True, description="课程ID"),
        "class_info_id": fields.Integer(required=True, description="班级ID"),
        "subject_id": fields.Integer(required=True, description="科目ID"),
        "day_of_week": fields.Integer(required=True, description="星期(0=周一~6=周日)"),
        "period_number": fields.Integer(required=True, description="节次编号"),
        "teacher_id": fields.Integer(description="教师ID"),
        "teacher_name": fields.String(description="教师姓名"),
        "classroom": fields.String(description="教室"),
        "description": fields.String(description="描述"),
        "color": fields.String(description="颜色"),
        "is_active": fields.Boolean(description="是否启用"),
    },
)

course_schedule_response = ns_course_schedule.model(
    "CourseScheduleResponse",
    {
        "id": fields.Integer(description="课程ID"),
        "class_info_id": fields.Integer(description="班级ID"),
        "class_name": fields.String(description="班级名称"),
        "subject_id": fields.Integer(description="科目ID"),
        "subject_name": fields.String(description="科目名称"),
        "subject_color": fields.String(description="科目颜色"),
        "day_of_week": fields.Integer(description="星期"),
        "day_of_week_text": fields.String(description="星期文本"),
        "period_number": fields.Integer(description="节次编号"),
        "period_name": fields.String(description="节次名称"),
        "period_time": fields.String(description="节次时间"),
        "teacher_id": fields.Integer(description="教师ID"),
        "teacher_name": fields.String(description="教师姓名"),
        "classroom": fields.String(description="教室"),
        "description": fields.String(description="描述"),
        "color": fields.String(description="颜色"),
        "is_active": fields.Boolean(description="是否启用"),
        "created_at": fields.String(description="创建时间"),
        "updated_at": fields.String(description="更新时间"),
    },
)


def format_day_of_week(day):
    days = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    return days[day] if 0 <= day <= 6 else "未知"


def get_period_info(period_number):
    period = ClassPeriod.query.filter_by(period_number=period_number).first()
    if period:
        return {
            "name": period.name,
            "time": (
                f"{period.start_hour:02d}:{period.start_minute:02d} - "
                f"{period.end_hour:02d}:{period.end_minute:02d}"
            ),
        }
    return {"name": f"第{period_number}节", "time": ""}


def check_conflicts(class_info_id, day_of_week, period_number, exclude_id=None):
    """
    检查课程时间冲突：
    1. 班级在同一时段已有课程（唯一约束）
    2. 教师在同一时段已有其他课程
    3. 教室在同一时段已有其他课程
    """
    conflicts = []

    # 检查班级冲突
    class_conflict = CourseSchedule.query.filter(
        CourseSchedule.id != (exclude_id or -1),
        CourseSchedule.class_info_id == class_info_id,
        CourseSchedule.day_of_week == day_of_week,
        CourseSchedule.period_number == period_number,
        CourseSchedule.is_active,
    ).first()

    if class_conflict:
        subject_name = class_conflict.subject.name if class_conflict.subject else ""
        conflicts.append(
            {
                "type": "class",
                "message": f"该班级此时段已有课程：{subject_name}",
                "schedule_id": class_conflict.id,
                "conflicting_class_name": (
                    class_conflict.class_info.name if class_conflict.class_info else ""
                ),
                "conflicting_subject_name": subject_name,
                "conflicting_teacher_name": class_conflict.teacher_name,
                "conflicting_classroom": class_conflict.classroom,
            }
        )

    return conflicts


def check_teacher_conflicts(
    teacher_name, day_of_week, period_number, exclude_id=None, exclude_class_id=None
):
    """检查教师时间冲突"""
    conflicts = []

    if not teacher_name:
        return conflicts

    teacher_conflicts = CourseSchedule.query.filter(
        CourseSchedule.id != (exclude_id or -1),
        CourseSchedule.teacher_name == teacher_name,
        CourseSchedule.day_of_week == day_of_week,
        CourseSchedule.period_number == period_number,
        CourseSchedule.is_active,
    ).all()

    for conflict in teacher_conflicts:
        # 如果是同一个班级的同一课程，不算冲突
        if exclude_class_id and conflict.class_info_id == exclude_class_id:
            continue

        class_name = conflict.class_info.name if conflict.class_info else ""
        subject_name = conflict.subject.name if conflict.subject else ""
        conflicts.append(
            {
                "type": "teacher",
                "message": f"教师 {teacher_name} 此时段已有课程：{class_name} - {subject_name}",
                "schedule_id": conflict.id,
                "conflicting_class_name": class_name,
                "conflicting_subject_name": subject_name,
                "conflicting_teacher_name": conflict.teacher_name,
                "conflicting_classroom": conflict.classroom,
            }
        )

    return conflicts


def check_classroom_conflicts(
    classroom, day_of_week, period_number, exclude_id=None, exclude_class_id=None
):
    """检查教室时间冲突"""
    conflicts = []

    if not classroom:
        return conflicts

    classroom_conflicts = CourseSchedule.query.filter(
        CourseSchedule.id != (exclude_id or -1),
        CourseSchedule.classroom == classroom,
        CourseSchedule.day_of_week == day_of_week,
        CourseSchedule.period_number == period_number,
        CourseSchedule.is_active,
    ).all()

    for conflict in classroom_conflicts:
        # 如果是同一个班级的同一课程，不算冲突
        if exclude_class_id and conflict.class_info_id == exclude_class_id:
            continue

        class_name = conflict.class_info.name if conflict.class_info else ""
        subject_name = conflict.subject.name if conflict.subject else ""
        conflicts.append(
            {
                "type": "classroom",
                "message": f"教室 {classroom} 此时段已有课程：{class_name} - {subject_name}",
                "schedule_id": conflict.id,
                "conflicting_class_name": class_name,
                "conflicting_subject_name": subject_name,
                "conflicting_teacher_name": conflict.teacher_name,
                "conflicting_classroom": conflict.classroom,
            }
        )

    return conflicts


@ns_course_schedule.route("/")
class CourseScheduleList(Resource):

    @ns_course_schedule.doc("list_course_schedule", description="获取课程表列表")
    @ns_course_schedule.response(200, "成功")
    @requires_permission("schedule.view")
    @cached_api(ttl=30)
    def get(self):
        """
        获取课程表列表。非管理员用户只能查看关联班级的课程表。

        可选参数：
        - class_info_id: 班级ID
        - day_of_week: 星期
        - teacher_name: 教师姓名
        - classroom: 教室
        """
        args = ns_course_schedule.parser.parse_args()
        class_info_id = args.get("class_info_id")
        day_of_week = args.get("day_of_week")
        teacher_name = args.get("teacher_name")
        classroom = args.get("classroom")

        query = CourseSchedule.query.filter_by(is_active=True)

        # 数据隔离：非管理员只能查看关联班级的课程
        admin = get_current_admin()
        allowed_classes = get_allowed_classes(admin.id) if admin else None
        if allowed_classes is not None:
            class_ids = [
                c.id for c in ClassInfo.query.filter(ClassInfo.name.in_(allowed_classes)).all()
            ]
            query = query.filter(CourseSchedule.class_info_id.in_(class_ids))

        if class_info_id:
            query = query.filter_by(class_info_id=class_info_id)
        if day_of_week is not None:
            query = query.filter_by(day_of_week=day_of_week)
        if teacher_name:
            query = query.filter(CourseSchedule.teacher_name.like(f"%{teacher_name}%"))
        if classroom:
            query = query.filter(CourseSchedule.classroom.like(f"%{classroom}%"))

        schedules = query.order_by(CourseSchedule.day_of_week, CourseSchedule.period_number).all()

        result = []  # noqa: F841
        for schedule in schedules:
            period_info = get_period_info(schedule.period_number)
            result.append(
                {
                    "id": schedule.id,
                    "class_info_id": schedule.class_info_id,
                    "class_name": schedule.class_info.name if schedule.class_info else "",
                    "subject_id": schedule.subject_id,
                    "subject_name": schedule.subject.name if schedule.subject else "",
                    "subject_color": schedule.subject.color if schedule.subject else schedule.color,
                    "day_of_week": schedule.day_of_week,
                    "day_of_week_text": format_day_of_week(schedule.day_of_week),
                    "period_number": schedule.period_number,
                    "period_name": period_info["name"],
                    "period_time": period_info["time"],
                    "teacher_name": schedule.teacher_name,
                    "classroom": schedule.classroom,
                    "description": schedule.description,
                    "color": schedule.color,
                    "is_active": schedule.is_active,
                    "created_at": schedule.created_at.isoformat() if schedule.created_at else None,
                    "updated_at": schedule.updated_at.isoformat() if schedule.updated_at else None,
                }
            )

        return APIResponse.success(data={"schedules": result})

    @ns_course_schedule.doc("create_course_schedule", description="创建课程安排", security="Bearer")
    @ns_course_schedule.expect(course_schedule_model)
    @ns_course_schedule.response(201, "创建成功", course_schedule_response)
    @requires_permission("schedule.manage")
    def post(self):
        """
        创建课程安排。需要课表管理权限。

        请求体：
        - class_info_id: 班级ID（必填）
        - subject_id: 科目ID（必填）
        - day_of_week: 星期(0-6)（必填）
        - period_number: 节次编号（必填）
        - teacher_name: 教师姓名（可选）
        - classroom: 教室（可选）
        - description: 描述（可选）
        - color: 颜色（可选）
        """
        data = ns_course_schedule.payload

        # 数据隔离：非管理员只能为关联班级创建课程
        admin = get_current_admin()
        allowed_classes = get_allowed_classes(admin.id) if admin else None
        if allowed_classes is not None:
            class_info = get_by_id(ClassInfo, data["class_info_id"])
            if class_info and class_info.name not in allowed_classes:
                return APIResponse.forbidden(message="无权为该班级创建课程安排")

        # 检查时间冲突
        conflicts = []
        conflicts.extend(
            check_conflicts(data["class_info_id"], data["day_of_week"], data["period_number"])
        )

        teacher_id = data.get("teacher_id")
        teacher_name = data.get("teacher_name")

        # 如果提供了teacher_id，验证教师存在并获取教师姓名
        if teacher_id:
            teacher = get_by_id(Admin, teacher_id)
            if teacher:
                teacher_name = teacher.real_name or teacher.username
            else:
                return APIResponse.bad_request(message=f'教师ID "{teacher_id}" 在系统中不存在')

        if teacher_name:
            conflicts.extend(
                check_teacher_conflicts(teacher_name, data["day_of_week"], data["period_number"])
            )

        classroom = data.get("classroom")
        if classroom:
            conflicts.extend(
                check_classroom_conflicts(classroom, data["day_of_week"], data["period_number"])
            )

        if conflicts:
            return APIResponse.bad_request(message="存在时间冲突", errors=conflicts)

        subject = get_by_id(Subject, data["subject_id"])  # noqa: F841
        color = data.get("color") or (subject.color if subject else "#3B82F6")

        schedule_id = academics_service.create_course_schedule(
            {
                "class_info_id": data["class_info_id"],
                "subject_id": data["subject_id"],
                "day_of_week": data["day_of_week"],
                "period_number": data["period_number"],
                "teacher_id": teacher_id,
                "teacher_name": teacher_name,
                "classroom": classroom,
                "description": data.get("description"),
                "color": color,
                "is_active": data.get("is_active", True),
            }
        )
        schedule = get_by_id(CourseSchedule, schedule_id)

        period_info = get_period_info(schedule.period_number)
        invalidate_cache("api:/api/course-schedules/*")
        return (
            APIResponse.success(
                data={
                    "id": schedule.id,
                    "class_info_id": schedule.class_info_id,
                    "class_name": schedule.class_info.name if schedule.class_info else "",
                    "subject_id": schedule.subject_id,
                    "subject_name": schedule.subject.name if schedule.subject else "",
                    "subject_color": schedule.subject.color if schedule.subject else schedule.color,
                    "day_of_week": schedule.day_of_week,
                    "day_of_week_text": format_day_of_week(schedule.day_of_week),
                    "period_number": schedule.period_number,
                    "period_name": period_info["name"],
                    "period_time": period_info["time"],
                    "teacher_id": schedule.teacher_id,
                    "teacher_name": schedule.teacher_name,
                    "classroom": schedule.classroom,
                    "description": schedule.description,
                    "color": schedule.color,
                    "is_active": schedule.is_active,
                    "created_at": schedule.created_at.isoformat() if schedule.created_at else None,
                },
                message="课程安排创建成功",
            ),
            201,
        )


@ns_course_schedule.route("/<int:id>")
@ns_course_schedule.param("id", "课程ID")
class CourseScheduleResource(Resource):

    @ns_course_schedule.doc("get_course_schedule", description="获取课程详情")
    @ns_course_schedule.response(200, "成功", course_schedule_response)
    @ns_course_schedule.response(404, "课程不存在")
    @requires_permission("schedule.view")
    def get(self, id):
        """获取课程详情。需要课表查看权限。"""
        schedule = CourseSchedule.query.get_or_404(id)

        # 数据隔离：非管理员只能查看关联班级的课程
        admin = get_current_admin()
        allowed_classes = get_allowed_classes(admin.id) if admin else None
        if allowed_classes is not None:
            if schedule.class_info and schedule.class_info.name not in allowed_classes:
                return APIResponse.forbidden(message="无权查看该课程")

        period_info = get_period_info(schedule.period_number)

        return APIResponse.success(
            data={
                "id": schedule.id,
                "class_info_id": schedule.class_info_id,
                "class_name": schedule.class_info.name if schedule.class_info else "",
                "subject_id": schedule.subject_id,
                "subject_name": schedule.subject.name if schedule.subject else "",
                "subject_color": schedule.subject.color if schedule.subject else schedule.color,
                "day_of_week": schedule.day_of_week,
                "day_of_week_text": format_day_of_week(schedule.day_of_week),
                "period_number": schedule.period_number,
                "period_name": period_info["name"],
                "period_time": period_info["time"],
                "teacher_name": schedule.teacher_name,
                "classroom": schedule.classroom,
                "description": schedule.description,
                "color": schedule.color,
                "is_active": schedule.is_active,
                "created_at": schedule.created_at.isoformat() if schedule.created_at else None,
                "updated_at": schedule.updated_at.isoformat() if schedule.updated_at else None,
            }
        )

    @ns_course_schedule.doc("update_course_schedule", description="更新课程安排", security="Bearer")
    @ns_course_schedule.expect(course_schedule_model)
    @ns_course_schedule.response(200, "更新成功")
    @ns_course_schedule.response(404, "课程不存在")
    @requires_permission("schedule.manage")
    def put(self, id):
        """更新课程安排。需要课表管理权限。"""
        schedule = CourseSchedule.query.get_or_404(id)

        # 数据隔离：非管理员只能修改关联班级的课程
        admin = get_current_admin()
        allowed_classes = get_allowed_classes(admin.id) if admin else None
        if allowed_classes is not None:
            if schedule.class_info and schedule.class_info.name not in allowed_classes:
                return APIResponse.forbidden(message="无权修改该课程")

        data = ns_course_schedule.payload

        # 获取更新后的字段值
        new_class_info_id = data.get("class_info_id", schedule.class_info_id)
        new_day_of_week = data.get("day_of_week", schedule.day_of_week)
        new_period_number = data.get("period_number", schedule.period_number)
        new_teacher_id = data.get("teacher_id", schedule.teacher_id)
        new_teacher_name = data.get("teacher_name", schedule.teacher_name)
        new_classroom = data.get("classroom", schedule.classroom)

        # 如果提供了teacher_id，验证教师存在并获取教师姓名
        if new_teacher_id:
            teacher = get_by_id(Admin, new_teacher_id)
            if teacher:
                new_teacher_name = teacher.real_name or teacher.username
            else:
                return APIResponse.bad_request(message=f'教师ID "{new_teacher_id}" 在系统中不存在')

        # 如果时间或班级发生变化，检查冲突
        if any(
            [
                new_class_info_id != schedule.class_info_id,
                new_day_of_week != schedule.day_of_week,
                new_day_of_week != schedule.day_of_week,
            ]
        ):

            conflicts = []
            conflicts.extend(
                check_conflicts(
                    new_class_info_id, new_day_of_week, new_period_number, exclude_id=id
                )
            )

            if new_teacher_name:
                conflicts.extend(
                    check_teacher_conflicts(
                        new_teacher_name,
                        new_day_of_week,
                        new_period_number,
                        exclude_id=id,
                        exclude_class_id=new_class_info_id,
                    )
                )

            if new_classroom:
                conflicts.extend(
                    check_classroom_conflicts(
                        new_classroom,
                        new_day_of_week,
                        new_period_number,
                        exclude_id=id,
                        exclude_class_id=new_class_info_id,
                    )
                )

            if conflicts:
                return APIResponse.bad_request(message="存在时间冲突", errors=conflicts)

        # 如果教师或教室发生变化，检查冲突
        if new_teacher_name != schedule.teacher_name:
            teacher_conflicts = check_teacher_conflicts(
                new_teacher_name,
                new_day_of_week,
                new_period_number,
                exclude_id=id,
                exclude_class_id=new_class_info_id,
            )
            if teacher_conflicts:
                return APIResponse.bad_request(message="教师时间冲突", errors=teacher_conflicts)

        if new_classroom != schedule.classroom:
            classroom_conflicts = check_classroom_conflicts(
                new_classroom,
                new_day_of_week,
                new_period_number,
                exclude_id=id,
                exclude_class_id=new_class_info_id,
            )
            if classroom_conflicts:
                return APIResponse.bad_request(message="教室时间冲突", errors=classroom_conflicts)

        new_subject_id = data.get("subject_id", schedule.subject_id)
        if "color" in data:
            final_color = data["color"]
        else:
            color_subject = get_by_id(Subject, new_subject_id)  # noqa: F841
            final_color = color_subject.color if color_subject else None

        academics_service.update_course_schedule(
            id,
            {
                "class_info_id": new_class_info_id,
                "subject_id": new_subject_id,
                "day_of_week": new_day_of_week,
                "period_number": new_period_number,
                "teacher_id": new_teacher_id,
                "teacher_name": new_teacher_name,
                "classroom": new_classroom,
                "description": data.get("description", schedule.description),
                "color": final_color,
                "is_active": data.get("is_active", schedule.is_active),
            },
        )
        schedule = get_by_id(CourseSchedule, id)

        period_info = get_period_info(schedule.period_number)
        invalidate_cache("api:/api/course-schedules/*")
        return APIResponse.success(
            data={
                "success": True,
                "message": "课程安排更新成功",
                "schedule": {
                    "id": schedule.id,
                    "class_info_id": schedule.class_info_id,
                    "class_name": schedule.class_info.name if schedule.class_info else "",
                    "subject_id": schedule.subject_id,
                    "subject_name": schedule.subject.name if schedule.subject else "",
                    "day_of_week": schedule.day_of_week,
                    "day_of_week_text": format_day_of_week(schedule.day_of_week),
                    "period_number": schedule.period_number,
                    "period_name": period_info["name"],
                    "period_time": period_info["time"],
                    "teacher_id": schedule.teacher_id,
                    "teacher_name": schedule.teacher_name,
                    "classroom": schedule.classroom,
                    "color": schedule.color,
                    "is_active": schedule.is_active,
                    "updated_at": schedule.updated_at.isoformat() if schedule.updated_at else None,
                },
            }
        )

    @ns_course_schedule.doc("delete_course_schedule", description="删除课程安排", security="Bearer")
    @ns_course_schedule.response(200, "删除成功")
    @ns_course_schedule.response(404, "课程不存在")
    @requires_permission("schedule.manage")
    def delete(self, id):
        """删除课程安排。需要课表管理权限。"""
        schedule = CourseSchedule.query.get_or_404(id)

        # 数据隔离：非管理员只能删除关联班级的课程
        admin = get_current_admin()
        allowed_classes = get_allowed_classes(admin.id) if admin else None
        if allowed_classes is not None:
            if schedule.class_info and schedule.class_info.name not in allowed_classes:
                return APIResponse.forbidden(message="无权删除该课程")

        academics_service.delete_course_schedule(id)
        invalidate_cache("api:/api/course-schedules/*")
        return APIResponse.success(message="课程安排删除成功")


@ns_course_schedule.route("/class/<int:class_info_id>")
@ns_course_schedule.param("class_info_id", "班级ID")
class CourseScheduleByClass(Resource):

    @ns_course_schedule.doc("get_course_schedule_by_class", description="获取班级课程表")
    @ns_course_schedule.response(200, "成功")
    @requires_permission("schedule.view")
    @cached_api(ttl=30)
    def get(self, class_info_id):
        """获取指定班级的完整课程表。非管理员用户只能查看关联班级的课程表。"""
        # 数据隔离：非管理员只能查看关联班级的课程
        admin = get_current_admin()
        allowed_classes = get_allowed_classes(admin.id) if admin else None
        if allowed_classes is not None:
            class_info = get_by_id(ClassInfo, class_info_id)
            if class_info and class_info.name not in allowed_classes:
                return APIResponse.forbidden(message="无权查看该班级课程表")

        schedules = CourseSchedule.query.filter(
            CourseSchedule.class_info_id == class_info_id, CourseSchedule.is_active
        ).order_by(CourseSchedule.day_of_week, CourseSchedule.period_number)

        class_info = get_by_id(ClassInfo, class_info_id)
        periods = ClassPeriod.query.filter_by(is_active=True).order_by(ClassPeriod.sort_order).all()

        result = {  # noqa: F841
            "class_info_id": class_info_id,
            "class_name": class_info.name if class_info else "",
            "periods": [p.to_dict() for p in periods],
            "schedules": [],
        }

        for schedule in schedules:
            period_info = get_period_info(schedule.period_number)
            result["schedules"].append(
                {
                    "id": schedule.id,
                    "class_info_id": schedule.class_info_id,
                    "class_name": schedule.class_info.name if schedule.class_info else "",
                    "subject_id": schedule.subject_id,
                    "subject_name": schedule.subject.name if schedule.subject else "",
                    "subject_color": schedule.subject.color if schedule.subject else schedule.color,
                    "day_of_week": schedule.day_of_week,
                    "day_of_week_text": format_day_of_week(schedule.day_of_week),
                    "period_number": schedule.period_number,
                    "period_name": period_info["name"],
                    "period_time": period_info["time"],
                    "teacher_name": schedule.teacher_name,
                    "classroom": schedule.classroom,
                    "description": schedule.description,
                    "color": schedule.color,
                    "is_active": schedule.is_active,
                }
            )

        return result


@ns_course_schedule.route("/now")
class CourseScheduleNow(Resource):
    @ns_course_schedule.doc("get_course_schedule_now", description="获取当前时刻班级上课状态")
    @ns_course_schedule.response(200, "成功")
    @requires_permission("schedule.view")
    def get(self):
        """返回当前是否处于上课时间、当前节次与科目。

        可传 class_info_id 精确查询某班；也可传 device_id 按设备反查班级
        （与下发端点 _resolve_class_from_device 口径一致，供"指定设备"模式的徽章使用）。
        """
        class_info_id = request.args.get("class_info_id", type=int)
        if not class_info_id:
            device_id = request.args.get("device_id", type=str)
            if device_id:
                try:
                    from models import Device

                    dev = Device.query.filter_by(device_id=str(device_id)).first()
                    if dev and dev.class_info_id:
                        class_info_id = dev.class_info_id
                except Exception:
                    class_info_id = None
        is_class_time, rule_info = ClassTimeChecker.is_during_class_time()
        period = ClassTimeChecker._current_period(datetime.now())
        data = {
            "is_during_class_time": is_class_time,
            "global_rule": rule_info,
            "period": None,
            "in_session": False,
            # 是否有任意班级此刻在上课——广播类下发的拦截依据，供前端徽章与后端 is_broadcast_blocked 保持一致
            "any_in_session": ClassTimeChecker.any_class_in_session(),
            "class_info_id": class_info_id,
            "class_name": "",
            "subject_name": "",
            "now": datetime.now().isoformat(),
        }
        if period:
            data["period"] = {
                "period_number": period.period_number,
                "name": period.name,
                "start": f"{period.start_hour:02d}:{period.start_minute:02d}",
                "end": f"{period.end_hour:02d}:{period.end_minute:02d}",
            }
        if class_info_id:
            in_session, info = ClassTimeChecker.check_class_in_session(class_info_id)
            data["in_session"] = in_session
            if in_session and info:
                data["class_name"] = info.get("class_name", "")
                data["subject_name"] = info.get("subject_name", "")
        return APIResponse.success(data=data)


@ns_course_schedule.route("/options")
class CourseScheduleOptions(Resource):

    @ns_course_schedule.doc("get_course_schedule_options", description="获取课程表选项")
    @ns_course_schedule.response(200, "成功")
    @requires_permission("view_classes")
    @cached_api(ttl=60)
    def get(self):
        """获取课程表相关选项（班级、科目、节次）"""
        classes = ClassInfo.query.filter_by(is_active=True).order_by(ClassInfo.name).all()
        subjects = Subject.query.filter_by(is_active=True).order_by(Subject.name).all()
        periods = ClassPeriod.query.filter_by(is_active=True).order_by(ClassPeriod.sort_order).all()

        return APIResponse.success(
            data={
                "classes": [
                    {
                        "id": c.id,
                        "name": c.name,
                        "grade": c.grade,
                    }
                    for c in classes
                ],
                "subjects": [
                    {
                        "id": s.id,
                        "name": s.name,
                        "color": s.color,
                        "description": s.description,
                    }
                    for s in subjects
                ],
                "periods": [
                    {
                        "number": p.period_number,
                        "name": p.name,
                        "start_time": f"{p.start_hour:02d}:{p.start_minute:02d}",
                        "end_time": f"{p.end_hour:02d}:{p.end_minute:02d}",
                    }
                    for p in periods
                ],
            }
        )


@ns_course_schedule.route("/check-conflict")
class CourseScheduleConflictCheck(Resource):

    @ns_course_schedule.doc("check_course_schedule_conflict", description="检查课程时间冲突")
    @ns_course_schedule.response(200, "成功")
    @requires_permission("schedule.view")
    def get(self):
        """检查课程时间冲突。支持按班级、教师、教室维度检测。"""
        args = ns_course_schedule.parser.parse_args()
        class_info_id = args.get("class_info_id")
        day_of_week = args.get("day_of_week")
        period_number = args.get("period_number")
        teacher_name = args.get("teacher_name")
        classroom = args.get("classroom")
        exclude_id = args.get("exclude_id")

        conflicts = []

        if class_info_id and day_of_week is not None and period_number:
            conflicts.extend(
                check_conflicts(class_info_id, day_of_week, period_number, exclude_id=exclude_id)
            )

        if teacher_name and day_of_week is not None and period_number:
            conflicts.extend(
                check_teacher_conflicts(
                    teacher_name, day_of_week, period_number, exclude_id=exclude_id
                )
            )

        if classroom and day_of_week is not None and period_number:
            conflicts.extend(
                check_classroom_conflicts(
                    classroom, day_of_week, period_number, exclude_id=exclude_id
                )
            )

        return APIResponse.success(
            data={"has_conflict": len(conflicts) > 0, "conflicts": conflicts}
        )


@ns_course_schedule.route("/export")
class CourseScheduleExport(Resource):

    @ns_course_schedule.doc("export_course_schedules", description="导出课程表数据")
    @requires_permission("schedule.view")
    def get(self):
        """导出课程表数据（支持JSON和Excel格式）"""
        args = ns_course_schedule.parser.parse_args()
        class_info_id = args.get("class_info_id")
        export_format = request.args.get("format", "json").lower()

        query = CourseSchedule.query.filter_by(is_active=True)

        if class_info_id:
            query = query.filter_by(class_info_id=class_info_id)

        schedules = query.order_by(
            CourseSchedule.class_info_id, CourseSchedule.day_of_week, CourseSchedule.period_number
        ).all()

        export_data = []

        for schedule in schedules:
            class_info = get_by_id(ClassInfo, schedule.class_info_id)
            subject = get_by_id(Subject, schedule.subject_id)  # noqa: F841

            export_data.append(
                {
                    "class_info_id": schedule.class_info_id,
                    "class_name": class_info.name if class_info else "",
                    "class_grade": class_info.grade if class_info else "",
                    "subject_id": schedule.subject_id,
                    "subject_name": subject.name if subject else "",
                    "subject_color": subject.color if subject else "",
                    "day_of_week": schedule.day_of_week,
                    "day_of_week_text": format_day_of_week(schedule.day_of_week),
                    "period_number": schedule.period_number,
                    "period_name": get_period_info(schedule.period_number)["name"],
                    "teacher_name": schedule.teacher_name,
                    "classroom": schedule.classroom,
                    "description": schedule.description,
                    "color": schedule.color,
                    "is_active": "是" if schedule.is_active else "否",
                    "created_at": schedule.created_at.isoformat() if schedule.created_at else None,
                    "updated_at": schedule.updated_at.isoformat() if schedule.updated_at else None,
                }
            )

        filename_prefix = f'course_schedules_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        if class_info_id and class_info:
            filename_prefix = (
                f'course_schedule_{class_info.name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
            )

        if export_format == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "课程表数据"

            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # 写入表头
            headers = [
                "班级名称",
                "班级年级",
                "科目名称",
                "星期",
                "节次",
                "节次名称",
                "教师",
                "教室",
                "备注",
                "是否启用",
                "创建时间",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            for row_idx, item in enumerate(export_data, 2):
                ws.cell(row=row_idx, column=1, value=item["class_name"]).border = thin_border
                ws.cell(row=row_idx, column=2, value=item["class_grade"]).border = thin_border
                ws.cell(row=row_idx, column=3, value=item["subject_name"]).border = thin_border
                ws.cell(row=row_idx, column=4, value=item["day_of_week_text"]).border = thin_border
                ws.cell(row=row_idx, column=5, value=item["period_number"]).border = thin_border
                ws.cell(row=row_idx, column=6, value=item["period_name"]).border = thin_border
                ws.cell(row=row_idx, column=7, value=item["teacher_name"]).border = thin_border
                ws.cell(row=row_idx, column=8, value=item["classroom"]).border = thin_border
                ws.cell(row=row_idx, column=9, value=item["description"]).border = thin_border
                ws.cell(row=row_idx, column=10, value=item["is_active"]).border = thin_border
                ws.cell(row=row_idx, column=11, value=item["created_at"]).border = thin_border

            # 调整列宽
            column_widths = [15, 10, 12, 10, 8, 15, 12, 12, 30, 10, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"].width = (
                    width
                )

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            return send_file(
                buf,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"{filename_prefix}.xlsx",
            )
        else:
            output = {
                "export_time": datetime.now().isoformat(),
                "total": len(export_data),
                "class_info_id": class_info_id,
                "data": export_data,
            }

            json_str = json.dumps(output, ensure_ascii=False, indent=2)
            buf = io.BytesIO(json_str.encode("utf-8"))
            buf.seek(0)

            return send_file(
                buf,
                mimetype="application/json",
                as_attachment=True,
                download_name=f"{filename_prefix}.json",
            )


@ns_course_schedule.route("/import")
class CourseScheduleImport(Resource):

    @ns_course_schedule.doc("import_course_schedules", description="导入课程表数据")
    @requires_permission("schedule.manage")
    def post(self):
        """从JSON或Excel文件导入课程表数据（支持配置映射）"""
        content_type = request.content_type or ""
        config_id = request.args.get("config_id", type=int)
        strategy_param = request.args.get("conflict_strategy", type=str)

        config = None
        if config_id:
            config = get_by_id(ImportConfig, config_id)
        else:
            config = ImportConfig.query.filter(
                ImportConfig.module_name == "course_schedule",
                ImportConfig.is_default,
                ImportConfig.is_active,
            ).first()

        default_mappings = [
            {
                "source_field": "班级名称",
                "target_field": "class_name",
                "field_type": "string",
                "required": True,
                "relation": "class_info",
            },
            {
                "source_field": "科目名称",
                "target_field": "subject_name",
                "field_type": "string",
                "required": True,
                "relation": "subject",
            },
            {
                "source_field": "星期",
                "target_field": "day_of_week",
                "field_type": "string",
                "required": True,
            },
            {
                "source_field": "节次",
                "target_field": "period_number",
                "field_type": "integer",
                "required": True,
            },
            {"source_field": "教师", "target_field": "teacher_name", "field_type": "string"},
            {"source_field": "教室", "target_field": "classroom", "field_type": "string"},
            {"source_field": "备注", "target_field": "description", "field_type": "string"},
            {"source_field": "是否启用", "target_field": "is_active", "field_type": "boolean"},
        ]

        field_mappings = config.field_mappings if config else default_mappings
        conflict_strategy = (
            strategy_param
            if strategy_param and strategy_param in ["skip", "update", "error"]
            else (config.conflict_strategy if config else "update")
        )
        default_values = config.default_values if config else {}

        day_text_map = {
            "周一": 0,
            "星期一": 0,
            "周二": 1,
            "星期二": 1,
            "周三": 2,
            "星期三": 2,
            "周四": 3,
            "星期四": 3,
            "周五": 4,
            "星期五": 4,
            "周六": 5,
            "星期六": 5,
            "周日": 6,
            "星期日": 6,
        }

        import_list = []

        if "multipart/form-data" in content_type:
            if "file" not in request.files:
                return APIResponse.bad_request(message="请上传文件")

            file = request.files["file"]
            if not file.filename:
                return APIResponse.bad_request(message="请选择文件")

            filename = file.filename.lower()
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                from openpyxl import load_workbook

                wb = load_workbook(file)
                ws = wb.active

                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)

                col_map = {}
                for idx, header in enumerate(headers):
                    if header:
                        col_map[header] = idx

                for row_idx in range(2, ws.max_row + 1):
                    row_data = {}
                    for header, col_idx in col_map.items():
                        row_data[header] = ws.cell(row=row_idx, column=col_idx + 1).value

                    mapped_item = {}
                    for mapping in field_mappings:
                        source_val = row_data.get(mapping["source_field"])
                        target_field = mapping["target_field"]
                        field_type = mapping.get("field_type", "string")

                        if source_val is None:
                            if mapping.get("required"):
                                break
                            source_val = mapping.get(
                                "default_value", default_values.get(target_field)
                            )

                        if field_type == "boolean":
                            if isinstance(source_val, str):
                                mapped_item[target_field] = source_val in [
                                    "是",
                                    "true",
                                    "True",
                                    "1",
                                ]
                            else:
                                mapped_item[target_field] = bool(source_val)
                        elif field_type == "integer":
                            mapped_item[target_field] = int(source_val) if source_val else None
                        elif target_field == "day_of_week" and isinstance(source_val, str):
                            mapped_item[target_field] = day_text_map.get(source_val, 0)
                        else:
                            mapped_item[target_field] = source_val

                    if mapped_item.get("class_name") and mapped_item.get("subject_name"):
                        import_list.append(mapped_item)
            else:
                return APIResponse.bad_request(message="仅支持 .xlsx 或 .xls 格式")
        elif "application/json" in content_type:
            data = request.json
            if not data or "data" not in data:
                return APIResponse.bad_request(message="导入数据格式错误")
            import_list = data["data"]
        else:
            return APIResponse.bad_request(message="不支持的文件格式")

        success_count = 0
        failed_count = 0
        messages = []
        errors = []
        creates = []
        updates = []

        day_text_map = {
            "周一": 0,
            "星期一": 0,
            "周二": 1,
            "星期二": 1,
            "周三": 2,
            "星期三": 2,
            "周四": 3,
            "星期四": 3,
            "周五": 4,
            "星期五": 4,
            "周六": 5,
            "星期六": 5,
            "周日": 6,
            "星期日": 6,
        }

        max_period = ClassPeriod.query.filter_by(is_active=True).count()

        def validate_item(item, row_idx=None):
            row_errors = []

            class_name = item.get("class_name")
            subject_name = item.get("subject_name")
            day_of_week = item.get("day_of_week")
            period_number = item.get("period_number")
            teacher_name = item.get("teacher_name")
            classroom = item.get("classroom")

            if not class_name:
                row_errors.append({"field": "class_name", "message": "班级名称不能为空"})
            elif not isinstance(class_name, str) or len(class_name.strip()) == 0:
                row_errors.append(
                    {"field": "class_name", "message": "班级名称格式无效，必须为非空字符串"}
                )
            elif len(class_name.strip()) > 100:
                row_errors.append(
                    {"field": "class_name", "message": "班级名称长度超过限制（最大100字符）"}
                )

            if not subject_name:
                row_errors.append({"field": "subject_name", "message": "科目名称不能为空"})
            elif not isinstance(subject_name, str) or len(subject_name.strip()) == 0:
                row_errors.append(
                    {"field": "subject_name", "message": "科目名称格式无效，必须为非空字符串"}
                )
            elif len(subject_name.strip()) > 50:
                row_errors.append(
                    {"field": "subject_name", "message": "科目名称长度超过限制（最大50字符）"}
                )

            if day_of_week is None:
                row_errors.append({"field": "day_of_week", "message": "星期不能为空"})
            elif isinstance(day_of_week, str):
                if day_of_week not in day_text_map:
                    row_errors.append(
                        {
                            "field": "day_of_week",
                            "message": f'星期值 "{day_of_week}" 无效，只能是"周一"到"周日"',
                        }
                    )
            elif not isinstance(day_of_week, int) or day_of_week < 0 or day_of_week > 6:
                row_errors.append(
                    {"field": "day_of_week", "message": "星期值无效，必须为0-6之间的整数"}
                )

            if period_number is None:
                row_errors.append({"field": "period_number", "message": "节次不能为空"})
            elif not isinstance(period_number, int):
                row_errors.append({"field": "period_number", "message": "节次格式无效，必须为整数"})
            elif period_number < 1 or (max_period > 0 and period_number > max_period):
                row_errors.append(
                    {"field": "period_number", "message": f"节次值无效，必须在1-{max_period}之间"}
                )

            if teacher_name:
                if not isinstance(teacher_name, str) or len(teacher_name.strip()) == 0:
                    row_errors.append(
                        {"field": "teacher_name", "message": "教师姓名格式无效，必须为非空字符串"}
                    )
                elif len(teacher_name.strip()) > 50:
                    row_errors.append(
                        {"field": "teacher_name", "message": "教师姓名长度超过限制（最大50字符）"}
                    )
                else:
                    admin = Admin.query.filter(Admin.real_name == teacher_name.strip()).first()
                    if not admin:
                        admin = Admin.query.filter(Admin.username == teacher_name.strip()).first()
                    if admin and admin.role not in ["admin", "teacher"]:
                        row_errors.append(
                            {
                                "field": "teacher_name",
                                "message": f'用户 "{teacher_name}" 的角色不是管理员或教师，无法担任授课教师',
                            }
                        )

            if classroom:
                if not isinstance(classroom, str) or len(classroom.strip()) == 0:
                    row_errors.append(
                        {"field": "classroom", "message": "教室名称格式无效，必须为非空字符串"}
                    )
                elif len(classroom.strip()) > 50:
                    row_errors.append(
                        {"field": "classroom", "message": "教室名称长度超过限制（最大50字符）"}
                    )

            return row_errors

        for row_idx, item in enumerate(import_list, start=2):
            try:
                row_errors = validate_item(item, row_idx)

                class_name = item.get("class_name")
                subject_name = item.get("subject_name")

                if class_name:
                    class_info = ClassInfo.query.filter_by(name=class_name.strip()).first()
                    if not class_info:
                        row_errors.append(
                            {
                                "field": "class_name",
                                "message": f'班级 "{class_name}" 在系统中不存在',
                            }
                        )

                if subject_name:
                    subject = Subject.query.filter_by(
                        name=subject_name.strip()
                    ).first()  # noqa: F841
                    if not subject:
                        row_errors.append(
                            {
                                "field": "subject_name",
                                "message": f'科目 "{subject_name}" 在系统中不存在',
                            }
                        )

                if row_errors:
                    failed_count += 1
                    error_msg = "; ".join(
                        [f'{err["field"]}: {err["message"]}' for err in row_errors]
                    )
                    messages.append(
                        {
                            "class_name": class_name or "未知",
                            "subject_name": subject_name or "未知",
                            "action": "failed",
                            "message": error_msg,
                            "row_data": item,
                            "error_fields": [err["field"] for err in row_errors],
                        }
                    )
                    errors.append(
                        {
                            "row": row_idx,
                            "message": error_msg,
                            "row_data": item,
                            "error_fields": [err["field"] for err in row_errors],
                        }
                    )
                    continue

                class_info = ClassInfo.query.filter_by(name=class_name.strip()).first()
                subject = Subject.query.filter_by(name=subject_name.strip()).first()  # noqa: F841

                day_of_week = item["day_of_week"]
                if isinstance(day_of_week, str):
                    day_of_week = day_text_map.get(day_of_week, 0)

                period_number = item["period_number"]
                teacher_name = item.get("teacher_name")
                classroom = item.get("classroom")

                conflicts = []
                conflicts.extend(check_conflicts(class_info.id, day_of_week, period_number))

                if teacher_name:
                    conflicts.extend(
                        check_teacher_conflicts(teacher_name, day_of_week, period_number)
                    )

                if classroom:
                    conflicts.extend(
                        check_classroom_conflicts(classroom, day_of_week, period_number)
                    )

                if conflicts:
                    conflict_messages = [c["message"] for c in conflicts]
                    failed_count += 1
                    error_msg = "; ".join(conflict_messages)
                    messages.append(
                        {
                            "class_name": class_info.name,
                            "subject_name": subject.name,
                            "action": "failed",
                            "message": error_msg,
                            "row_data": item,
                            "error_fields": ["conflict"],
                        }
                    )
                    errors.append(
                        {
                            "row": row_idx,
                            "message": error_msg,
                            "row_data": item,
                            "error_fields": ["conflict"],
                        }
                    )
                    continue

                existing = CourseSchedule.query.filter(
                    CourseSchedule.class_info_id == class_info.id,
                    CourseSchedule.day_of_week == day_of_week,
                    CourseSchedule.period_number == period_number,
                ).first()

                if existing:
                    if conflict_strategy == "skip":
                        messages.append(
                            {
                                "class_name": class_info.name,
                                "subject_name": subject.name,
                                "action": "skipped",
                                "message": (
                                    f"{class_info.name} {format_day_of_week(day_of_week)}"
                                    f"第{period_number}节课程已存在，已跳过"
                                ),
                            }
                        )
                        continue
                    elif conflict_strategy == "update":
                        updates.append(
                            (
                                existing.id,
                                {
                                    "subject_id": subject.id,
                                    "teacher_name": teacher_name or existing.teacher_name,
                                    "classroom": classroom or existing.classroom,
                                    "description": item.get("description", existing.description),
                                    "color": item.get("color", existing.color),
                                    "is_active": item.get("is_active", existing.is_active),
                                },
                            )
                        )

                        messages.append(
                            {
                                "class_name": class_info.name,
                                "subject_name": subject.name,
                                "action": "updated",
                                "message": f"{class_info.name} {format_day_of_week(day_of_week)}第{period_number}节课程已更新",
                            }
                        )
                    elif conflict_strategy == "error":
                        failed_count += 1
                        error_msg = f"{class_info.name} {format_day_of_week(day_of_week)}第{period_number}节课程已存在，与导入数据冲突"
                        messages.append(
                            {
                                "class_name": class_info.name,
                                "subject_name": subject.name,
                                "action": "failed",
                                "message": error_msg,
                                "row_data": item,
                                "error_fields": ["conflict"],
                            }
                        )
                        errors.append(
                            {
                                "row": row_idx,
                                "message": error_msg,
                                "row_data": item,
                                "error_fields": ["conflict"],
                            }
                        )
                        continue
                else:
                    creates.append(
                        {
                            "class_info_id": class_info.id,
                            "subject_id": subject.id,
                            "day_of_week": day_of_week,
                            "period_number": period_number,
                            "teacher_name": teacher_name,
                            "classroom": classroom,
                            "description": item.get("description"),
                            "color": item.get("color", subject.color),
                            "is_active": item.get("is_active", True),
                        }
                    )

                    messages.append(
                        {
                            "class_name": class_info.name,
                            "subject_name": subject.name,
                            "action": "created",
                            "message": f"{class_info.name} {format_day_of_week(day_of_week)}第{period_number}节课程已创建",
                        }
                    )

                success_count += 1
            except Exception as e:
                failed_count += 1
                error_msg = str(e)
                messages.append(
                    {
                        "class_name": item.get("class_name", "未知"),
                        "subject_name": item.get("subject_name", "未知"),
                        "action": "failed",
                        "message": error_msg,
                        "row_data": item,
                        "error_fields": ["system"],
                    }
                )
                errors.append(
                    {
                        "row": row_idx,
                        "message": error_msg,
                        "row_data": item,
                        "error_fields": ["system"],
                    }
                )

        academics_service.apply_course_schedule_import(creates, updates)
        invalidate_cache("api:/api/course-schedules/*")

        return APIResponse.success(
            data={
                "success": True,
                "total": len(import_list),
                "success_count": success_count,
                "failed_count": failed_count,
                "messages": messages,
            }
        )
