import time
import json
import threading
import io
import openpyxl
from flask_restx import Namespace, Resource, fields
from models import db, Device, DeviceHeartbeat, ClassInfo, Admin, get_by_id
from sqlalchemy.orm import joinedload
from utils.permission import requires_permission, get_current_admin, get_admin_class_ids
from utils.validation import validate_device_id, validate_name
from utils.response import APIResponse
from services.mqtt_service import publish_mqtt
from services.cache_service import cache_service
from services.heartbeat_service import is_device_online
from datetime import datetime
from sqlalchemy import func

from models import DeviceAlert

from flask import request

from flask import send_file
from flask import Response

ns_devices = Namespace("devices", description="设备管理相关操作")


def send_ota_upgrade_command(firmware_url, version="", force=False, device_id=None):
    """
    发送OTA升级指令的公共函数

    Args:
        firmware_url: 固件下载URL
        version: 目标固件版本
        force: 是否强制升级
        device_id: 指定设备ID（None表示广播到所有在线设备）

    Returns:
        APIResponse对象
    """
    if not firmware_url:
        return APIResponse.bad_request(message="需要提供固件下载URL")

    if device_id:
        device = Device.query.filter_by(device_id=device_id).first()
        if not device:
            return APIResponse.bad_request(message="设备不存在")

        ota_payload = {
            "action": "update",
            "url": firmware_url,
            "version": version,
            "force": force,
            "timestamp": int(datetime.now().timestamp()),
        }

        ota_topic = f"phonebox/ota/{device_id}"
        result = publish_mqtt(ota_topic, json.dumps(ota_payload))  # noqa: F841

        if result:
            return APIResponse.success(
                data={
                    "success": True,
                    "message": "OTA升级指令已发送，设备将自动下载并升级",
                    "device_id": device_id,
                    "firmware_url": firmware_url,
                    "version": version,
                    "force": force,
                }
            )
        else:
            return APIResponse.server_error(message="MQTT发送失败，请检查连接")
    else:
        online_devices = Device.query.filter_by(status="online").all()

        if not online_devices:
            return APIResponse.bad_request(message="没有在线设备")

        ota_payload = {
            "action": "update",
            "url": firmware_url,
            "version": version,
            "force": force,
            "timestamp": int(datetime.now().timestamp()),
        }

        ota_topic = "phonebox/ota"
        result = publish_mqtt(ota_topic, json.dumps(ota_payload))  # noqa: F841

        if result:
            return APIResponse.success(
                data={
                    "success": True,
                    "message": f"OTA升级指令已发送到 {len(online_devices)} 个在线设备",
                    "online_count": len(online_devices),
                    "firmware_url": firmware_url,
                    "version": version,
                    "force": force,
                }
            )
        else:
            return APIResponse.server_error(message="MQTT发送失败，请检查连接")


device_model = ns_devices.model(
    "Device",
    {
        "id": fields.Integer(readOnly=True, description="设备ID"),
        "device_id": fields.String(required=True, description="设备标识"),
        "name": fields.String(description="设备名称"),
        "status": fields.String(readOnly=True, description="状态（online/offline/error）"),
        "wifi_signal": fields.Integer(description="WiFi信号强度"),
        "uptime": fields.Integer(description="运行时间（秒）"),
    },
)

device_list_response = ns_devices.model(
    "DeviceListResponse",
    {
        "id": fields.Integer(description="设备ID"),
        "device_id": fields.String(description="设备标识"),
        "name": fields.String(description="设备名称"),
        "status": fields.String(description="状态"),
        "is_online": fields.Boolean(description="是否在线"),
        "last_heartbeat": fields.String(description="最后心跳时间"),
        "wifi_signal": fields.Integer(description="WiFi信号强度"),
        "uptime": fields.Integer(description="运行时间"),
        "box_a_status": fields.String(description="Box A状态"),
        "box_b_status": fields.String(description="Box B状态"),
        "system_state": fields.String(description="系统状态"),
        "class_info_id": fields.Integer(description="班级ID"),
        "class_name": fields.String(description="班级名称"),
        "admin_id": fields.Integer(description="管理员ID"),
        "admin_name": fields.String(description="管理员姓名"),
        "admin_username": fields.String(description="管理员用户名"),
        "created_at": fields.String(description="创建时间"),
        "updated_at": fields.String(description="更新时间"),
    },
)

device_stats_response = ns_devices.model(
    "DeviceStatsResponse",
    {
        "total_devices": fields.Integer(description="设备总数"),
        "online_devices": fields.Integer(description="在线设备数"),
        "offline_devices": fields.Integer(description="离线设备数"),
        "error_devices": fields.Integer(description="故障设备数"),
        "today_heartbeats": fields.Integer(description="今日心跳数"),
        "recent_activity": fields.List(fields.Raw, description="最近活动"),
    },
)


def get_devices_for_admin(admin):
    """根据管理员权限获取设备查询对象"""
    if not admin:
        return Device.query
    if admin.role in ("admin", "super_admin"):
        return Device.query
    class_ids = get_admin_class_ids(admin.id)
    if class_ids:
        return Device.query.filter((Device.class_info_id.in_(class_ids)) | (Device.admin_id == admin.id))

    return Device.query.filter(Device.admin_id == admin.id)


@ns_devices.route("/")
class DeviceList(Resource):

    @ns_devices.doc("list_devices", description="获取设备列表", security="Bearer")
    @ns_devices.param("page", "页码（默认1）")
    @ns_devices.param("per_page", "每页数量（默认20）")
    @ns_devices.param("device_id", "设备标识（模糊搜索）")
    @ns_devices.param("name", "设备名称（模糊搜索）")
    @ns_devices.param("status", "状态（online/offline/error）")
    @ns_devices.param("class_id", "班级ID")
    @ns_devices.response(200, "成功", device_list_response)
    @requires_permission("device.view")
    def get(self):
        """
        获取设备列表

        获取当前管理员有权访问的所有设备列表，支持分页和筛选。
        超级管理员可以看到所有设备，普通管理员只能看到自己班级或绑定到自己的设备。
        """
        admin = get_current_admin()
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 20, type=int)
        device_id = request.args.get("device_id")
        name = request.args.get("name")
        status = request.args.get("status")
        class_id = request.args.get("class_id", type=int)

        query = get_devices_for_admin(admin)

        if device_id:
            query = query.filter(Device.device_id.like(f"%{device_id}%"))
        if name:
            query = query.filter(Device.name.like(f"%{name}%"))
        if status:
            query = query.filter(Device.status == status)
        if class_id:
            query = query.filter(Device.class_info_id == class_id)

        pagination = query.order_by(Device.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
        devices = pagination.items

        return APIResponse.success(
            data={
                "devices": [
                    {
                        "id": d.id,
                        "device_id": d.device_id,
                        "name": d.name,
                        "status": d.status,
                        # is_online 以 last_heartbeat 时效性为准（status 字段可能陈旧，无心跳仍显示 online）
                        "is_online": is_device_online(d),
                        "last_heartbeat": d.last_heartbeat.isoformat() if d.last_heartbeat else None,
                        "wifi_signal": d.wifi_signal,
                        "uptime": d.uptime,
                        "box_a_status": d.box_a_status,
                        "box_b_status": d.box_b_status,
                        "system_state": d.system_state,
                        "class_info_id": d.class_info_id,
                        "class_name": d.class_info.name if d.class_info else None,
                        "admin_id": d.admin_id,
                        "admin_name": d.admin.real_name if d.admin else None,
                        "admin_username": d.admin.username if d.admin else None,
                        "created_at": d.created_at.isoformat() if d.created_at else None,
                        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
                    }
                    for d in devices
                ],
                "total": pagination.total,
                "page": page,
                "per_page": per_page,
                "pages": pagination.pages,
            }
        )

    @ns_devices.doc("create_device", description="创建设备", security="Bearer")
    @ns_devices.expect(device_model)
    @ns_devices.response(201, "创建成功")
    @requires_permission("device.edit")
    def post(self):
        """
        创建设备

        创建新的设备，需要管理员权限。

        请求体：
        - device_id: 设备标识（必填）
        - name: 设备名称（可选，默认"设备 {device_id}"）
        """
        data = ns_devices.payload
        device = Device(device_id=data.get("device_id"), name=data.get("name", f'设备 {data.get("device_id")}'))
        db.session.add(device)
        db.session.commit()
        return APIResponse.created(data={"device_id": device.id}, message="设备创建成功")


@ns_devices.route("/<int:id>")
@ns_devices.param("id", "设备ID")
class DeviceResource(Resource):

    @ns_devices.doc("get_device", description="获取单个设备详情")
    @ns_devices.response(200, "成功")
    @ns_devices.response(404, "设备不存在")
    @requires_permission("device.view")
    def get(self, id):
        """
        获取单个设备详情

        根据设备ID获取设备的详细信息。
        """
        device = Device.query.get_or_404(id)
        return APIResponse.success(
            data={
                "id": device.id,
                "device_id": device.device_id,
                "name": device.name,
                "status": device.status,
                # is_online 以 last_heartbeat 时效性为准（避免无心跳却显示在线）
                "is_online": is_device_online(device),
                "last_heartbeat": device.last_heartbeat.isoformat() if device.last_heartbeat else None,
                "wifi_signal": device.wifi_signal,
                "uptime": device.uptime,
                "box_a_status": device.box_a_status,
                "box_b_status": device.box_b_status,
                "system_state": device.system_state,
                "class_info_id": device.class_info_id,
                "class_name": device.class_info.name if device.class_info else None,
                "admin_id": device.admin_id,
                "admin_name": device.admin.real_name if device.admin else None,
                "admin_username": device.admin.username if device.admin else None,
                "created_at": device.created_at.isoformat() if device.created_at else None,
                "updated_at": device.updated_at.isoformat() if device.updated_at else None,
            }
        )

    @ns_devices.doc("update_device", description="更新设备", security="Bearer")
    @ns_devices.expect(device_model)
    @ns_devices.response(200, "更新成功")
    @ns_devices.response(404, "设备不存在")
    @requires_permission("device.edit")
    def put(self, id):
        """
        更新设备

        更新指定设备的信息，需要管理员权限。
        """
        device = Device.query.get_or_404(id)
        data = ns_devices.payload
        device.name = data.get("name", device.name)
        device.updated_at = datetime.now()
        db.session.commit()
        return APIResponse.success(message="设备更新成功")

    @ns_devices.doc("delete_device", description="删除设备", security="Bearer")
    @ns_devices.response(200, "删除成功")
    @ns_devices.response(404, "设备不存在")
    @requires_permission("device.delete")
    def delete(self, id):
        """
        删除设备

        删除指定的设备，需要管理员权限。
        """
        device = Device.query.get_or_404(id)
        db.session.delete(device)
        db.session.commit()
        return APIResponse.success(message="设备删除成功")


@ns_devices.route("/<int:id>/heartbeats")
@ns_devices.param("id", "设备ID")
class DeviceHeartbeats(Resource):
    @ns_devices.doc(
        "get_device_heartbeats",
        description="获取设备心跳记录",
        params={"page": "页码（默认1）", "per_page": "每页数量（默认50）"},
    )
    @ns_devices.response(200, "成功")
    @ns_devices.response(404, "设备不存在")
    @requires_permission("device.view")
    def get(self, id):
        """
        获取设备心跳记录

        获取指定设备的所有心跳历史记录，支持分页。
        """
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 50, type=int)

        device = Device.query.get_or_404(id)
        pagination = (
            DeviceHeartbeat.query.filter_by(device_id=device.device_id)
            .order_by(DeviceHeartbeat.received_at.desc())
            .paginate(page=page, per_page=per_page, error_out=False)
        )

        return APIResponse.success(
            data={
                "heartbeats": [
                    {
                        "id": h.id,
                        "timestamp": h.timestamp,
                        "status": h.status,
                        "wifi_signal": h.wifi_signal,
                        "uptime": h.uptime,
                        "box_a_status": h.box_a_status,
                        "box_b_status": h.box_b_status,
                        "system_state": h.system_state,
                        "received_at": h.received_at.isoformat() if h.received_at else None,
                    }
                    for h in pagination.items
                ],
                "total": pagination.total,
                "page": page,
                "per_page": per_page,
                "pages": pagination.pages,
            }
        )


@ns_devices.route("/device/<string:device_id>/heartbeats")
@ns_devices.param("device_id", "设备标识")
class DeviceHeartbeatsByDeviceId(Resource):
    @ns_devices.doc(
        "get_device_heartbeats_by_device_id",
        description="通过设备标识获取心跳记录",
        params={"page": "页码（默认1）", "per_page": "每页数量（默认50）"},
    )
    @ns_devices.response(200, "成功")
    @ns_devices.response(404, "设备不存在")
    @requires_permission("device.view")
    def get(self, device_id):
        """
        通过设备标识获取心跳记录

        使用设备标识（如 phonebox_001）获取心跳历史记录，支持分页。
        """
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 50, type=int)

        device = Device.query.filter_by(device_id=device_id).first_or_404()
        pagination = (
            DeviceHeartbeat.query.filter_by(device_id=device.device_id)
            .order_by(DeviceHeartbeat.received_at.desc())
            .paginate(page=page, per_page=per_page, error_out=False)
        )

        return APIResponse.success(
            data={
                "heartbeats": [
                    {
                        "id": h.id,
                        "timestamp": h.timestamp,
                        "status": h.status,
                        "wifi_signal": h.wifi_signal,
                        "uptime": h.uptime,
                        "box_a_status": h.box_a_status,
                        "box_b_status": h.box_b_status,
                        "system_state": h.system_state,
                        "received_at": h.received_at.isoformat() if h.received_at else None,
                    }
                    for h in pagination.items
                ],
                "total": pagination.total,
                "page": page,
                "per_page": per_page,
                "pages": pagination.pages,
            }
        )


@ns_devices.route("/stats")
class DeviceStats(Resource):

    @ns_devices.doc("get_device_stats", description="获取设备统计信息")
    @ns_devices.response(200, "成功", device_stats_response)
    @requires_permission("device.view")
    def get(self):
        """
        获取设备统计信息

        获取所有设备的统计数据，包括在线/离线数量、今日心跳数等。
        """
        from services.cache_service import cache_service

        cache_key = "device_stats"
        cached = cache_service.get(cache_key)
        if cached:
            return APIResponse.success(data=cached)

        total = Device.query.count()
        online = Device.query.filter_by(status="online").count()
        offline = Device.query.filter_by(status="offline").count()
        error = Device.query.filter_by(status="error").count()

        today = datetime.now().date()
        today_heartbeats = DeviceHeartbeat.query.filter(
            DeviceHeartbeat.received_at >= datetime.combine(today, datetime.min.time())
        ).count()

        recent_heartbeats = DeviceHeartbeat.query.order_by(DeviceHeartbeat.received_at.desc()).limit(100).all()

        result = {  # noqa: F841
            "total_devices": total,
            "online_devices": online,
            "offline_devices": offline,
            "error_devices": error,
            "today_heartbeats": today_heartbeats,
            "recent_activity": [
                {
                    "device_id": h.device_id,
                    "status": h.status,
                    "received_at": h.received_at.isoformat() if h.received_at else None,
                }
                for h in recent_heartbeats[:10]
            ],
        }

        cache_service.set(cache_key, result, ttl=300)

        return APIResponse.success(data=result)


@ns_devices.route("/online")
class OnlineDevices(Resource):

    @ns_devices.doc("get_online_devices", description="获取在线设备列表")
    @ns_devices.response(200, "成功")
    @requires_permission("device.view")
    def get(self):
        """
        获取在线设备列表

        获取所有当前在线的设备列表。
        """
        devices = (
            Device.query.filter_by(status="online")
            .options(joinedload(Device.class_info), joinedload(Device.admin))
            .all()
        )
        return [
            {
                "id": d.id,
                "device_id": d.device_id,
                "name": d.name,
                "status": d.status,
                "is_online": True,
                "last_heartbeat": d.last_heartbeat.isoformat() if d.last_heartbeat else None,
                "wifi_signal": d.wifi_signal,
                "class_info_id": d.class_info_id,
                "class_name": d.class_info.name if d.class_info else None,
                "admin_id": d.admin_id,
                "admin_name": d.admin.real_name if d.admin else None,
            }
            for d in devices
        ]


bind_class_model = ns_devices.model(
    "BindClassRequest", {"class_id": fields.Integer(description="班级ID（设为null可解绑）")}
)


@ns_devices.route("/<int:id>/bind-class")
@ns_devices.param("id", "设备ID")
class BindDeviceClass(Resource):

    @ns_devices.doc("bind_device_class", description="绑定设备到班级", security="Bearer")
    @ns_devices.expect(bind_class_model)
    @ns_devices.response(200, "绑定成功")
    @ns_devices.response(403, "无权绑定到该班级")
    @ns_devices.response(404, "班级不存在")
    @requires_permission("device.edit")
    def post(self, id):
        """
        绑定设备到班级

        将设备绑定到指定的班级，需要设备编辑权限。
        非管理员只能绑定到自己管理的班级。

        请求体：
        - class_id: 班级ID（设为null可解绑设备与班级的关联）
        """
        admin = get_current_admin()
        device = Device.query.get_or_404(id)
        data = request.get_json()
        class_id = data.get("class_id")

        if admin.role not in ("admin", "super_admin"):
            class_ids = get_admin_class_ids(admin.id)
            if class_id and class_id not in class_ids:
                return APIResponse.forbidden(message="无权绑定到该班级")

        if class_id:
            class_info = get_by_id(ClassInfo, class_id)
            if not class_info:
                return APIResponse.not_found(message="班级不存在")
            device.class_info_id = class_id
        else:
            device.class_info_id = None

        device.updated_at = datetime.now()
        db.session.commit()

        return APIResponse.success(
            data={
                "class_info_id": device.class_info_id,
                "class_name": device.class_info.name if device.class_info else None,
            },
            message="设备绑定班级成功",
        )


bind_admin_model = ns_devices.model(
    "BindAdminRequest", {"admin_id": fields.Integer(description="管理员ID（设为null可解绑）")}
)


@ns_devices.route("/<int:id>/bind-admin")
@ns_devices.param("id", "设备ID")
class BindDeviceAdmin(Resource):

    @ns_devices.doc("bind_device_admin", description="绑定设备到管理员", security="Bearer")
    @ns_devices.expect(bind_admin_model)
    @ns_devices.response(200, "绑定成功")
    @ns_devices.response(403, "只有超级管理员可以绑定管理员")
    @ns_devices.response(404, "管理员不存在")
    @requires_permission("device.edit")
    def post(self, id):
        """
        绑定设备到管理员

        将设备绑定到指定的管理员，只有管理员可以执行此操作。

        请求体：
        - admin_id: 管理员ID（设为null可解绑设备与管理员的关联）
        """
        admin = get_current_admin()

        if admin.role not in ("admin", "super_admin"):
            return APIResponse.forbidden(message="只有管理员可以绑定管理员")

        device = Device.query.get_or_404(id)
        data = request.get_json()
        admin_id = data.get("admin_id")

        if admin_id:
            target_admin = get_by_id(Admin, admin_id)
            if not target_admin:
                return APIResponse.not_found(message="管理员不存在")
            device.admin_id = admin_id
        else:
            device.admin_id = None

        device.updated_at = datetime.now()
        db.session.commit()

        return APIResponse.success(
            data={
                "admin_id": device.admin_id,
                "admin_name": device.admin.real_name if device.admin else None,
                "admin_username": device.admin.username if device.admin else None,
            },
            message="设备绑定管理员成功",
        )


@ns_devices.route("/class/<int:class_id>")
@ns_devices.param("class_id", "班级ID")
class DevicesByClass(Resource):

    @ns_devices.doc("get_devices_by_class", description="获取班级的设备列表")
    @ns_devices.response(200, "成功")
    @requires_permission("device.view")
    def get(self, class_id):
        """
        获取班级的设备列表

        获取绑定到指定班级的所有设备。
        """
        devices = Device.query.filter_by(class_info_id=class_id).options(joinedload(Device.admin)).all()
        return [
            {
                "id": d.id,
                "device_id": d.device_id,
                "name": d.name,
                "status": d.status,
                "is_online": d.status == "online",
                "last_heartbeat": d.last_heartbeat.isoformat() if d.last_heartbeat else None,
                "wifi_signal": d.wifi_signal,
                "admin_name": d.admin.real_name if d.admin else None,
                "updated_at": d.updated_at.isoformat() if d.updated_at else None,
            }
            for d in devices
        ]


@ns_devices.route("/admin/<int:admin_id>")
@ns_devices.param("admin_id", "管理员ID")
class DevicesByAdmin(Resource):

    @ns_devices.doc("get_devices_by_admin", description="获取管理员的设备列表")
    @ns_devices.response(200, "成功")
    @requires_permission("device.view")
    def get(self, admin_id):
        """
        获取管理员的设备列表

        获取绑定到指定管理员的所有设备。
        """
        devices = Device.query.filter_by(admin_id=admin_id).options(joinedload(Device.class_info)).all()
        return [
            {
                "id": d.id,
                "device_id": d.device_id,
                "name": d.name,
                "status": d.status,
                "is_online": d.status == "online",
                "last_heartbeat": d.last_heartbeat.isoformat() if d.last_heartbeat else None,
                "wifi_signal": d.wifi_signal,
                "class_name": d.class_info.name if d.class_info else None,
                "updated_at": d.updated_at.isoformat() if d.updated_at else None,
            }
            for d in devices
        ]


@ns_devices.route("/alerts")
class DeviceAlerts(Resource):

    @ns_devices.doc("get_device_alerts", description="获取设备告警列表")
    @ns_devices.param("resolved", "是否已解决（true/false）")
    @ns_devices.param("severity", "告警级别（info/warning/error/critical）")
    @ns_devices.param("page", "页码（默认1）")
    @ns_devices.param("per_page", "每页数量（默认50）")
    @ns_devices.response(200, "成功")
    @requires_permission("device.view")
    def get(self):
        """
        获取设备告警列表

        获取所有设备的告警记录，支持按状态和级别筛选，支持分页。
        """
        resolved = request.args.get("resolved", "false").lower() == "true"
        severity = request.args.get("severity")
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 50, type=int)

        query = DeviceAlert.query
        if resolved:
            query = query.filter(DeviceAlert.is_resolved)
        else:
            query = query.filter(DeviceAlert.is_resolved)

        if severity:
            query = query.filter(DeviceAlert.severity == severity)

        pagination = query.order_by(DeviceAlert.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        alerts = pagination.items

        device_ids = {a.device_id for a in alerts}
        devices = Device.query.filter(Device.device_id.in_(device_ids)).all()
        device_map = {d.device_id: d.name for d in devices}

        return APIResponse.success(
            data={
                "alerts": [
                    {
                        "id": a.id,
                        "device_id": a.device_id,
                        "device_name": device_map.get(a.device_id),
                        "alert_type": a.alert_type,
                        "severity": a.severity,
                        "message": a.message,
                        "is_resolved": a.is_resolved,
                        "resolved_at": a.resolved_at.isoformat() if a.resolved_at else None,
                        "created_at": a.created_at.isoformat() if a.created_at else None,
                    }
                    for a in alerts
                ],
                "total": pagination.total,
                "page": page,
                "per_page": per_page,
                "pages": pagination.pages,
                "unresolved_count": DeviceAlert.query.filter_by(is_resolved=False).count(),
            }
        )


@ns_devices.route("/<int:id>/resolve-alert/<int:alert_id>")
@ns_devices.param("id", "设备ID")
@ns_devices.param("alert_id", "告警ID")
class ResolveDeviceAlert(Resource):

    @ns_devices.doc("resolve_device_alert", description="解决设备告警", security="Bearer")
    @ns_devices.response(200, "成功")
    @requires_permission("device.edit")
    def post(self, id, alert_id):
        """
        解决设备告警

        将指定告警标记为已解决。
        """
        alert = DeviceAlert.query.get_or_404(alert_id)
        alert.is_resolved = True
        alert.resolved_at = datetime.now()
        db.session.commit()
        return APIResponse.success(message="告警已解决")


@ns_devices.route("/<int:id>/alerts/<int:alert_id>/resolve")
@ns_devices.param("id", "设备ID")
@ns_devices.param("alert_id", "告警ID")
class ResolveDeviceAlertAlt(Resource):

    @ns_devices.doc("resolve_device_alert_alt", description="解决设备告警（备用路径）", security="Bearer")
    @ns_devices.response(200, "成功")
    @requires_permission("device.edit")
    def post(self, id, alert_id):
        """
        解决设备告警（备用路径）

        将指定告警标记为已解决。
        """
        alert = DeviceAlert.query.get_or_404(alert_id)
        alert.is_resolved = True
        alert.resolved_at = datetime.now()
        db.session.commit()
        return APIResponse.success(message="告警已解决")


@ns_devices.route("/<int:id>/alerts")
@ns_devices.param("id", "设备ID")
class DeviceAlertHistory(Resource):

    @ns_devices.doc("get_device_alert_history", description="获取设备告警历史")
    @ns_devices.response(200, "成功")
    @requires_permission("device.view")
    def get(self, id):
        """
        获取设备的告警历史记录

        获取指定设备的所有告警记录。
        """
        device = Device.query.get_or_404(id)
        alerts = (
            DeviceAlert.query.filter_by(device_id=device.device_id)
            .order_by(DeviceAlert.created_at.desc())
            .limit(50)
            .all()
        )

        return APIResponse.success(
            data={
                "alerts": [
                    {
                        "id": a.id,
                        "alert_type": a.alert_type,
                        "severity": a.severity,
                        "message": a.message,
                        "is_resolved": a.is_resolved,
                        "resolved_at": a.resolved_at.isoformat() if a.resolved_at else None,
                        "created_at": a.created_at.isoformat() if a.created_at else None,
                    }
                    for a in alerts
                ],
                "total": len(alerts),
                "unresolved_count": DeviceAlert.query.filter_by(device_id=device.device_id, is_resolved=False).count(),
            }
        )


@ns_devices.route("/<int:id>/remote-control")
@ns_devices.param("id", "设备ID")
class DeviceRemoteControl(Resource):
    @ns_devices.doc("device_remote_control", description="设备远程控制", security="Bearer")
    @ns_devices.expect(
        ns_devices.model(
            "RemoteControl",
            {"action": fields.String(required=True, description="操作类型：restart/reboot/unlock_a/unlock_b")},
        )
    )
    @ns_devices.response(200, "成功")
    @ns_devices.response(400, "设备不在线")
    @requires_permission("device.edit")
    def post(self, id):
        """
        设备远程控制

        对指定设备执行远程操作，包括重启和远程开锁。
        需要设备在线才能执行操作。

        操作类型：
        - restart: 重启设备
        - unlock_a: 打开A箱（班主任远程开锁，无需验证）
        - unlock_b: 打开B箱（需要验证积分）
        """
        device = Device.query.get_or_404(id)
        data = request.get_json()
        action = data.get("action")

        if not action:
            return APIResponse.bad_request(message="需要提供操作类型")

        if action in ["restart", "unlock_a", "unlock_b"]:
            if device.status != "online":
                return APIResponse.bad_request(message="设备不在线，无法执行远程操作")

        if action == "restart":
            restart_topic = "phonebox/control/restart"
            result = publish_mqtt(restart_topic, '{"command": "restart"}')  # noqa: F841
            if result:
                return APIResponse.success(
                    message="重启指令已发送", data={"action": action, "device_id": device.device_id}
                )
            else:
                return APIResponse.server_error(message="MQTT发送失败，请检查连接")

        elif action == "unlock_a":
            # 智能开锁：增加重试机制，确保指令能被设备接收
            # ESP32设备只在IDLE状态时响应A箱开锁指令
            # 通过多次发送指令，覆盖设备状态转换的时间窗口

            def smart_unlock_a():
                unlock_topic = "phonebox/unlock/A"

                # 多次发送开锁指令，确保设备在可接收状态时能收到
                for attempt in range(3):
                    publish_mqtt(unlock_topic, "")
                    if attempt < 2:
                        time.sleep(0.5)  # 等待500ms后重试

            # 在后台线程执行智能开锁流程，避免阻塞响应
            thread = threading.Thread(target=smart_unlock_a)
            thread.daemon = True
            thread.start()

            return APIResponse.success(
                message="A箱智能开锁指令已发送（后台执行，共发送3次）",
                data={"action": action, "device_id": device.device_id},
            )

        elif action == "unlock_b":
            # 智能开锁：增加重试机制，确保指令能被设备接收
            # ESP32设备只在IDLE或SHOWING_CARD状态时响应开锁指令
            # 通过多次发送指令，覆盖设备状态转换的时间窗口

            def smart_unlock_b():
                unlock_topic = "phonebox/unlock/B"
                payload = '{"result": "true", "reason": "manual", "current_score": 999}'

                # 多次发送开锁指令，确保设备在可接收状态时能收到
                # 发送间隔：500ms，共发送3次
                for attempt in range(3):
                    publish_mqtt(unlock_topic, payload)
                    if attempt < 2:
                        time.sleep(0.5)  # 等待500ms后重试

            # 在后台线程执行智能开锁流程，避免阻塞响应
            thread = threading.Thread(target=smart_unlock_b)
            thread.daemon = True
            thread.start()

            return APIResponse.success(
                message="B箱智能开锁指令已发送（后台执行，共发送3次）",
                data={"action": action, "device_id": device.device_id},
            )

        else:
            return APIResponse.bad_request(message=f"不支持的操作类型: {action}")


@ns_devices.route("/advanced-stats")
class DeviceAdvancedStats(Resource):

    @ns_devices.doc("get_device_advanced_stats", description="获取设备高级统计")
    @ns_devices.response(200, "成功")
    @requires_permission("device.view")
    def get(self):
        """
        获取设备高级统计信息

        获取更详细的设备统计，包括信号强度分布、在线时长等。
        """

        cache_key = "device_advanced_stats"
        cached = cache_service.get(cache_key)
        if cached:
            return APIResponse.success(data=cached)

        total = Device.query.count()
        online = Device.query.filter_by(status="online").count()
        offline = Device.query.filter_by(status="offline").count()
        error = Device.query.filter_by(status="error").count()

        avg_signal = db.session.query(func.avg(Device.wifi_signal)).filter(Device.wifi_signal.isnot(None)).scalar() or 0

        alert_count = DeviceAlert.query.filter_by(is_resolved=False).count()
        critical_alerts = DeviceAlert.query.filter_by(is_resolved=False, severity="critical").count()

        today = datetime.now().date()
        today_start = datetime.combine(today, datetime.min.time())
        today_heartbeats = DeviceHeartbeat.query.filter(DeviceHeartbeat.received_at >= today_start).count()

        devices_with_signal = Device.query.filter(Device.wifi_signal.isnot(None)).all()

        signal_distribution = {"excellent": 0, "good": 0, "fair": 0, "poor": 0}

        for d in devices_with_signal:
            if d.wifi_signal >= -50:
                signal_distribution["excellent"] += 1
            elif d.wifi_signal >= -70:
                signal_distribution["good"] += 1
            elif d.wifi_signal >= -80:
                signal_distribution["fair"] += 1
            else:
                signal_distribution["poor"] += 1

        result = {  # noqa: F841
            "total_devices": total,
            "online_devices": online,
            "offline_devices": offline,
            "error_devices": error,
            "online_rate": round(online / total * 100, 1) if total > 0 else 0,
            "avg_signal_strength": round(avg_signal, 1),
            "signal_distribution": signal_distribution,
            "today_heartbeats": today_heartbeats,
            "unresolved_alerts": alert_count,
            "critical_alerts": critical_alerts,
        }

        cache_service.set(cache_key, result, ttl=300)

        return result


@ns_devices.route("/heartbeat-timeout-check")
class HeartbeatTimeoutCheck(Resource):

    @ns_devices.doc("check_heartbeat_timeout", description="检查心跳超时设备")
    @ns_devices.response(200, "成功")
    @requires_permission("device.view")
    def get(self):
        """
        检查心跳超时的设备

        遍历所有设备，检查是否有设备超过心跳间隔未响应。
        返回超时的设备列表，并自动创建告警。
        """
        from services.heartbeat_service import check_heartbeat_timeout

        result = check_heartbeat_timeout()  # noqa: F841
        return APIResponse.success(data=result)


@ns_devices.route("/<int:id>/settings")
@ns_devices.param("id", "设备ID")
class DeviceSettings(Resource):
    @ns_devices.doc("update_device_settings", description="更新设备设置", security="Bearer")
    @ns_devices.expect(
        ns_devices.model(
            "DeviceSettings",
            {
                "alert_enabled": fields.Boolean(description="是否启用告警"),
                "heartbeat_timeout": fields.Integer(description="心跳超时时间（秒）"),
                "name": fields.String(description="设备名称"),
            },
        )
    )
    @ns_devices.response(200, "成功")
    @requires_permission("device.edit")
    def put(self, id):
        """
        更新设备设置

        更新指定设备的配置选项。
        """
        device = Device.query.get_or_404(id)
        data = request.get_json()

        if "alert_enabled" in data:
            device.alert_enabled = data["alert_enabled"]
        if "heartbeat_timeout" in data:
            device.heartbeat_timeout = data["heartbeat_timeout"]
        if "name" in data:
            device.name = data["name"]

        device.updated_at = datetime.now()
        db.session.commit()

        return APIResponse.success(
            data={
                "success": True,
                "message": "设备设置已更新",
                "settings": {
                    "alert_enabled": device.alert_enabled,
                    "heartbeat_timeout": device.heartbeat_timeout,
                    "name": device.name,
                },
            }
        )


@ns_devices.route("/batch-control")
class BatchDeviceControl(Resource):
    @ns_devices.doc("batch_device_control", description="批量设备控制", security="Bearer")
    @ns_devices.expect(
        ns_devices.model(
            "BatchControl",
            {
                "device_ids": fields.List(fields.Integer, required=True, description="设备ID列表"),
                "action": fields.String(required=True, description="操作类型：restart/unlock"),
            },
        )
    )
    @ns_devices.response(200, "成功")
    @ns_devices.response(400, "设备不在线")
    @requires_permission("device.edit")
    def post(self):
        """
        批量设备控制

        对多个设备同时执行远程操作。
        只对在线设备执行操作。

        操作类型：
        - restart: 重启设备
        - unlock: 打开所有箱门
        """
        data = request.get_json()
        device_ids = data.get("device_ids", [])
        action = data.get("action")

        if not device_ids:
            return APIResponse.success(data={"total": 0, "online_count": 0, "offline_count": 0, "results": []})

        devices = Device.query.filter(Device.id.in_(device_ids)).all()
        device_map = {d.id: d for d in devices}

        results = []
        for device_id in device_ids:
            device = device_map.get(device_id)
            if not device:
                results.append({"device_id": device_id, "success": False, "message": "设备不存在"})
                continue

            if device.status == "online":
                if action == "restart":
                    restart_topic = "phonebox/control/restart"
                    result = publish_mqtt(restart_topic, '{"command": "restart"}')  # noqa: F841
                elif action == "unlock":
                    unlock_topic_a = "phonebox/unlock/A"
                    publish_mqtt(unlock_topic_a, "")
                    unlock_topic_b = "phonebox/unlock/B"
                    result = publish_mqtt(
                        unlock_topic_b, '{"result": "true", "reason": "manual", "current_score": 0}'
                    )  # noqa: F841

                if result:
                    results.append(
                        {
                            "device_id": device_id,
                            "device_name": device.name,
                            "success": True,
                            "message": f"指令已发送: {action}",
                        }
                    )
                else:
                    results.append(
                        {
                            "device_id": device_id,
                            "device_name": device.name,
                            "success": False,
                            "message": "MQTT发送失败",
                        }
                    )
            else:
                results.append(
                    {"device_id": device_id, "device_name": device.name, "success": False, "message": "设备不在线"}
                )

        return APIResponse.success(
            data={
                "success": True,
                "total": len(device_ids),
                "online_count": sum(1 for r in results if r["success"]),
                "offline_count": sum(1 for r in results if not r["success"]),
                "results": results,
            }
        )


ota_upgrade_model = ns_devices.model(
    "OTAUpgrade",
    {
        "firmware_url": fields.String(required=True, description="固件下载URL"),
        "version": fields.String(description="目标固件版本"),
        "force": fields.Boolean(description="是否强制升级（忽略版本检查）", default=False),
    },
)


@ns_devices.route("/<int:id>/ota-upgrade")
@ns_devices.param("id", "设备ID")
class DeviceOTAUpgrade(Resource):

    @ns_devices.doc("device_ota_upgrade", description="设备OTA固件升级", security="Bearer")
    @ns_devices.expect(ota_upgrade_model)
    @ns_devices.response(200, "成功")
    @ns_devices.response(400, "设备不在线")
    @requires_permission("device.edit")
    def post(self, id):
        """
        设备OTA固件升级

        向指定设备发送OTA固件升级指令。
        需要设备在线才能执行升级。

        请求体：
        - firmware_url: 固件下载URL（必填）
        - version: 目标固件版本（可选）
        - force: 是否强制升级，忽略版本检查（可选，默认false）
        """
        device = Device.query.get_or_404(id)

        if device.status != "online":
            return APIResponse.bad_request(message="设备不在线，无法执行OTA升级")

        data = request.get_json()
        firmware_url = data.get("firmware_url")
        version = data.get("version", "")
        force = data.get("force", False)

        return send_ota_upgrade_command(firmware_url, version, force, device.device_id)


@ns_devices.route("/ota-upgrade-all")
class DeviceOTAUpgradeAll(Resource):

    @ns_devices.doc("device_ota_upgrade_all", description="批量OTA固件升级", security="Bearer")
    @ns_devices.expect(ota_upgrade_model)
    @ns_devices.response(200, "成功")
    @requires_permission("device.edit")
    def post(self):
        """
        批量OTA固件升级

        向所有在线设备发送OTA固件升级指令。

        请求体：
        - firmware_url: 固件下载URL（必填）
        - version: 目标固件版本（可选）
        - force: 是否强制升级，忽略版本检查（可选，默认false）
        """
        data = request.get_json()
        firmware_url = data.get("firmware_url")
        version = data.get("version", "")
        force = data.get("force", False)

        return send_ota_upgrade_command(firmware_url, version, force)


@ns_devices.route("/bulk-ota-upgrade")
class DeviceBulkOTAUpgrade(Resource):

    @ns_devices.doc("device_bulk_ota_upgrade", description="批量OTA固件升级（别名）", security="Bearer")
    @ns_devices.expect(ota_upgrade_model)
    @ns_devices.response(200, "成功")
    @requires_permission("device.edit")
    def post(self):
        """
        批量OTA固件升级（别名接口）

        向所有在线设备发送OTA固件升级指令。
        """
        data = request.get_json()
        firmware_url = data.get("firmware_url")
        version = data.get("version", "")
        force = data.get("force", False)

        return send_ota_upgrade_command(firmware_url, version, force)


@ns_devices.route("/import")
class DeviceImport(Resource):

    @ns_devices.doc("device_import", description="批量导入设备", security="Bearer")
    @requires_permission("device.edit")
    def post(self):
        """
        批量导入设备

        通过Excel文件批量导入设备信息。

        支持的字段：设备标识(device_id)、设备名称(name)、班级名称(class_name)、管理员姓名(admin_name)

        返回导入结果统计。
        """
        try:
            if "file" not in request.files:
                return APIResponse.bad_request(message="没有上传文件")

            file = request.files["file"]

            if not file.filename.endswith((".xlsx", ".xls")):
                return APIResponse.bad_request(message="仅支持Excel文件格式")

            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            success_count = 0
            failed_count = 0
            messages = []

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    row_dict = dict(zip(headers, row))

                    device_id = row_dict.get("设备标识") or row_dict.get("device_id") or row_dict.get("设备ID")
                    name = row_dict.get("设备名称") or row_dict.get("name")
                    class_name = row_dict.get("班级名称") or row_dict.get("class_name")
                    admin_name = row_dict.get("管理员姓名") or row_dict.get("admin_name")

                    row_errors = []

                    if not device_id:
                        row_errors.append({"field": "device_id", "message": "设备标识不能为空"})
                    elif not isinstance(device_id, (int, str)) or len(str(device_id).strip()) == 0:
                        row_errors.append({"field": "device_id", "message": "设备标识格式无效"})
                    elif len(str(device_id).strip()) > 100:
                        row_errors.append({"field": "device_id", "message": "设备标识长度超过限制（最大100字符）"})
                    else:
                        device_id_str = str(device_id).strip()
                        is_valid, msg = validate_device_id(device_id_str)
                        if not is_valid:
                            row_errors.append({"field": "device_id", "message": msg})

                    if name and (not isinstance(name, str) or len(name.strip()) > 200):
                        row_errors.append({"field": "name", "message": "设备名称长度超过限制（最大200字符）"})
                    elif name:
                        is_valid, msg = validate_name(name.strip())
                        if not is_valid:
                            row_errors.append({"field": "name", "message": msg})

                    existing_device = Device.query.filter_by(device_id=str(device_id)).first()
                    if existing_device:
                        row_errors.append({"field": "device_id", "message": f'设备 "{str(device_id)}" 已存在'})

                    class_info = None
                    if class_name:
                        if not isinstance(class_name, str) or len(class_name.strip()) == 0:
                            row_errors.append({"field": "class_name", "message": "班级名称格式无效，必须为非空字符串"})
                        elif len(class_name.strip()) > 100:
                            row_errors.append({"field": "class_name", "message": "班级名称长度超过限制（最大100字符）"})
                        else:
                            class_info = ClassInfo.query.filter_by(name=class_name.strip()).first()
                            if not class_info:
                                row_errors.append(
                                    {"field": "class_name", "message": f'班级 "{class_name}" 在系统中不存在'}
                                )

                    admin = None
                    if admin_name:
                        if not isinstance(admin_name, str) or len(admin_name.strip()) == 0:
                            row_errors.append(
                                {"field": "admin_name", "message": "管理员姓名格式无效，必须为非空字符串"}
                            )
                        elif len(admin_name.strip()) > 50:
                            row_errors.append(
                                {"field": "admin_name", "message": "管理员姓名长度超过限制（最大50字符）"}
                            )
                        else:
                            admin = Admin.query.filter(Admin.real_name == admin_name.strip()).first()
                            if not admin:
                                admin = Admin.query.filter(Admin.username == admin_name.strip()).first()
                            if not admin:
                                row_errors.append(
                                    {"field": "admin_name", "message": f'管理员 "{admin_name}" 在系统中不存在'}
                                )
                            else:
                                if admin.role not in ["admin", "teacher"]:
                                    row_errors.append(
                                        {
                                            "field": "admin_name",
                                            "message": f'用户 "{admin_name}" 的角色不是管理员或教师，无法担任设备管理员',
                                        }
                                    )

                    if row_errors:
                        failed_count += 1
                        messages.append(
                            {
                                "action": "失败",
                                "message": "; ".join([f'{err["field"]}: {err["message"]}' for err in row_errors]),
                                "row_data": row_dict,
                                "error_fields": [err["field"] for err in row_errors],
                            }
                        )
                        continue

                    new_device = Device(
                        device_id=str(device_id),
                        name=name or str(device_id),
                        class_info_id=class_info.id if class_info else None,
                        admin_id=admin.id if admin else None,
                        status="offline",
                    )

                    db.session.add(new_device)
                    success_count += 1
                    messages.append({"action": "成功", "message": f"创建设备 {str(device_id)}", "row_data": row_dict})

                except Exception as e:
                    failed_count += 1
                    messages.append(
                        {
                            "action": "失败",
                            "message": str(e),
                            "row_data": dict(zip(headers, row)) if row else None,
                            "error_fields": ["system"],
                        }
                    )

            db.session.commit()

            return APIResponse.success(
                data={
                    "success": True,
                    "total": success_count + failed_count,
                    "success_count": success_count,
                    "failed_count": failed_count,
                    "messages": messages,
                }
            )

        except Exception as e:
            db.session.rollback()
            return APIResponse.server_error(message=str(e))


@ns_devices.route("/export")
class DeviceExport(Resource):

    @ns_devices.doc("device_export", description="导出设备数据", security="Bearer")
    @requires_permission("device.view")
    def get(self):
        """
        导出设备数据

        支持JSON和Excel格式导出。

        参数：
        - format: 导出格式（json 或 excel，默认excel）

        返回设备数据文件下载。
        """
        try:
            format_type = request.args.get("format", "excel")

            devices = Device.query.all()

            if format_type == "json":
                device_list = []
                for device in devices:
                    device_data = {
                        "id": device.id,
                        "device_id": device.device_id,
                        "name": device.name,
                        "status": device.status,
                        "is_online": device.is_online,
                        "last_heartbeat": (
                            device.last_heartbeat.strftime("%Y-%m-%d %H:%M:%S") if device.last_heartbeat else ""
                        ),
                        "wifi_signal": device.wifi_signal,
                        "uptime": device.uptime,
                        "box_a_status": device.box_a_status,
                        "box_b_status": device.box_b_status,
                        "system_state": device.system_state,
                        "class_info_id": device.class_info_id,
                        "class_name": device.class_info.name if device.class_info else "",
                        "admin_id": device.admin_id,
                        "admin_name": device.admin.name if device.admin else "",
                        "created_at": device.created_at.strftime("%Y-%m-%d %H:%M:%S") if device.created_at else "",
                        "updated_at": device.updated_at.strftime("%Y-%m-%d %H:%M:%S") if device.updated_at else "",
                    }
                    device_list.append(device_data)

                json_str = json.dumps(device_list, ensure_ascii=False, indent=2)
                return Response(
                    json_str,
                    mimetype="application/json",
                    headers={
                        "Content-Disposition": f'attachment; filename=devices_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'  # noqa: E501
                    },
                )

            else:
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "设备数据"

                headers = [
                    "设备标识",
                    "设备名称",
                    "状态",
                    "是否在线",
                    "最后心跳",
                    "WiFi信号",
                    "班级名称",
                    "管理员姓名",
                    "创建时间",
                    "更新时间",
                ]
                sheet.append(headers)

                header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                header_fill = openpyxl.styles.PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
                for col in range(1, len(headers) + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill

                for device in devices:
                    row_data = [
                        device.device_id,
                        device.name,
                        device.status,
                        "是" if device.is_online else "否",
                        device.last_heartbeat.strftime("%Y-%m-%d %H:%M:%S") if device.last_heartbeat else "",
                        device.wifi_signal,
                        device.class_info.name if device.class_info else "",
                        device.admin.name if device.admin else "",
                        device.created_at.strftime("%Y-%m-%d %H:%M:%S") if device.created_at else "",
                        device.updated_at.strftime("%Y-%m-%d %H:%M:%S") if device.updated_at else "",
                    ]
                    sheet.append(row_data)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                return send_file(
                    output,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    download_name=f'devices_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                )

        except Exception as e:
            return APIResponse.server_error(message=str(e))
