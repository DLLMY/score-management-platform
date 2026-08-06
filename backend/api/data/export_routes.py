from flask import request, send_file
import io
from flask_restx import Namespace, Resource, fields
from models import User, ScoreRule, Device, ScoreRecord, ScoreCategory
from utils.permission import requires_permission
from utils.response import APIResponse
from services.export_service import export_service
from datetime import datetime
from sqlalchemy.orm import joinedload

"""
数据导出API路由
支持Excel和PDF格式的数据导出
"""
ns_export = Namespace("export", description="数据导出相关操作")
export_format_model = ns_export.model(
    "ExportFormat",
    {
        "format": fields.String(required=True, description="导出格式：excel 或 pdf"),
        "type": fields.String(required=True, description="导出类型：users/rules/devices/records/summary"),
    },
)


@ns_export.route("/")
class ExportData(Resource):
    @ns_export.doc("export_data", description="导出数据", security="Bearer")
    @ns_export.expect(export_format_model)
    @ns_export.response(200, "导出成功")
    @ns_export.response(400, "参数错误")
    @requires_permission("report.export")
    def post(self):
        """
        导出数据
        支持导出学生、规则、设备、积分记录等数据，支持Excel和PDF格式。
        请求体：
        - format: 导出格式（excel 或 pdf）
        - type: 导出类型（users/rules/devices/records/summary）
        返回对应的文件下载。
        """
        data = request.get_json()
        export_format = data.get("format", "excel").lower()
        export_type = data.get("type", "users").lower()
        if export_format not in ["excel", "pdf"]:
            return APIResponse.bad_request(message="不支持的导出格式，支持 excel 和 pdf")
        if export_type not in ["users", "rules", "devices", "records", "summary"]:
            return APIResponse.bad_request(message="不支持的导出类型")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            if export_type == "users":
                users = User.query.all()
                user_data = [
                    {
                        "id": u.id,
                        "name": u.name,
                        "gender": u.gender,
                        "class_name": u.class_name,
                        "phone": u.phone,
                        "card_id": u.card_id,
                        "current_score": u.current_score,
                        "created_at": u.created_at.isoformat() if u.created_at else None,
                    }
                    for u in users
                ]
                if export_format == "excel":
                    output = export_service.export_users_to_excel(user_data)
                    filename = f"users_{timestamp}.xlsx"
                    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    output = export_service.export_users_to_pdf(user_data, "学生列表报告")
                    filename = f"users_{timestamp}.pdf"
                    mimetype = "application/pdf"
            elif export_type == "rules":
                rules = ScoreRule.query.all()
                rule_data = [
                    {
                        "id": r.id,
                        "name": r.name,
                        "description": r.description,
                        "category_id": r.category_id,
                        "category_name": r.category.name if r.category else None,
                        "score": r.score,
                        "is_active": r.is_active,
                        "daily_limit": r.daily_limit,
                        "min_interval": r.min_interval,
                        "created_at": r.created_at.isoformat() if r.created_at else None,
                    }
                    for r in rules
                ]
                if export_format == "excel":
                    output = export_service.export_rules_to_excel(rule_data)
                    filename = f"rules_{timestamp}.xlsx"
                    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    output = export_service.export_rules_to_pdf(rule_data, "积分规则报告")
                    filename = f"rules_{timestamp}.pdf"
                    mimetype = "application/pdf"
            elif export_type == "devices":
                devices = Device.query.all()
                device_data = [
                    {
                        "id": d.id,
                        "device_id": d.device_id,
                        "name": d.name,
                        "status": d.status,
                        "is_online": d.status == "online",
                        "wifi_signal": d.wifi_signal,
                        "class_name": d.class_info.name if d.class_info else None,
                        "admin_name": d.admin.real_name if d.admin else None,
                        "created_at": d.created_at.isoformat() if d.created_at else None,
                    }
                    for d in devices
                ]
                if export_format == "excel":
                    output = export_service.export_devices_to_excel(device_data)
                    filename = f"devices_{timestamp}.xlsx"
                    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    output = export_service.export_devices_to_pdf(device_data, "设备列表报告")
                    filename = f"devices_{timestamp}.pdf"
                    mimetype = "application/pdf"
            elif export_type == "records":
                # 性能优化：使用 joinedload 预加载关联数据，消除 N+1 查询
                records = (
                    ScoreRecord.query.options(joinedload(ScoreRecord.user), joinedload(ScoreRecord.rule))
                    .order_by(ScoreRecord.created_at.desc())
                    .all()
                )
                record_data = [
                    {
                        "id": r.id,
                        "user_id": r.user_id,
                        "user_name": r.user.name if r.user else None,
                        "card_id": r.user.card_id if r.user else None,
                        "score_change": r.score_change,
                        "new_score": r.new_score,
                        "rule_id": r.rule_id,
                        "rule_name": r.rule.name if r.rule else None,
                        "category_name": r.rule.category.name if r.rule and r.rule.category else None,
                        "description": r.description,
                        "created_at": r.created_at.isoformat() if r.created_at else None,
                        "operator": r.operator,
                    }
                    for r in records
                ]
                if export_format == "excel":
                    output = export_service.export_records_to_excel(record_data)
                    filename = f"records_{timestamp}.xlsx"
                    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    output = export_service.export_records_to_pdf(record_data, "积分记录报告")
                    filename = f"records_{timestamp}.pdf"
                    mimetype = "application/pdf"
            elif export_type == "summary":
                users_count = User.query.count()
                rules_count = ScoreRule.query.count()
                devices_count = Device.query.count()
                online_devices = Device.query.filter_by(status="online").count()
                records_count = ScoreRecord.query.count()
                output = export_service.export_summary_report(
                    users_count, rules_count, devices_count, online_devices, records_count
                )
                filename = f"summary_{timestamp}.pdf"
                mimetype = "application/pdf"
            return send_file(output, mimetype=mimetype, as_attachment=True, download_name=filename)
        except Exception as e:
            return APIResponse.server_error(message=f"导出失败: {str(e)}")


@ns_export.route("/users")
class ExportUsers(Resource):
    @ns_export.doc(
        "export_users",
        description="导出学生数据",
        security="Bearer",
        params={"format": "导出格式：excel（默认）或 pdf"},
    )
    @ns_export.response(200, "导出成功")
    @requires_permission("report.export")
    def get(self):
        """
        导出学生数据
        查询参数：
        - format: 导出格式（excel 或 pdf，默认excel）
        返回学生数据文件下载。
        """
        export_format = request.args.get("format", "excel").lower()
        if export_format not in ["excel", "pdf"]:
            return APIResponse.bad_request(message="不支持的导出格式")
        users = User.query.all()
        user_data = [
            {
                "id": u.id,
                "name": u.name,
                "gender": u.gender,
                "class_name": u.class_name,
                "phone": u.phone,
                "card_id": u.card_id,
                "current_score": u.current_score,
                "created_at": u.created_at.isoformat() if u.created_at else None,
            }
            for u in users
        ]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            if export_format == "excel":
                output = export_service.export_users_to_excel(user_data)
                filename = f"users_{timestamp}.xlsx"
                mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                output = export_service.export_users_to_pdf(user_data, "学生列表报告")
                filename = f"users_{timestamp}.pdf"
                mimetype = "application/pdf"
            return send_file(output, mimetype=mimetype, as_attachment=True, download_name=filename)
        except Exception as e:
            return APIResponse.server_error(message=f"导出失败: {str(e)}")


@ns_export.route("/rules")
class ExportRules(Resource):
    @ns_export.doc(
        "export_rules",
        description="导出积分规则",
        security="Bearer",
        params={"format": "导出格式：excel（默认）或 pdf"},
    )
    @ns_export.response(200, "导出成功")
    @requires_permission("report.export")
    def get(self):
        """
        导出积分规则
        查询参数：
        - format: 导出格式（excel 或 pdf，默认excel）
        返回积分规则文件下载。
        """
        export_format = request.args.get("format", "excel").lower()
        if export_format not in ["excel", "pdf"]:
            return APIResponse.bad_request(message="不支持的导出格式")
        rules = ScoreRule.query.all()
        rule_data = [
            {
                "id": r.id,
                "name": r.name,
                "description": r.description,
                "category_id": r.category_id,
                "category_name": r.category.name if r.category else None,
                "score": r.score,
                "is_active": r.is_active,
                "daily_limit": r.daily_limit,
                "min_interval": r.min_interval,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in rules
        ]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            if export_format == "excel":
                output = export_service.export_rules_to_excel(rule_data)
                filename = f"rules_{timestamp}.xlsx"
                mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                output = export_service.export_rules_to_pdf(rule_data, "积分规则报告")
                filename = f"rules_{timestamp}.pdf"
                mimetype = "application/pdf"
            return send_file(output, mimetype=mimetype, as_attachment=True, download_name=filename)
        except Exception as e:
            return APIResponse.server_error(message=f"导出失败: {str(e)}")


@ns_export.route("/devices")
class ExportDevices(Resource):
    @ns_export.doc(
        "export_devices",
        description="导出设备数据",
        security="Bearer",
        params={"format": "导出格式：excel（默认）或 pdf"},
    )
    @ns_export.response(200, "导出成功")
    @requires_permission("report.export")
    def get(self):
        """
        导出设备数据
        查询参数：
        - format: 导出格式（excel 或 pdf，默认excel）
        返回设备数据文件下载。
        """
        export_format = request.args.get("format", "excel").lower()
        if export_format not in ["excel", "pdf"]:
            return APIResponse.bad_request(message="不支持的导出格式")
        devices = Device.query.all()
        device_data = [
            {
                "id": d.id,
                "device_id": d.device_id,
                "name": d.name,
                "status": d.status,
                "is_online": d.status == "online",
                "wifi_signal": d.wifi_signal,
                "class_name": d.class_info.name if d.class_info else None,
                "admin_name": d.admin.real_name if d.admin else None,
                "created_at": d.created_at.isoformat() if d.created_at else None,
            }
            for d in devices
        ]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            if export_format == "excel":
                output = export_service.export_devices_to_excel(device_data)
                filename = f"devices_{timestamp}.xlsx"
                mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                output = export_service.export_devices_to_pdf(device_data, "设备列表报告")
                filename = f"devices_{timestamp}.pdf"
                mimetype = "application/pdf"
            return send_file(output, mimetype=mimetype, as_attachment=True, download_name=filename)
        except Exception as e:
            return APIResponse.server_error(message=f"导出失败: {str(e)}")


@ns_export.route("/records")
class ExportRecords(Resource):
    @ns_export.doc(
        "export_records",
        description="导出积分记录",
        security="Bearer",
        params={"format": "导出格式：excel（默认）或 pd", "limit": "限制导出记录数量（默认10000）"},
    )
    @ns_export.response(200, "导出成功")
    @requires_permission("report.export")
    def get(self):
        """
        导出积分记录
        查询参数：
        - format: 导出格式（excel 或 pdf，默认excel）
        - limit: 限制导出记录数量（默认10000）
        返回积分记录文件下载。
        """
        export_format = request.args.get("format", "excel").lower()
        limit = request.args.get("limit", 10000, type=int)
        if export_format not in ["excel", "pdf"]:
            return APIResponse.bad_request(message="不支持的导出格式")
        records = ScoreRecord.query.order_by(ScoreRecord.created_at.desc()).limit(limit).all()
        record_data = [
            {
                "id": r.id,
                "user_id": r.user_id,
                "user_name": r.user.name if r.user else None,
                "card_id": r.user.card_id if r.user else None,
                "score_change": r.score_change,
                "new_score": r.new_score,
                "rule_id": r.rule_id,
                "rule_name": r.rule.name if r.rule else None,
                "category_name": r.rule.category.name if r.rule and r.rule.category else None,
                "description": r.description,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "operator": r.operator,
            }
            for r in records
        ]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            if export_format == "excel":
                output = export_service.export_records_to_excel(record_data)
                filename = f"records_{timestamp}.xlsx"
                mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                output = export_service.export_records_to_pdf(record_data, "积分记录报告")
                filename = f"records_{timestamp}.pdf"
                mimetype = "application/pdf"
            return send_file(output, mimetype=mimetype, as_attachment=True, download_name=filename)
        except Exception as e:
            return APIResponse.server_error(message=f"导出失败: {str(e)}")


@ns_export.route("/categories")
class ExportCategories(Resource):
    @ns_export.doc(
        "export_categories",
        description="导出分类数据",
        security="Bearer",
        params={"format": "导出格式：excel（默认）或 pdf"},
    )
    @ns_export.response(200, "导出成功")
    @requires_permission("report.export")
    def get(self):
        """
        导出分类数据
        查询参数：
        - format: 导出格式（excel 或 pdf，默认excel）
        返回分类数据文件下载。
        """
        export_format = request.args.get("format", "excel").lower()
        if export_format not in ["excel", "pdf"]:
            return APIResponse.bad_request(message="不支持的导出格式")
        categories = ScoreCategory.query.all()
        category_data = [
            {
                "id": c.id,
                "name": c.name,
                "description": c.description,
                "color": c.color,
                "created_at": c.created_at.isoformat() if c.created_at else None,
            }
            for c in categories
        ]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        headers = ["ID", "名称", "描述", "颜色", "创建时间"]
        try:
            if export_format == "excel":
                output = export_service.export_to_excel(category_data, headers)
                filename = f"categories_{timestamp}.xlsx"
                mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                output = export_service.export_to_pdf("积分分类报告", category_data, headers)
                filename = f"categories_{timestamp}.pdf"
                mimetype = "application/pdf"
            return send_file(output, mimetype=mimetype, as_attachment=True, download_name=filename)
        except Exception as e:
            return APIResponse.server_error(message=f"导出失败: {str(e)}")


@ns_export.route("/summary")
class ExportSummary(Resource):
    @ns_export.doc("export_summary", description="导出系统数据汇总报告", security="Bearer")
    @ns_export.response(200, "导出成功")
    @requires_permission("report.export")
    def get(self):
        """
        导出系统数据汇总报告（PDF格式）
        返回包含学生总数、规则数、设备数、在线设备数、积分记录数等统计数据的汇总报告。
        """
        users_count = User.query.count()
        rules_count = ScoreRule.query.count()
        devices_count = Device.query.count()
        online_devices = Device.query.filter_by(status="online").count()
        records_count = ScoreRecord.query.count()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            output = export_service.export_summary_report(
                users_count, rules_count, devices_count, online_devices, records_count
            )
            filename = f"summary_{timestamp}.pdf"
            return send_file(output, mimetype="application/pdf", as_attachment=True, download_name=filename)
        except Exception as e:
            return APIResponse.server_error(message=f"导出失败: {str(e)}")


@ns_export.route("/errors")
class ExportErrors(Resource):
    @ns_export.doc("export_errors", description="导出导入错误数据", security="Bearer")
    @ns_export.response(200, "导出成功")
    @ns_export.response(400, "参数错误")
    @requires_permission("report.export")
    def post(self):
        """
        导出导入错误数据
        接收导入失败的错误数据列表，将其导出为Excel文件，方便用户修正后重新导入。
        请求体：
        - errors: 错误数据列表（从导入API返回的messages中筛选失败记录）
        - module: 模块名称（users/devices/classes/exams/subjects）
        返回错误数据Excel文件下载。
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        data = request.get_json()
        errors = data.get("errors", [])
        module = data.get("module", "")
        if not errors:
            return APIResponse.bad_request(message="没有错误数据可导出")
        wb = Workbook()
        ws = wb.active
        ws.title = "导入错误数据"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        all_keys = set()
        for error in errors:
            if isinstance(error.get("row_data"), dict):
                all_keys.update(error["row_data"].keys())
        base_columns = ["行号", "错误字段", "错误信息"]
        data_columns = sorted(list(all_keys))
        headers = base_columns + data_columns
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        for row_idx, error in enumerate(errors, 2):
            row_num = error.get("row", "")
            error_fields = ", ".join(error.get("error_fields", []))
            message = error.get("message", "")
            ws.cell(row=row_idx, column=1, value=row_num)
            ws.cell(row=row_idx, column=2, value=error_fields)
            ws.cell(row=row_idx, column=3, value=message)
            row_data = error.get("row_data", {})
            if isinstance(row_data, dict):
                for col_idx, key in enumerate(data_columns, 4):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"import_errors_{module}_{timestamp}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
