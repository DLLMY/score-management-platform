import logging

import json
import io
import openpyxl
from flask import Blueprint, request, send_file
from utils.response import APIResponse

logger = logging.getLogger(__name__)

"""
文件下载路由模块
"""
download_bp = Blueprint("download", __name__)


@download_bp.route("/api/scores/template/download")
def download_score_template():
    """
    下载成绩导入Excel模板
    参数:
    - exam_id: 考试ID（可选）
    - class_name: 班级名称（可选）
    """
    from models import Exam, User, get_by_id
    from utils.permission import requires_permission

    @requires_permission("system.settings")
    def generate_template():
        try:
            exam_id = request.args.get("exam_id", type=int)
            class_name = request.args.get("class_name")
            class_id = request.args.get("class_id", type=int)
            exam = None
            if exam_id:
                exam = get_by_id(Exam, exam_id)
            # 获取科目列表
            if exam and exam.subjects:
                subjects = (
                    json.loads(exam.subjects) if isinstance(exam.subjects, str) else exam.subjects
                )
            else:
                subjects = ["语文", "数学", "英语"]
            # 获取学生列表（兼容 class_name / class_id）
            if class_id:
                students = (
                    User.query.filter(User.class_info_id == class_id).order_by(User.card_id).all()
                )
            elif class_name:
                students = User.query.filter_by(class_name=class_name).order_by(User.card_id).all()
            else:
                students = User.query.order_by(User.class_name, User.card_id).all()
            wb = openpyxl.Workbook()
            # 创建成绩导入表
            sheet_data = wb.active
            sheet_data.title = "成绩导入"
            # 创建表头
            headers = ["学号", "姓名", "班级"] + subjects
            for col_idx, header in enumerate(headers, 1):
                sheet_data.cell(row=1, column=col_idx, value=header)
            # 填充学生数据
            for row_idx, student in enumerate(students, 2):
                sheet_data.cell(row=row_idx, column=1, value=student.card_id)
                sheet_data.cell(row=row_idx, column=2, value=student.name)
                sheet_data.cell(row=row_idx, column=3, value=student.class_name)
                # 科目分数列留空，由教师填写
                for col_offset in range(len(subjects)):
                    sheet_data.cell(row=row_idx, column=4 + col_offset, value="")
            # 设置列宽
            sheet_data.column_dimensions["A"].width = 15
            sheet_data.column_dimensions["B"].width = 10
            sheet_data.column_dimensions["C"].width = 15
            for i in range(len(subjects)):
                col_letter = openpyxl.utils.get_column_letter(4 + i)
                sheet_data.column_dimensions[col_letter].width = 12
            # 创建说明表
            sheet_notes = wb.create_sheet(title="填写说明")
            notes_data = [
                ["列名", "说明", "填写方式", "示例"],
                ["学号", "学生的学号，系统自动填入，请勿修改", "系统自动", "202401001"],
                ["姓名", "学生姓名，系统自动填入，仅作参考", "系统自动", "张三"],
                ["班级", "班级名称，系统自动填入", "系统自动", "高一(1)班"],
                ["科目", "科目名称，对应考试科目", "系统自动", "语文"],
                ["分数", "学生成绩，教师必须填写，必须为数字(0-100)", "教师填写", "85"],
            ]
            for row_idx, row_data in enumerate(notes_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    sheet_notes.cell(row=row_idx, column=col_idx, value=cell_value)
            # 设置说明表列宽
            sheet_notes.column_dimensions["A"].width = 12
            sheet_notes.column_dimensions["B"].width = 40
            sheet_notes.column_dimensions["C"].width = 12
            sheet_notes.column_dimensions["D"].width = 15
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"score_import_template_{class_name or 'all'}.xlsx"
            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
        except Exception as e:
            logger.error("%s: %s", "生成模板失败", e)
            return APIResponse.error(message="生成模板失败", status_code=500)

    return generate_template()
