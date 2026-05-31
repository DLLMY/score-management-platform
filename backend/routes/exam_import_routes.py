from flask import request, jsonify
from flask_restx import Namespace, Resource, fields
from models import db, Exam, Score, User, ClassInfo, Admin
from utils.permission import requires_admin
from datetime import datetime
import openpyxl
from io import BytesIO

ns_exam_import = Namespace('exam-import', description='成绩导入增强功能')

exam_import_request = ns_exam_import.model('ExamImportRequest', {
    'exam_id': fields.Integer(required=True, description='考试ID'),
    'entered_by': fields.Integer(description='录入人ID'),
    'update_existing': fields.Boolean(description='是否更新已存在的成绩', default=True),
    'validate_score': fields.Boolean(description='是否验证分数范围', default=True)
})


class ScoreImportHelper:
    @staticmethod
    def validate_excel_format(sheet) -> dict:
        """
        验证Excel格式是否正确

        Returns:
            dict: {valid: bool, errors: list, headers: list}
        """
        errors = []

        if sheet.max_row < 2:
            errors.append('Excel文件没有数据行')
            return {'valid': False, 'errors': errors, 'headers': []}

        headers = [cell.value for cell in sheet[1]]

        required_headers = ['card_id', 'subject', 'score']
        optional_headers = ['full_score', 'student_name', 'class_name']

        header_lower = [h.lower().strip() if h else '' for h in headers]

        for req in required_headers:
            if req not in header_lower:
                matched = False
                for h in header_lower:
                    if req.replace('_', '') in h.replace('_', ''):
                        matched = True
                        break
                if not matched:
                    errors.append(f'缺少必需列: {req}')

        return {'valid': len(errors) == 0, 'errors': errors, 'headers': headers}

    @staticmethod
    def find_column_index(headers, target_column: str) -> int:
        """查找列的索引，支持模糊匹配"""
        for i, header in enumerate(headers):
            if header and target_column.lower() in header.lower():
                return i
        return -1

    @staticmethod
    def parse_score_value(value) -> float:
        """解析分数值"""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def validate_score_range(score: float, full_score: float = 100) -> tuple:
        """验证分数是否在合理范围内"""
        if score is None:
            return False, '分数为空'

        if score < 0:
            return False, '分数不能为负数'

        if score > full_score * 1.5:
            return False, f'分数超过满分的150% ({full_score * 1.5})'

        return True, 'valid'


@ns_exam_import.route('/validate')
class ValidateImportFile(Resource):
    @ns_exam_import.doc('validate_import_file', description='验证导入文件格式')
    @requires_admin
    def post(self):
        """
        验证导入文件格式

        上传Excel文件后，先验证格式是否正确。
        """
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传文件'}), 400

        file = request.files['file']
        exam_id = request.args.get('exam_id', type=int)

        if not exam_id:
            return jsonify({'success': False, 'message': '缺少考试ID'}), 400

        try:
            exam = Exam.query.get(exam_id)
            if not exam:
                return jsonify({'success': False, 'message': '考试不存在'}), 404

            file_content = file.read()
            wb = openpyxl.load_workbook(BytesIO(file_content))
            sheet = wb.active

            validation = ScoreImportHelper.validate_excel_format(sheet)

            if not validation['valid']:
                return jsonify({
                    'success': False,
                    'message': '文件格式验证失败',
                    'errors': validation['errors']
                }), 400

            headers = validation['headers']
            data_preview = []
            card_id_idx = ScoreImportHelper.find_column_index(headers, 'card_id')
            subject_idx = ScoreImportHelper.find_column_index(headers, 'subject')
            score_idx = ScoreImportHelper.find_column_index(headers, 'score')

            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=6, values_only=True)):
                if all(cell is None for cell in row):
                    continue

                card_id = row[card_id_idx] if card_id_idx >= 0 else None
                subject = row[subject_idx] if subject_idx >= 0 else None
                score_val = row[score_idx] if score_idx >= 0 else None

                student = User.query.filter_by(card_id=card_id).first() if card_id else None

                data_preview.append({
                    'row': i + 2,
                    'card_id': str(card_id) if card_id else None,
                    'student_name': student.name if student else '未找到',
                    'subject': subject,
                    'score': score_val,
                    'status': 'ready' if student else 'student_not_found'
                })

            return jsonify({
                'success': True,
                'message': '文件格式验证通过',
                'exam_name': exam.name,
                'subjects': exam.subjects,
                'preview': data_preview,
                'total_rows': sheet.max_row - 1
            })

        except Exception as e:
            return jsonify({'success': False, 'message': f'验证失败: {str(e)}'}), 500


@ns_exam_import.route('/preview')
class PreviewImportData(Resource):
    @ns_exam_import.doc('preview_import_data', description='预览导入数据')
    @requires_admin
    def post(self):
        """
        预览导入数据

        返回导入数据的预览，不实际写入数据库。
        """
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传文件'}), 400

        file = request.files['file']
        exam_id = request.form.get('exam_id', type=int)

        if not exam_id:
            return jsonify({'success': False, 'message': '缺少考试ID'}), 400

        try:
            exam = Exam.query.get(exam_id)
            if not exam:
                return jsonify({'success': False, 'message': '考试不存在'}), 404

            file_content = file.read()
            wb = openpyxl.load_workbook(BytesIO(file_content))
            sheet = wb.active

            validation = ScoreImportHelper.validate_excel_format(sheet)
            if not validation['valid']:
                return jsonify({
                    'success': False,
                    'message': '文件格式错误',
                    'errors': validation['errors']
                }), 400

            headers = validation['headers']
            card_id_idx = ScoreImportHelper.find_column_index(headers, 'card_id')
            subject_idx = ScoreImportHelper.find_column_index(headers, 'subject')
            score_idx = ScoreImportHelper.find_column_index(headers, 'score')
            full_score_idx = ScoreImportHelper.find_column_index(headers, 'full_score')

            results = []
            errors = []
            total_rows = sheet.max_row - 1

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                if all(cell is None for cell in row):
                    continue

                card_id = str(row[card_id_idx]).strip() if card_id_idx >= 0 and row[card_id_idx] else None
                subject = row[subject_idx] if subject_idx >= 0 else None
                score_val = ScoreImportHelper.parse_score_value(row[score_idx]) if score_idx >= 0 else None
                full_score = ScoreImportHelper.parse_score_value(row[full_score_idx]) if full_score_idx >= 0 else 100

                if not card_id:
                    errors.append(f'行{i+2}: 学号为空')
                    continue

                student = User.query.filter_by(card_id=card_id).first()
                if not student:
                    errors.append(f'行{i+2}: 学号{card_id}不存在')
                    continue

                if not subject:
                    errors.append(f'行{i+2}: 科目为空')
                    continue

                is_valid, msg = ScoreImportHelper.validate_score_range(score_val, full_score)
                if not is_valid:
                    errors.append(f'行{i+2}: {student.name}-{subject} - {msg}')

                existing_score = Score.query.filter_by(
                    exam_id=exam_id,
                    student_id=student.id,
                    subject=subject
                ).first()

                results.append({
                    'row': i + 2,
                    'card_id': card_id,
                    'student_name': student.name,
                    'class_name': student.class_name,
                    'subject': subject,
                    'score': score_val,
                    'full_score': full_score,
                    'will_update': existing_score is not None,
                    'will_insert': existing_score is None
                })

            return jsonify({
                'success': True,
                'message': f'预览完成，共{total_rows}行',
                'results': results[:50],
                'errors': errors[:20],
                'summary': {
                    'total': len(results),
                    'will_insert': sum(1 for r in results if r['will_insert']),
                    'will_update': sum(1 for r in results if r['will_update']),
                    'error_count': len(errors)
                }
            })

        except Exception as e:
            return jsonify({'success': False, 'message': f'预览失败: {str(e)}'}), 500


@ns_exam_import.route('/execute')
class ExecuteImport(Resource):
    @ns_exam_import.doc('execute_import', description='执行成绩导入')
    @requires_admin
    def post(self):
        """
        执行成绩导入

        实际写入数据库。
        """
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传文件'}), 400

        file = request.files['file']
        exam_id = request.form.get('exam_id', type=int)
        entered_by = request.form.get('entered_by', type=int, default=1)
        update_existing = request.form.get('update_existing', 'true').lower() == 'true'
        validate_score = request.form.get('validate_score', 'true').lower() == 'true'

        if not exam_id:
            return jsonify({'success': False, 'message': '缺少考试ID'}), 400

        try:
            exam = Exam.query.get(exam_id)
            if not exam:
                return jsonify({'success': False, 'message': '考试不存在'}), 404

            if exam.status != 'published':
                return jsonify({'success': False, 'message': '考试未发布，无法导入成绩'}), 400

            file_content = file.read()
            wb = openpyxl.load_workbook(BytesIO(file_content))
            sheet = wb.active

            validation = ScoreImportHelper.validate_excel_format(sheet)
            if not validation['valid']:
                return jsonify({
                    'success': False,
                    'message': '文件格式错误',
                    'errors': validation['errors']
                }), 400

            headers = validation['headers']
            card_id_idx = ScoreImportHelper.find_column_index(headers, 'card_id')
            subject_idx = ScoreImportHelper.find_column_index(headers, 'subject')
            score_idx = ScoreImportHelper.find_column_index(headers, 'score')
            full_score_idx = ScoreImportHelper.find_column_index(headers, 'full_score')

            success_count = 0
            update_count = 0
            insert_count = 0
            failed_count = 0
            errors = []

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                if all(cell is None for cell in row):
                    continue

                try:
                    card_id = str(row[card_id_idx]).strip() if card_id_idx >= 0 and row[card_id_idx] else None
                    subject = row[subject_idx] if subject_idx >= 0 else None
                    score_val = ScoreImportHelper.parse_score_value(row[score_idx]) if score_idx >= 0 else None
                    full_score = ScoreImportHelper.parse_score_value(row[full_score_idx]) if full_score_idx >= 0 else 100

                    if not card_id or not subject:
                        failed_count += 1
                        errors.append(f'行{i+2}: 必需字段为空')
                        continue

                    student = User.query.filter_by(card_id=card_id).first()
                    if not student:
                        failed_count += 1
                        errors.append(f'行{i+2}: 学号{card_id}不存在')
                        continue

                    if validate_score:
                        is_valid, msg = ScoreImportHelper.validate_score_range(score_val, full_score)
                        if not is_valid:
                            failed_count += 1
                            errors.append(f'行{i+2}: {student.name}-{subject} - {msg}')
                            continue

                    existing_score = Score.query.filter_by(
                        exam_id=exam_id,
                        student_id=student.id,
                        subject=subject
                    ).first()

                    if existing_score:
                        if update_existing:
                            existing_score.score = score_val
                            existing_score.full_score = full_score
                            existing_score.status = 'pending'
                            existing_score.entered_by = entered_by
                            update_count += 1
                        else:
                            failed_count += 1
                            errors.append(f'行{i+2}: {student.name}-{subject}已存在')
                            continue
                    else:
                        score = Score(
                            exam_id=exam_id,
                            student_id=student.id,
                            subject=subject,
                            score=score_val,
                            full_score=full_score,
                            status='pending',
                            entered_by=entered_by
                        )
                        db.session.add(score)
                        insert_count += 1

                    success_count += 1

                except Exception as e:
                    failed_count += 1
                    errors.append(f'行{i+2}: {str(e)}')

            db.session.commit()

            return jsonify({
                'success': True,
                'message': f'导入完成',
                'imported_count': success_count,
                'insert_count': insert_count,
                'update_count': update_count,
                'failed_count': failed_count,
                'errors': errors[:50]
            })

        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500


@ns_exam_import.route('/template')
class DownloadTemplate(Resource):
    @ns_exam_import.doc('download_template', description='下载导入模板')
    @requires_admin
    def get(self):
        """
        下载成绩导入Excel模板
        """
        exam_id = request.args.get('exam_id', type=int)

        if exam_id:
            exam = Exam.query.get(exam_id)
            subjects = exam.subjects if exam else ['语文', '数学', '英语']
        else:
            subjects = ['语文', '数学', '英语']

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '成绩导入'

        headers = ['card_id', 'student_name', 'subject', 'score', 'full_score']
        sheet.append(headers)

        sample_students = User.query.limit(5).all()
        for student in sample_students:
            for subject in subjects[:2]:
                sheet.append([
                    student.card_id,
                    student.name,
                    subject,
                    '',
                    100
                ])

        for col in ['A', 'B', 'C', 'D', 'E']:
            sheet.column_dimensions[col].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            attachment_filename=f'score_import_template.xlsx'
        )


@ns_exam_import.route('/history')
class ImportHistory(Resource):
    @ns_exam_import.doc('get_import_history', description='获取导入历史')
    @requires_admin
    def get(self):
        """
        获取成绩导入历史记录
        """
        exam_id = request.args.get('exam_id', type=int)
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)

        query = Score.query.filter(Score.entered_by.isnot(None))

        if exam_id:
            query = query.filter_by(exam_id=exam_id)

        pagination = query.order_by(Score.entered_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )

        results = []
        for score in pagination.items:
            student = User.query.get(score.student_id)
            exam = Exam.query.get(score.exam_id)
            entered_by_admin = Admin.query.get(score.entered_by) if score.entered_by else None

            results.append({
                'id': score.id,
                'exam_name': exam.name if exam else None,
                'student_name': student.name if student else None,
                'student_card_id': student.card_id if student else None,
                'subject': score.subject,
                'score': score.score,
                'full_score': score.full_score,
                'status': score.status,
                'rank': score.rank,
                'entered_by': entered_by_admin.username if entered_by_admin else None,
                'entered_at': score.entered_at.isoformat() if score.entered_at else None
            })

        return jsonify({
            'success': True,
            'data': results,
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })
