import io
import csv
from pathlib import Path
from typing import List, Dict, Any

from flask import send_file

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelUtils:
    """Excel文件处理工具类"""

    @staticmethod
    def create_workbook(sheets: List[Dict[str, Any]]) -> Workbook:
        """
        创建Excel工作簿

        :param sheets: 工作表列表，每个字典包含 'name' 和 'data'
        :return: Workbook对象
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装，请先安装: pip install openpyxl")

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_data in sheets:
            sheet_name = sheet_data.get("name", "Sheet")
            data = sheet_data.get("data", [])
            headers = sheet_data.get("headers", [])

            ws = wb.create_sheet(title=sheet_name[:31])

            if headers:
                ExcelUtils._apply_header_style(ws, headers, 1)

            for row_idx, row in enumerate(data, start=2):
                for col_idx, cell_value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            ExcelUtils._auto_adjust_columns(ws)

        return wb

    @staticmethod
    def _apply_header_style(ws, headers: List[str], row_num: int):
        """应用表头样式"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    @staticmethod
    def _auto_adjust_columns(ws):
        """自动调整列宽"""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    # 逐单元格列宽计算的热循环：失败（如 formula/error cell）仅跳过该列宽估算，
                    # 属可预期降级。若改为 logger 会在大表下刷屏，故保留静默并显式说明（T9 评估结论）。
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    @staticmethod
    def workbook_to_bytes(wb: Workbook) -> bytes:
        """将工作簿转换为字节流"""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_excel(sheets: List[Dict[str, Any]], filename: str = None) -> bytes:
        """
        导出数据到Excel文件

        :param sheets: 工作表数据
        :param filename: 文件名（可选）
        :return: Excel文件字节流
        """
        wb = ExcelUtils.create_workbook(sheets)
        return ExcelUtils.workbook_to_bytes(wb)

    @staticmethod
    def export_to_csv(data: List[List[Any]], headers: List[str]) -> bytes:
        """
        导出数据到CSV文件

        :param data: 数据行列表
        :param headers: 表头列表
        :return: CSV文件字节流（UTF-8 BOM编码）
        """
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(data)
        output.seek(0)
        return output.getvalue().encode("utf-8-sig")

    @staticmethod
    def read_excel(file_bytes: bytes, sheet_name: str = None) -> Dict[str, Any]:
        """
        读取Excel文件

        :param file_bytes: 文件字节流
        :param sheet_name: 工作表名称（可选，默认读取第一个）
        :return: 包含headers和data的字典
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装")

        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)

            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            headers = []
            data = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(cell) if cell is not None else "" for cell in row]
                else:
                    data.append([cell for cell in row])

            return {"headers": headers, "data": data, "sheet_name": ws.title}
        except Exception as e:
            raise ValueError(f"读取Excel文件失败: {str(e)}")

    @staticmethod
    def read_csv(file_bytes: bytes) -> Dict[str, Any]:
        """
        读取CSV文件

        :param file_bytes: 文件字节流
        :return: 包含headers和data的字典
        """
        try:
            content = file_bytes.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)

            if not rows:
                return {"headers": [], "data": []}

            headers = rows[0]
            data = rows[1:]

            return {"headers": headers, "data": data}
        except Exception as e:
            raise ValueError(f"读取CSV文件失败: {str(e)}")

    @staticmethod
    def detect_file_type(file_bytes: bytes, filename: str) -> str:
        """
        检测文件类型

        :param file_bytes: 文件字节流
        :param filename: 文件名
        :return: 'xlsx', 'xls', 'csv' 或 None
        """
        lower_name = filename.lower()

        if lower_name.endswith(".xlsx"):
            return "xlsx"
        elif lower_name.endswith(".xls"):
            return "xls"
        elif lower_name.endswith(".csv"):
            return "csv"

        # 尝试通过内容检测
        if file_bytes[:4] == b"PK\x03\x04":
            return "xlsx"
        elif file_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            return "xls"

        return None


class ExcelTemplateGenerator:
    """Excel模板生成器"""

    USER_TEMPLATE = {
        "headers": ["姓名", "性别", "班级", "联系电话", "饭卡号", "备注"],
        "examples": [
            ["张三", "男", "25电气五年制", "13800138001", "CARD001", "测试学生"],
            ["李四", "女", "25电气五年制", "13800138002", "CARD002", ""],
            ["王五", "男", "24计算机三班", "13800138003", "CARD003", "转学生"],
        ],
        "description": [
            "1. 姓名：必填，学生真实姓名",
            "2. 性别：男/女",
            "3. 班级：学生所在班级名称",
            "4. 联系电话：11位手机号码",
            "5. 饭卡号：唯一标识，不可重复",
            "6. 备注：选填，其他说明信息",
        ],
    }

    RULE_TEMPLATE = {
        "headers": [
            "规则名称",
            "描述",
            "分类名称",
            "分数",
            "是否启用",
            "每日上限",
            "最小间隔(分钟)",
        ],
        "examples": [
            ["按时到校", "每天按时到校打卡", "日常行为", 10, "是", 1, 0],
            ["课堂表现优秀", "课堂积极回答问题", "学习成绩", 5, "是", 3, 30],
            ["迟到", "上课迟到", "日常行为", -5, "是", 3, 0],
        ],
        "description": [
            "1. 规则名称：必填，规则的名称",
            "2. 描述：规则的详细说明",
            "3. 分类名称：必须是已存在的分类名称",
            "4. 分数：正整数为加分，负整数为扣分",
            "5. 是否启用：是/否",
            "6. 每日上限：每天最多可获得/扣除的次数",
            "7. 最小间隔：两次触发的最小时间间隔（分钟）",
        ],
    }

    CATEGORY_TEMPLATE = {
        "headers": ["分类名称", "描述", "颜色"],
        "examples": [
            ["日常行为", "学生日常行为表现", "#3B82F6"],
            ["学习成绩", "学习相关的成绩表现", "#10B981"],
            ["纪律表现", "纪律遵守情况", "#F59E0B"],
        ],
        "description": [
            "1. 分类名称：必填，唯一标识",
            "2. 描述：分类的详细说明",
            "3. 颜色：十六进制颜色代码，如 #3B82F6",
        ],
    }

    @staticmethod
    def generate_template(template_type: str) -> bytes:
        """
        生成指定类型的Excel模板

        :param template_type: 'user', 'rule', 'category'
        :return: Excel文件字节流
        """
        templates = {
            "user": ExcelTemplateGenerator.USER_TEMPLATE,
            "rule": ExcelTemplateGenerator.RULE_TEMPLATE,
            "category": ExcelTemplateGenerator.CATEGORY_TEMPLATE,
        }

        if template_type not in templates:
            raise ValueError(f"不支持的模板类型: {template_type}")

        template = templates[template_type]

        sheets = [
            {"name": "导入数据", "headers": template["headers"], "data": template["examples"]},
            {
                "name": "填写说明",
                "headers": ["填写说明"],
                "data": [[desc] for desc in template["description"]],
            },
        ]

        return ExcelUtils.export_to_excel(sheets)

    @staticmethod
    def validate_import_data(
        template_type: str, headers: List[str], data: List[List[Any]]
    ) -> Dict[str, Any]:
        """
        验证导入数据格式

        :param template_type: 模板类型
        :param headers: 表头
        :param data: 数据
        :return: 验证结果 {'valid': bool, 'errors': list, 'warnings': list}
        """
        templates = {
            "user": ExcelTemplateGenerator.USER_TEMPLATE,
            "rule": ExcelTemplateGenerator.RULE_TEMPLATE,
            "category": ExcelTemplateGenerator.CATEGORY_TEMPLATE,
        }

        if template_type not in templates:
            return {"valid": False, "errors": ["不支持的模板类型"], "warnings": []}

        expected_headers = templates[template_type]["headers"]
        errors = []
        warnings = []

        # 验证表头
        actual_headers = [h.strip() for h in headers]
        expected_headers_stripped = [h.strip() for h in expected_headers]

        if actual_headers != expected_headers_stripped:
            errors.append(f"表头不匹配！期望: {expected_headers}, 实际: {actual_headers}")

        # 验证数据行数
        if len(data) == 0:
            errors.append("没有数据需要导入")

        # 逐行验证
        for row_idx, row in enumerate(data, start=2):
            row_errors = []

            if template_type == "user":
                # 验证必填字段
                if not row[0] or str(row[0]).strip() == "":
                    row_errors.append("姓名不能为空")
                if row[1] and str(row[1]).strip() not in ["男", "女"]:
                    row_errors.append('性别必须为"男"或"女"')
                if row[3] and len(str(row[3]).strip()) != 11:
                    row_errors.append("联系电话必须是11位")

            elif template_type == "rule":
                if not row[0] or str(row[0]).strip() == "":
                    row_errors.append("规则名称不能为空")
                if row[3] and not ExcelTemplateGenerator._is_number(row[3]):
                    row_errors.append("分数必须是数字")

            elif template_type == "category":
                if not row[0] or str(row[0]).strip() == "":
                    row_errors.append("分类名称不能为空")

            if row_errors:
                errors.append(f'第{row_idx}行: {", ".join(row_errors)}')

        return {"valid": len(errors) == 0, "errors": errors, "warnings": warnings}

    @staticmethod
    def _is_number(value) -> bool:
        """检查是否为数字"""
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False


# ============================================================
# B6（2026-08-23 追加）：附件下载响应构造公共 helper
# 收敛散落各路由的 send_file(output, mimetype=..., as_attachment=True, download_name=...)
# 三元组样板。只做响应构造，不碰业务字节流，纯提取零行为变化。
# ============================================================

# 常见导出格式 -> mimetype 映射（显式传 mimetype 时优先使用显式值）
ATTACHMENT_MIME_BY_EXT = {
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".csv": "text/csv;charset=utf-8",
    ".pdf": "application/pdf",
    ".json": "application/json",
    ".bin": "application/octet-stream",
    ".txt": "text/plain;charset=utf-8",
}


def build_attachment_response(output, filename: str, mimetype: str = None):
    """构造附件下载响应。

    - mimetype 显式传入时原样使用（保持各路由既有语义）
    - 未传时按 filename 扩展名推断；未知扩展名回退 application/octet-stream
    - 与 send_file(output, mimetype=..., as_attachment=True, download_name=filename) 完全等价
    """
    if mimetype is None:
        mimetype = ATTACHMENT_MIME_BY_EXT.get(
            Path(filename).suffix.lower(), "application/octet-stream"
        )
    return send_file(output, mimetype=mimetype, as_attachment=True, download_name=filename)
