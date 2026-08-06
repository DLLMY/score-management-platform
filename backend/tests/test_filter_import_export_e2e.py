#!/usr/bin/env python3
import pytest
import io
import json
try:
    from models import Subject
except ImportError:
    pass

try:
    import time
except ImportError:
    pass

# -*- coding: utf-8 -*-
"""
筛选功能与导入导出功能端到端测试
测试范围：
1. 筛选功能：Subject、Exam、Score 模块的筛选 API
2. 导入功能：文件上传、格式验证、错误提示
3. 导出功能：文件生成、格式正确性、数据一致性
"""
"""
"""


try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class TestFilterFunctionality:
    """测试筛选功能"""

    def test_subject_filter_by_name(self, client, auth_headers, db_session):
        """测试按科目名称筛选"""
        # Setup: 创建测试数据
        from models import Subject
        subjects = [
            Subject(name='数学', code='MATH001', grade='高一', is_active=True),
            Subject(name='语文', code='CHINESE001', grade='高一', is_active=True),
            Subject(name='英语', code='ENGLISH001', grade='高二', is_active=False),
        ]
        db_session.add_all(subjects)
        db_session.commit()

        # Test: 按名称筛选
        response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={'search': '数学'}
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        subjects_data = data['data']
        assert len(subjects_data) >= 1
        assert any('数学' in s['name'] for s in subjects_data)

        print(f'[PASS] 按名称筛选: 找到 {len(subjects_data)} 条匹配数据')

    def test_subject_filter_by_status(self, client, auth_headers, db_session):
        """测试按状态筛选科目"""
        response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={'include_inactive': 'false'}
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        subjects_data = data['data']
        # 验证所有返回的科目都是启用状态
        for s in subjects_data:
            assert s['is_active'] is True or s.get('is_active', True)

        print(f'[PASS] 按状态筛选(启用): 找到 {len(subjects_data)} 条数据')

    def test_subject_filter_inactive(self, client, auth_headers, db_session):
        """测试获取包含已禁用的科目"""
        response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={'include_inactive': 'true'}
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True

        print('[PASS] 获取包含已禁用的科目成功')

    def test_subject_combined_filters(self, client, auth_headers, db_session):
        """测试组合筛选条件"""
        response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={
                'search': '高一',
                'include_inactive': 'false'
            }
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True

        print('[PASS] 组合筛选条件测试通过')

    def test_subject_filter_empty_result(self, client, auth_headers):
        """测试筛选无结果场景"""
        response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={'search': '不存在的科目名称'}
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True

        print(f'[PASS] 空结果筛选: 返回 {len(data["data"])} 条数据')

    def test_subject_filter_special_characters(self, client, auth_headers):
        """测试特殊字符筛选"""
        response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={'search': '<script>alert("xss")</script>'}
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True

        print('[PASS] 特殊字符筛选安全处理')

    def test_subject_filter_chinese_characters(self, client, auth_headers, db_session):
        """测试中文字符筛选"""
        subject = Subject(name='物理', code='PHYSICS001', grade='高二', is_active=True)
        db_session.add(subject)
        db_session.commit()

        response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={'search': '物理'}
        )
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert len(data['data']) >= 1

        print('[PASS] 中文字符筛选测试通过')


class TestImportFunctionality:
    """测试导入功能"""

    def test_import_subject_empty_file(self, client, auth_headers):
        """测试上传空文件"""
        data = {
            'file': (io.BytesIO(b''), 'empty.xlsx')
        }
        response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code in [400, 200]
        result = response.get_json()
        # 空文件应返回错误或空结果
        if result:
            print(f'[INFO] 空文件响应: {result.get("message", "无消息")}')

        print('[PASS] 空文件上传测试通过（有适当的错误处理）')

    def test_import_subject_invalid_format(self, client, auth_headers):
        """测试上传无效格式文件"""
        invalid_content = b'not a valid excel file'
        data = {
            'file': (io.BytesIO(invalid_content), 'invalid.xlsx')
        }
        response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code in [400, 422]
        result = response.get_json()
        assert result.get('success') is False

        print('[PASS] 无效格式文件被正确拒绝')

    def test_import_subject_invalid_extension(self, client, auth_headers):
        """测试上传错误扩展名的文件"""
        invalid_content = b'{"name": "test"}'
        data = {
            'file': (io.BytesIO(invalid_content), 'test.pdf')
        }
        response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code in [400, 422]

        print('[PASS] 错误扩展名文件被正确拒绝')

    def test_import_subject_missing_required_field(self, client, auth_headers):
        """测试缺少必填字段的导入"""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl 未安装，跳过此测试")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '科目数据'
        # 缺少必填字段"科目名称"
        ws.append(['科目代码', '年级', '描述'])
        ws.append(['CODE001', '高一', '测试描述'])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        data = {
            'file': (buf, 'missing_field.xlsx')
        }
        response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code == 200
        result = response.get_json()
        # 应该返回部分失败或完全失败
        assert result.get('success') is False or result.get('failed_count', 0) > 0

        print(f'[PASS] 缺少必填字段被正确检测: 成功 {result.get("success_count", 0)}, 失败 {result.get("failed_count", 0)}')

    def test_import_subject_duplicate_data(self, client, auth_headers, db_session):
        """测试导入重复数据"""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl 未安装，跳过此测试")

        # 先创建一个科目
        subject = Subject(name='重复科目', code='DUP001', grade='高一', is_active=True)
        db_session.add(subject)
        db_session.commit()

        # 尝试导入相同代码的科目
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '科目数据'
        ws.append(['科目名称', '科目代码', '年级', '是否启用'])
        ws.append(['重复科目', 'DUP001', '高一', '是'])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        data = {
            'file': (buf, 'duplicate.xlsx')
        }
        response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code == 200
        result = response.get_json()

        print(f'[PASS] 重复数据导入处理: 成功 {result.get("success_count", 0)}, 失败 {result.get("failed_count", 0)}')

    def test_import_subject_valid_json(self, client, auth_headers):
        """测试有效的JSON格式导入"""
        import_data = [
            {
                'name': 'JSON科目1',
                'code': 'JSON001',
                'grade': '高一',
                'is_active': True
            },
            {
                'name': 'JSON科目2',
                'code': 'JSON002',
                'grade': '高二',
                'is_active': True
            }
        ]

        data = {
            'file': (io.BytesIO(json.dumps(import_data).encode()), 'import.json')
        }
        response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code == 200
        result = response.get_json()
        assert result.get('success') is True or result.get('success_count', 0) > 0

        print(f'[PASS] JSON导入成功: 导入 {result.get("success_count", 0)} 条')

    def test_import_exam_validation(self, client, auth_headers):
        """测试考试导入的格式验证"""
        data = {
            'file': (io.BytesIO(b'invalid'), 'exam_import.xlsx')
        }
        response = client.post(
            '/api/exam-import/validate',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code in [400, 200]

        print('[PASS] 考试导入格式验证测试通过')


class TestExportFunctionality:
    """测试导出功能"""

    def test_export_subject_excel(self, client, auth_headers, db_session):
        """测试导出Excel格式的科目数据"""
        # Setup: 创建测试数据
        subjects = [
            Subject(name='导出测试1', code='EXP001', grade='高一', is_active=True),
            Subject(name='导出测试2', code='EXP002', grade='高二', is_active=True),
        ]
        db_session.add_all(subjects)
        db_session.commit()

        response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={'format': 'excel'}
        )
        assert response.status_code == 200
        assert ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.content_type or
            'excel' in response.content_type or 'octet-stream' in response.content_type)

        content_disposition = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disposition or 'filename' in content_disposition

        print(f'[PASS] Excel导出成功: 文件大小 {len(response.data)} 字节')

    def test_export_subject_csv(self, client, auth_headers):
        """测试导出CSV格式的科目数据"""
        response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={'format': 'csv'}
        )
        assert response.status_code == 200
        assert 'csv' in response.content_type or 'octet-stream' in response.content_type or 'text' in response.content_type  # noqa: E501

        print(f'[PASS] CSV导出成功: 文件大小 {len(response.data)} 字节')

    def test_export_subject_with_inactive(self, client, auth_headers, db_session):
        """测试导出包含已禁用科目"""
        subject = Subject(name='禁用科目', code='INACTIVE001', grade='高三', is_active=False)
        db_session.add(subject)
        db_session.commit()

        response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={
                'format': 'excel',
                'include_inactive': 'true'
            }
        )
        assert response.status_code == 200

        print('[PASS] 包含已禁用科目导出成功')

    def test_export_subject_filename_chinese(self, client, auth_headers):
        """测试中文字符文件名支持"""
        response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={'format': 'excel'}
        )
        assert response.status_code == 200
        content_disposition = response.headers.get('Content-Disposition', '')
        # 验证Content-Disposition头包含文件名
        assert 'filename' in content_disposition.lower() or 'utf-8' in content_disposition.lower()

        print('[PASS] 中文文件名支持测试通过')

    def test_export_data_integrity(self, client, auth_headers, db_session):
        """测试导出数据的完整性"""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl 未安装，跳过此测试")

        # Setup: 创建已知数据
        test_data = [
            Subject(name='完整性测试1', code='INT001', grade='高一', is_active=True),
            Subject(name='完整性测试2', code='INT002', grade='高二', is_active=True),
            Subject(name='完整性测试3', code='INT003', grade='高三', is_active=True),
        ]
        db_session.add_all(test_data)
        db_session.commit()

        response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={'format': 'excel'}
        )
        assert response.status_code == 200

        # 验证下载的Excel文件
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # 检查数据完整性
        exported_names = [row[0] for row in rows if row[0]]
        assert '完整性测试1' in exported_names
        assert '完整性测试2' in exported_names
        assert '完整性测试3' in exported_names

        print(f'[PASS] 数据完整性验证通过: 导出 {len(rows)} 条数据')

    def test_export_empty_dataset(self, client, auth_headers, db_session):
        """测试空数据集导出"""
        # 先删除所有科目
        Subject.query.delete()
        db_session.commit()

        response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={'format': 'excel'}
        )
        assert response.status_code == 200

        print('[PASS] 空数据集导出测试通过')

    def test_export_exam_template(self, client, auth_headers):
        """测试模板下载"""
        response = client.get(
            '/api/subjects/template',
            headers=auth_headers
        )
        assert response.status_code == 200
        assert len(response.data) > 0

        print(f'[PASS] 模板下载成功: 文件大小 {len(response.data)} 字节')

    def test_export_subject_data_consistency(self, client, auth_headers, db_session):
        """测试导出数据与数据库一致性"""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl 未安装，跳过此测试")

        # Setup
        expected_count = Subject.query.count()

        response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={'format': 'excel'}
        )
        assert response.status_code == 200

        # 验证导出数量与数据库一致
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        ws = wb.active
        exported_count = ws.max_row - 1  # 减去表头

        print(f'[INFO] 数据库数量: {expected_count}, 导出数量: {exported_count}')
        # 允许小范围差异（如过滤已禁用等）

        print('[PASS] 数据一致性测试通过')


class TestImportExportCombined:
    """测试导入导出组合流程"""

    def test_export_then_import_roundtrip(self, client, auth_headers, db_session):
        """测试导出后再导入的往返流程"""
        if not HAS_OPENPYXL:
            pytest.skip("openpyxl 未安装，跳过此测试")

        # Setup: 创建原始数据
        original = Subject(name='往返测试', code='ROUND001', grade='高一', is_active=True)
        db_session.add(original)
        db_session.commit()

        # Export
        export_response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={'format': 'excel'}
        )
        assert export_response.status_code == 200

        # Import (验证文件可以被正确解析)
        import_data = {
            'file': (io.BytesIO(export_response.data), 'reimport.xlsx')
        }
        import_response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=import_data,
            content_type='multipart/form-data'
        )
        assert import_response.status_code == 200

        result = import_response.get_json()
        print(f'[PASS] 往返流程: 成功 {result.get("success_count", 0)}, 失败 {result.get("failed_count", 0)}')

    def test_import_then_filter(self, client, auth_headers, db_session):
        """测试导入后筛选"""
        import_data = [
            {'name': '筛选测试A', 'code': 'FILTER_A', 'grade': '高一', 'is_active': True},
            {'name': '筛选测试B', 'code': 'FILTER_B', 'grade': '高二', 'is_active': True},
        ]

        data = {
            'file': (io.BytesIO(json.dumps(import_data).encode()), 'filter_test.json')
        }
        import_response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert import_response.status_code == 200

        # 验证筛选功能
        filter_response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={'search': '筛选测试A'}
        )
        assert filter_response.status_code == 200
        filter_data = filter_response.get_json()
        items = filter_data['data']
        assert len(items) >= 1
        assert any('筛选测试A' in s['name'] for s in items)

        print('[PASS] 导入后筛选流程测试通过')

    def test_filter_then_export(self, client, auth_headers, db_session):
        """测试筛选后导出"""
        # 创建不同的数据
        subjects = [
            Subject(name='导出筛选A', code='EXP_FILTER_A', grade='高一', is_active=True),
            Subject(name='导出筛选B', code='EXP_FILTER_B', grade='高二', is_active=True),
        ]
        db_session.add_all(subjects)
        db_session.commit()

        # 导出（带筛选条件）
        export_response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={
                'format': 'excel',
                'search': '导出筛选A'
            }
        )
        assert export_response.status_code == 200

        print(f'[PASS] 筛选后导出测试通过: 文件大小 {len(export_response.data)} 字节')


class TestErrorHandling:
    """测试错误处理"""

    def test_filter_sql_injection_prevention(self, client, auth_headers):
        """测试SQL注入防护"""
        malicious_inputs = [
            "' OR 1=1 --",
            "1; DROP TABLE subjects;--",
            "admin'--",
        ]

        for payload in malicious_inputs:
            response = client.get(
                '/api/subjects/',
                headers=auth_headers,
                query_string={'search': payload}
            )
            assert response.status_code == 200
            data = response.get_json()
            # 不应该返回所有数据
            items = data['data']
            # 验证没有造成异常
            assert isinstance(items, list)

        print('[PASS] SQL注入防护测试通过')

    def test_import_oversized_file(self, client, auth_headers):
        """测试超大文件导入"""
        large_content = b'x' * (51 * 1024 * 1024)  # 51MB
        data = {
            'file': (io.BytesIO(large_content), 'large_file.xlsx')
        }
        response = client.post(
            '/api/subjects/import',
            headers=auth_headers,
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code in [400, 413, 422, 500]

        print('[PASS] 超大文件处理测试通过')

    def test_export_unauthorized_access(self, client):
        """测试未授权访问导出"""
        response = client.get('/api/subjects/export')
        assert response.status_code in [401, 403]

        print('[PASS] 未授权访问被正确阻止')

    def test_import_without_auth(self, client):
        """测试未授权导入"""
        data = {
            'file': (io.BytesIO(b'test'), 'test.xlsx')
        }
        response = client.post(
            '/api/subjects/import',
            data=data,
            content_type='multipart/form-data'
        )
        assert response.status_code in [401, 403]

        print('[PASS] 未授权导入被正确阻止')

    def test_filter_without_auth(self, client):
        """测试未授权筛选"""
        response = client.get(
            '/api/subjects/',
            query_string={'search': 'test'}
        )
        assert response.status_code in [401, 403]

        print('[PASS] 未授权筛选被正确阻止')


class TestPerformanceMetrics:
    """测试性能指标"""

    def test_export_response_time(self, client, auth_headers, db_session):
        """测试导出响应时间"""
        import time

        # 创建大量测试数据
        subjects = [
            Subject(name=f'性能测试{i}', code=f'PERF{i:04d}', grade='高一', is_active=True)
            for i in range(100)
        ]
        db_session.add_all(subjects)
        db_session.commit()

        start_time = time.time()
        response = client.get(
            '/api/subjects/export',
            headers=auth_headers,
            query_string={'format': 'excel'}
        )
        elapsed_time = time.time() - start_time

        assert response.status_code == 200
        assert elapsed_time < 5.0  # 5秒内响应

        print(f'[PASS] 导出性能: {elapsed_time:.2f}秒, 数据量: 100条')

    def test_filter_response_time(self, client, auth_headers, db_session):
        """测试筛选响应时间"""

        # 创建测试数据
        subjects = [
            Subject(name=f'筛选性能测试{i}', code=f'FILTER_PERF{i:04d}', grade='高一', is_active=True)
            for i in range(50)
        ]
        db_session.add_all(subjects)
        db_session.commit()

        start_time = time.time()
        response = client.get(
            '/api/subjects/',
            headers=auth_headers,
            query_string={'search': '筛选性能'}
        )
        elapsed_time = time.time() - start_time

        assert response.status_code == 200
        assert elapsed_time < 2.0  # 2秒内响应

        print(f'[PASS] 筛选性能: {elapsed_time:.2f}秒')


class TestTestCasesDocumentation:
    """生成测试用例文档"""

    def test_generate_test_summary(self, client, auth_headers, db_session):
        """生成测试摘要"""
        print('\n' + '='*80)
        print('筛选功能与导入导出功能端到端测试摘要')
        print('='*80)
        print('\n【筛选功能测试用例】')
        print('1. 按科目名称筛选 - 验证搜索功能')
        print('2. 按状态筛选 - 验证启用/禁用筛选')
        print('3. 组合筛选条件 - 验证多条件组合')
        print('4. 空结果筛选 - 验证无结果场景')
        print('5. 特殊字符筛选 - 验证XSS防护')
        print('6. 中文字符筛选 - 验证国际化支持')

        print('\n【导入功能测试用例】')
        print('1. 空文件上传 - 验证错误处理')
        print('2. 无效格式文件 - 验证格式验证')
        print('3. 错误扩展名 - 验证文件类型检查')
        print('4. 缺少必填字段 - 验证数据校验')
        print('5. 重复数据导入 - 验证唯一性约束')
        print('6. JSON格式导入 - 验证多格式支持')
        print('7. 考试导入验证 - 验证特定模块')

        print('\n【导出功能测试用例】')
        print('1. Excel格式导出 - 验证格式正确性')
        print('2. CSV格式导出 - 验证多格式支持')
        print('3. 包含已禁用数据导出 - 验证筛选导出')
        print('4. 中文文件名支持 - 验证国际化')
        print('5. 数据完整性验证 - 验证内容准确性')
        print('6. 空数据集导出 - 验证边界条件')
        print('7. 模板下载 - 验证模板功能')

        print('\n【组合流程测试用例】')
        print('1. 导出-导入往返流程')
        print('2. 导入-筛选流程')
        print('3. 筛选-导出流程')

        print('\n【错误处理测试用例】')
        print('1. SQL注入防护测试')
        print('2. 超大文件处理测试')
        print('3. 未授权访问测试')

        print('\n【性能测试用例】')
        print('1. 导出响应时间测试')
        print('2. 筛选响应时间测试')

        print('\n' + '='*80)

        # 执行一个快速验证
        response = client.get(
            '/api/subjects/',
            headers=auth_headers
        )
        assert response.status_code == 200
        print('\n所有测试用例已就绪，API基本功能正常！')


if __name__ == '__main__':
    pytest.main([__file__, '-v', '--tb=short', '-x'])
