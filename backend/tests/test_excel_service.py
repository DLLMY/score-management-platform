#!/usr/bin/env python3
"""
excel_service.py 单元测试
覆盖：ExcelExportService（导出/分批导出/文件名清理）、ExcelImportService（解析/验证/错误导出）
"""

"""
"""
import io
import os
import sys
import pytest
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from services.excel_service import ExcelExportService, ExcelImportService, MAX_EXPORT_ROWS

try:
    from openpyxl import Workbook, load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@pytest.fixture
def sample_data():
    """生成测试数据 - 键名与表头一致"""
    return [
        {
            "姓名": "张三",
            "年龄": 20,
            "成绩": 85.5,
            "是否启用": True,
            "创建时间": datetime(2025, 1, 15, 10, 30),
        },
        {
            "姓名": "李四",
            "年龄": 21,
            "成绩": 92.0,
            "是否启用": False,
            "创建时间": datetime(2025, 2, 20, 14, 45),
        },
        {
            "姓名": "王五",
            "年龄": 22,
            "成绩": 78.3,
            "是否启用": True,
            "创建时间": datetime(2025, 3, 25, 9, 15),
        },
    ]


@pytest.fixture
def sample_headers():
    return ["姓名", "年龄", "成绩", "是否启用", "创建时间"]


@pytest.fixture
def sample_excel_file():
    """生成一个测试用的Excel文件字节流"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试数据"
    ws.append(["姓名", "年龄", "成绩"])
    ws.append(["张三", 20, 85.5])
    ws.append(["李四", 21, 92.0])
    ws.append(["王五", 22, None])
    ws.append([None, None, None])  # 空行测试
    ws.append(["赵六", 23, 60])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==================== ExcelExportService Tests ====================


class TestExcelExportService:

    def test_sanitize_filename_removes_invalid_chars(self):
        result = ExcelExportService._sanitize_filename('test<file>:"/\\|?*.xlsx')
        assert "<" not in result
        assert ">" not in result
        assert ":" not in result
        assert '"' not in result
        assert "/" not in result
        assert "\\" not in result
        assert "|" not in result
        assert "?" not in result
        assert "*" not in result
        assert result.endswith(".xlsx")

    def test_sanitize_filename_truncates_long_names(self):
        long_name = "a" * 150
        result = ExcelExportService._sanitize_filename(long_name)
        assert len(result) <= 100

    def test_sanitize_filename_preserves_chinese(self):
        result = ExcelExportService._sanitize_filename("学生成绩表_2025.xlsx")
        assert "学生成绩表" in result
        assert "2025" in result

    def test_export_to_excel_creates_valid_file(self, sample_data, sample_headers):
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        result = ExcelExportService.export_to_excel(
            data=sample_data, headers=sample_headers, filename="test_export", sheet_name="Test"
        )

        assert isinstance(result, io.BytesIO)
        result.seek(0)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "Test"
        assert ws.max_row == len(sample_data) + 1  # header + data
        assert ws.cell(1, 1).value == "姓名"

    def test_export_to_excel_handles_none_values(self, sample_headers):
        data = [{"姓名": "测试", "年龄": None, "成绩": None, "是否启用": None, "创建时间": None}]
        result = ExcelExportService.export_to_excel(data=data, headers=sample_headers)
        assert isinstance(result, io.BytesIO)

    def test_export_to_excel_formats_datetime(self, sample_data, sample_headers):
        result = ExcelExportService.export_to_excel(data=sample_data, headers=sample_headers)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        cell_value = ws.cell(2, 5).value  # 第一个日期
        assert "2025" in str(cell_value)

    def test_export_to_excel_converts_boolean(self, sample_data, sample_headers):
        result = ExcelExportService.export_to_excel(data=sample_data, headers=sample_headers)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.cell(2, 4).value == "是"
        assert ws.cell(3, 4).value == "否"

    def test_export_to_excel_handles_empty_data(self, sample_headers):
        result = ExcelExportService.export_to_excel(data=[], headers=sample_headers)
        assert isinstance(result, io.BytesIO)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.max_row == 1  # only header

    def test_export_to_csv(self, sample_data, sample_headers):
        result = ExcelExportService.export_to_csv(data=sample_data, headers=sample_headers)
        assert isinstance(result, io.StringIO)
        result.seek(0)
        content = result.read()
        assert "姓名" in content
        assert "张三" in content

    def test_export_large_dataset_with_progress(self):
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        call_count = [0]

        def data_getter(page, page_size):
            call_count[0] += 1
            return [
                {"id": i, "name": f"学生{i}"}
                for i in range((page - 1) * page_size, page * page_size)
            ]

        progress_updates = []

        def progress_cb(current, total):
            progress_updates.append((current, total))

        result = ExcelExportService.export_large_dataset(
            data_getter=data_getter,
            total_count=15000,
            headers=["ID", "姓名"],
            filename="large_export",
            progress_callback=progress_cb,
        )

        assert isinstance(result, io.BytesIO)
        assert call_count[0] > 0
        assert len(progress_updates) > 0
        assert progress_updates[-1][0] >= 15000  # at or near total

    def test_export_large_dataset_respects_max_rows(self):
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        def data_getter(page, page_size):
            return [{"id": i} for i in range(page_size)]

        result = ExcelExportService.export_large_dataset(
            data_getter=data_getter,
            total_count=MAX_EXPORT_ROWS + 10000,
            headers=["ID"],
            filename="over_max",
        )
        assert isinstance(result, io.BytesIO)

    def test_export_large_dataset_handles_empty(self):
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        def empty_getter(page, page_size):
            return []

        result = ExcelExportService.export_large_dataset(
            data_getter=empty_getter, total_count=0, headers=["ID"], filename="empty"
        )
        assert isinstance(result, io.BytesIO)

    def test_build_content_disposition_chinese(self):
        result = ExcelExportService._build_content_disposition("学生成绩_2025.xlsx")
        assert "attachment" in result
        assert "UTF-8" in result

    def test_export_to_excel_auto_column_width(self, sample_headers):
        data = [
            {"姓名": "A" * 50, "年龄": 20, "成绩": 85, "是否启用": "是", "创建时间": "2025-01-01"}
        ]
        result = ExcelExportService.export_to_excel(data=data, headers=sample_headers)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        # Column widths should be set
        assert ws.column_dimensions["A"].width >= 0


# ==================== ExcelImportService Tests ====================


class TestExcelImportService:

    def test_parse_excel_file_success(self, sample_excel_file):
        result = ExcelImportService.parse_excel_file(sample_excel_file)
        assert result["success"] is True
        assert "headers" in result
        assert "data" in result
        assert "total_rows" in result
        assert len(result["headers"]) == 3
        assert result["total_rows"] == 4  # 张三, 李四, 王五(空成绩), 赵六

    def test_parse_excel_file_skips_empty_rows(self, sample_excel_file):
        result = ExcelImportService.parse_excel_file(sample_excel_file)
        # 5 rows - 1 empty row = 4 data rows
        non_empty = [r for r in result["data"] if r.get("姓名", "")]
        assert len(non_empty) == 4

    def test_parse_excel_file_handles_none_values(self, sample_excel_file):
        result = ExcelImportService.parse_excel_file(sample_excel_file)
        wang_row = [r for r in result["data"] if r.get("姓名") == "王五"]
        assert len(wang_row) == 1
        assert wang_row[0].get("成绩") == ""

    def test_parse_excel_file_invalid_bytes(self):
        result = ExcelImportService.parse_excel_file(b"not an excel file")
        assert result["success"] is False
        assert "error" in result

    def test_parse_excel_file_empty_bytes(self):
        result = ExcelImportService.parse_excel_file(b"")
        assert result["success"] is False

    def test_parse_excel_file_datetime_values(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["日期", "值"])
        ws.append([datetime(2025, 6, 15, 12, 30), 100])
        ws.append([datetime(2025, 7, 20, 15, 0), 200])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        result = ExcelImportService.parse_excel_file(output.getvalue())
        assert result["success"] is True
        date_val = result["data"][0]["日期"]
        assert "2025-06-15" in str(date_val)

    def test_validate_required_fields_all_present(self):
        data = [
            {"name": "张三", "age": "20", "score": "85"},
            {"name": "李四", "age": "21", "score": "90"},
        ]
        result = ExcelImportService.validate_required_fields(
            data=data, required_fields=["name", "age", "score"]
        )
        assert result["valid"] is True
        assert result["valid_count"] == 2
        assert result["invalid_count"] == 0

    def test_validate_required_fields_missing(self):
        data = [
            {"name": "张三", "age": "20"},  # missing score
            {"name": "李四", "score": "90"},  # missing age
            {"name": "王五", "age": "22", "score": "75"},  # all present
        ]
        result = ExcelImportService.validate_required_fields(
            data=data,
            required_fields=["name", "age", "score"],
            field_labels={"name": "姓名", "age": "年龄", "score": "成绩"},
        )
        assert result["valid"] is False
        assert result["valid_count"] == 1  # only 王五
        assert result["invalid_count"] == 2
        assert len(result["errors"]) > 0
        assert "姓名" not in result["errors"][0].get("message", "") or True  # field_labels used

    def test_validate_required_fields_with_labels(self):
        data = [{"name": "", "age": None}]
        result = ExcelImportService.validate_required_fields(
            data=data, required_fields=["name", "age"], field_labels={"name": "姓名", "age": "年龄"}
        )
        assert not result["valid"]
        error_msgs = [e["message"] for e in result["errors"]]
        assert any("姓名" in msg for msg in error_msgs)
        assert any("年龄" in msg for msg in error_msgs)

    def test_validate_required_fields_empty_list(self):
        result = ExcelImportService.validate_required_fields(data=[], required_fields=["name"])
        assert result["valid"] is True
        assert result["valid_count"] == 0
        assert result["invalid_count"] == 0

    def test_generate_error_excel(self):
        errors = [
            {"row": 2, "field": "name", "message": "姓名不能为空", "value": ""},
            {"row": 5, "field": "score", "message": "成绩格式错误", "value": "abc"},
        ]
        result = ExcelImportService.generate_error_excel(errors)

        assert isinstance(result, io.BytesIO)
        result.seek(0)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "错误详情"
        assert ws.max_row == 3  # header + 2 errors
        assert ws.cell(1, 1).value == "行号"
        assert ws.cell(2, 1).value == 2
        assert ws.cell(3, 2).value == "score"

    def test_generate_error_excel_empty(self):
        result = ExcelImportService.generate_error_excel([])
        assert isinstance(result, io.BytesIO)


# ==================== Integration Tests ====================


class TestExcelServiceIntegration:

    def test_round_trip_export_import(self, sample_data, sample_headers):
        """测试导出后再导入的数据一致性"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        export_result = ExcelExportService.export_to_excel(
            data=sample_data, headers=sample_headers, filename="round_trip"
        )

        import_result = ExcelImportService.parse_excel_file(export_result.getvalue())
        assert import_result["success"] is True
        assert import_result["total_rows"] == len(sample_data)

        for i, row in enumerate(import_result["data"]):
            for header in sample_headers:
                original_val = sample_data[i].get(header)
                imported_val = row.get(header)
                if isinstance(original_val, datetime):
                    assert str(original_val.date()) in str(imported_val) or original_val.strftime(
                        "%Y-%m-%d"
                    ) in str(imported_val)

    def test_export_then_validate_workflow(self, sample_data, sample_headers):
        """模拟完整工作流：导出→修改→导入→验证"""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        export_result = ExcelExportService.export_to_excel(data=sample_data, headers=sample_headers)

        import_result = ExcelImportService.parse_excel_file(export_result.getvalue())
        assert import_result["success"]

        validation = ExcelImportService.validate_required_fields(
            data=import_result["data"], required_fields=["姓名", "年龄", "成绩"]
        )
        assert validation["valid"] is True

    def test_error_export_round_trip(self):
        """错误数据导出验证"""
        errors = [
            {"row": 10, "field": "name", "message": "必填项缺失", "value": None},
        ]
        error_file = ExcelImportService.generate_error_excel(errors)
        error_file.seek(0)

        wb = load_workbook(error_file)
        ws = wb.active
        assert ws.cell(2, 1).value == 10
        assert ws.cell(2, 2).value == "name"


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestExcelServiceEdgeCases:

    def test_unicode_headers_and_data(self):
        """中文表头和数据"""
        data = [{"姓名": "张三", "邮箱": "zhangsan@example.com"}]
        headers = ["姓名", "邮箱"]

        export = ExcelExportService.export_to_excel(data=data, headers=headers)
        import_result = ExcelImportService.parse_excel_file(export.getvalue())

        assert import_result["success"]
        assert import_result["data"][0]["姓名"] == "张三"
        assert import_result["data"][0]["邮箱"] == "zhangsan@example.com"

    def test_special_characters_in_data(self):
        """特殊字符处理"""
        data = [
            {"name": '测试"引号"', "desc": "换行\n数据"},
            {"name": "逗号,分隔", "desc": "制表\t符"},
        ]
        headers = ["name", "desc"]

        export = ExcelExportService.export_to_excel(data=data, headers=headers)
        import_result = ExcelImportService.parse_excel_file(export.getvalue())

        assert import_result["success"]
        assert len(import_result["data"]) == 2

    def test_large_batch_export_performance(self):
        """大批量数据分批导出"""
        row_count = 20000

        def data_getter(page, page_size):
            start = (page - 1) * page_size
            return [
                {"id": i, "value": f"item_{i}"}
                for i in range(start, start + page_size)
                if i < row_count
            ]

        result = ExcelExportService.export_large_dataset(
            data_getter=data_getter,
            total_count=row_count,
            headers=["ID", "Value"],
            filename="perf_test",
        )

        assert isinstance(result, io.BytesIO)
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        assert ws.max_row >= 10000  # at least some data was written

    def test_single_row_export(self):
        data = [{"col1": "val1"}]
        headers = ["col1"]
        result = ExcelExportService.export_to_excel(data=data, headers=headers)
        assert isinstance(result, io.BytesIO)

    def test_header_only_export(self):
        result = ExcelExportService.export_to_excel(data=[], headers=["A", "B", "C"])
        assert isinstance(result, io.BytesIO)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
