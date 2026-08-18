"""班级学期报告算法摘要测试

覆盖 services/report_summary_service.py 与 /api/reports/class-semester 摘要接入：
- build_class_summary 结构完整（参与度/风险/归因三维）
- 单维异常隔离（mock 参与度失败不影响风险/归因）
- 空班级不炸
- summary_to_rows 转行格式
- Excel 导出含「算法摘要」sheet；CSV 导出含 # 摘要行

注意：摘要的 batch 接口按 User.class_name 过滤，seed 必须同时设置 class_name。
"""

import io
import uuid
from datetime import datetime, timedelta, date
from unittest import mock

import pytest
from openpyxl import load_workbook

from models import (
    db,
    ClassInfo,
    Exam,
    Score,
    Subject,
    User,
    Attendance,
    HomeworkAssignment,
    HomeworkSubmission,
    ScoreRecord,
)
from services.report_summary_service import build_class_summary, summary_to_rows


@pytest.fixture
def klass(app_context):
    name = "SumClass_" + uuid.uuid4().hex[:6]
    c = ClassInfo(name=name)
    db.session.add(c)
    db.session.commit()
    c.name = name  # 确保 class_name 字段同步
    db.session.commit()
    return c


@pytest.fixture
def students(app_context, klass):
    out = []
    for i in range(2):
        u = User(
            name=f"sum_stu_{i}",
            card_id="SUM" + uuid.uuid4().hex[:10],
            is_active=True,
            role="student",  # risk/score_predict batch 按 role=="student" 过滤
            class_info_id=klass.id,
            class_name=klass.name,  # 摘要 batch 按 class_name 过滤，必须设置
            current_score=80 + i,
        )
        db.session.add(u)
        out.append(u)
    db.session.commit()
    return out


def _seed_engagement(app, sid, cid, base=None):
    """参与度种子：出勤 + 作业 + 积分，使 participation 有数据。"""
    base = base or date.today()
    with app.app_context():
        for k in range(8):
            db.session.add(
                Attendance(
                    student_id=sid,
                    class_id=cid,
                    date=base - timedelta(days=k),
                    status="present",
                )
            )
        ha = HomeworkAssignment(
            id=5000 + sid,
            class_id=cid,
            title="hw",
            assigned_date=base - timedelta(days=4),
            due_date=base,
        )
        db.session.add(ha)
        db.session.add(
            HomeworkSubmission(
                assignment_id=5000 + sid,
                student_id=sid,
                is_submitted=True,
                is_late=False,
            )
        )
        for k in range(4):
            db.session.add(
                ScoreRecord(
                    user_id=sid,
                    score_change=2,
                    created_at=datetime.combine(base, datetime.min.time()) - timedelta(days=k),
                )
            )
        db.session.commit()


class TestBuildClassSummary:
    def test_summary_structure_with_data(self, app, klass, students):
        for s in students:
            _seed_engagement(app, s.id, klass.id)
        with app.app_context():
            summary = build_class_summary(klass.name, 30)
        # 参与度
        assert summary["participation"] is not None
        assert summary["participation"]["total"] == 2
        assert summary["participation"]["valid_students"] == 2
        assert summary["participation"]["avg_score"] > 0
        dist = summary["participation"]["level_distribution"]
        assert sum(dist.values()) == 2
        # 风险（结构存在）
        assert summary["risk"] is not None
        assert summary["risk"]["total"] == 2
        assert isinstance(summary["risk"]["risk_students"], list)
        # 归因（结构存在）
        assert summary["attribution"] is not None
        assert summary["attribution"]["total"] == 2
        assert isinstance(summary["attribution"]["top_factors"], list)
        # 元信息
        assert summary["class_name"] == klass.name
        assert summary["days"] == 30
        assert summary["generated_at"]

    def test_summary_isolates_dimension_failure(self, app, klass, students):
        for s in students:
            _seed_engagement(app, s.id, klass.id)
        with app.app_context():
            with mock.patch(
                "services.report_summary_service.batch_rank",
                side_effect=RuntimeError("boom"),
            ):
                summary = build_class_summary(klass.name, 30)
        # 参与度维度失败 → None，不影响风险/归因
        assert summary["participation"] is None
        assert summary["risk"] is not None
        assert summary["attribution"] is not None

    def test_summary_empty_class_no_crash(self, app):
        with app.app_context():
            summary = build_class_summary("不存在的班", 30)
        assert summary["class_name"] == "不存在的班"
        # 空班级各维度返回结构（可为 None 或空结构，但不抛异常）
        assert "participation" in summary and "risk" in summary and "attribution" in summary

    def test_summary_to_rows_format(self, app, klass, students):
        for s in students:
            _seed_engagement(app, s.id, klass.id)
        with app.app_context():
            summary = build_class_summary(klass.name, 30)
        rows = summary_to_rows(summary)
        assert rows and len(rows) >= 4
        for row in rows:
            assert len(row) == 2  # [项目, 内容]
            assert row[0]
        # 关键维度行存在
        labels = [r[0] for r in rows]
        assert any("参与度" in l for l in labels)
        assert any("风险" in l for l in labels)
        assert any("归因" in l or "波动" in l for l in labels)


class TestSemesterReportWithSummary:
    def test_export_excel_contains_summary_sheet(
        self, client, auth_headers, klass, students, exam_factory
    ):
        exam_factory()
        resp = client.get(f"/api/reports/class-semester?class_id={klass.id}", headers=auth_headers)
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type
        wb = load_workbook(io.BytesIO(resp.data), read_only=True)
        assert "算法摘要" in wb.sheetnames
        ws = wb["算法摘要"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows and len(rows) >= 4
        wb.close()

    def test_export_csv_contains_summary_lines(
        self, client, auth_headers, klass, students, exam_factory
    ):
        exam_factory()
        resp = client.get(
            f"/api/reports/class-semester?class_id={klass.id}&format=csv", headers=auth_headers
        )
        assert resp.status_code == 200
        assert "text/csv" in resp.content_type
        text = resp.data.decode("utf-8-sig")
        assert text.startswith("# 算法摘要")
        assert "参与度" in text and "风险" in text


@pytest.fixture
def exam_factory(app_context, klass, students):
    """创建考试并给两名学生录分（触发主表格渲染）。"""

    def _make():
        e = Exam(name="SumExam_" + uuid.uuid4().hex[:6], status="published", class_id=klass.id)
        db.session.add(e)
        math = Subject(name="数学", code="SX")
        db.session.add(math)
        db.session.commit()
        for i, s in enumerate(students):
            db.session.add(
                Score(exam_id=e.id, student_id=s.id, subject_id=math.id, score=90 - i * 5)
            )
        db.session.commit()
        return e

    return _make
