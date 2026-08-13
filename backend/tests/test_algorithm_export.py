"""算法结果导出 Excel 测试（/api/algorithm/export）

覆盖 GET /api/algorithm/export?tab=engagement|attribution|risk：
- 三 tab 各导出 200 + spreadsheetml + openpyxl 读 sheet 验证表头/行数
- 非法 tab → 400
- 缺鉴权 → 401（student token 不能调）
- 空 class_name（全部班级）不炸
"""
import io
import uuid
from datetime import datetime, timedelta, date

import pytest
from openpyxl import load_workbook

from models import db, ClassInfo, User, Attendance, HomeworkAssignment, HomeworkSubmission, ScoreRecord
from utils.security import generate_student_token


@pytest.fixture
def klass(app_context):
    name = "ExpClass_" + uuid.uuid4().hex[:6]
    c = ClassInfo(name=name)
    db.session.add(c)
    db.session.commit()
    c.name = name
    db.session.commit()
    return c


@pytest.fixture
def students(app_context, klass):
    out = []
    for i in range(2):
        u = User(
            name=f"exp_stu_{i}",
            card_id="EXP" + uuid.uuid4().hex[:10],
            is_active=True,
            role="student",
            class_info_id=klass.id,
            class_name=klass.name,
            current_score=80 + i,
        )
        db.session.add(u)
        out.append(u)
    db.session.commit()
    return out


def _seed_engagement(app, sid, cid, base=None):
    base = base or date.today()
    with app.app_context():
        for k in range(8):
            db.session.add(Attendance(
                student_id=sid, class_id=cid, date=base - timedelta(days=k), status="present",
            ))
        ha = HomeworkAssignment(
            id=6000 + sid, class_id=cid, title="hw",
            assigned_date=base - timedelta(days=4), due_date=base,
        )
        db.session.add(ha)
        db.session.add(HomeworkSubmission(
            assignment_id=6000 + sid, student_id=sid, is_submitted=True, is_late=False,
        ))
        for k in range(4):
            db.session.add(ScoreRecord(
                user_id=sid, score_change=2,
                created_at=datetime.combine(base, datetime.min.time()) - timedelta(days=k),
            ))
        db.session.commit()


class TestAlgorithmExport:
    def test_export_engagement_excel(self, app, client, auth_headers, klass, students):
        for s in students:
            _seed_engagement(app, s.id, klass.id)
        resp = client.get(
            f"/api/algorithm/export?tab=engagement&class_name={klass.name}",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type
        wb = load_workbook(io.BytesIO(resp.data), read_only=True)
        assert "参与度排名" in wb.sheetnames
        ws = wb["参与度排名"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "排名" and rows[0][3] == "参与度"
        assert len(rows) == 3  # 表头 + 2 学生
        assert rows[1][1]  # 学生名非空
        wb.close()

    def test_export_attribution_excel(self, client, auth_headers, klass, students):
        resp = client.get(
            f"/api/algorithm/export?tab=attribution&class_name={klass.name}",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type
        wb = load_workbook(io.BytesIO(resp.data), read_only=True)
        assert "班级归因" in wb.sheetnames
        ws = wb["班级归因"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "姓名"
        assert len(rows) >= 1
        wb.close()

    def test_export_risk_excel(self, client, auth_headers, klass, students):
        resp = client.get(
            f"/api/algorithm/export?tab=risk&class_name={klass.name}",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type
        wb = load_workbook(io.BytesIO(resp.data), read_only=True)
        assert "风险评估" in wb.sheetnames
        ws = wb["风险评估"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "姓名"
        assert len(rows) == 3  # 表头 + 2 学生
        wb.close()

    def test_export_invalid_tab(self, client, auth_headers):
        resp = client.get(
            "/api/algorithm/export?tab=bogus", headers=auth_headers
        )
        assert resp.status_code == 400

    def test_export_requires_admin(self, client, students):
        s = students[0]
        token = generate_student_token(s.id, s.name, s.card_id)
        resp = client.get(
            f"/api/algorithm/export?tab=engagement&class_name={s.class_name}",
            headers={"Authorization": f"Bearer {token}"},
        )
        # student token(type=student) 不能调 algorithm.view 端点
        assert resp.status_code in (401, 403)

    def test_export_without_class_name(self, client, auth_headers, klass, students):
        # 不传 class_name（全部班级）也应正常导出
        resp = client.get("/api/algorithm/export?tab=risk", headers=auth_headers)
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type
